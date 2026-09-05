import sys
import os
import requests
import time
import json
import logging
import random
import math
import functools
import inspect
import traceback
from datetime import datetime, timedelta
from threading import Lock
from collections import defaultdict
from flask import Flask, request, jsonify
from apscheduler.schedulers.background import BackgroundScheduler
from threading import Thread

# ============================================================
# ИМПОРТЫ ИЗ ПРОЕКТА (С app.)
# ============================================================
from app.config import Config
from app.database.storage import storage
from app.telegram.handlers import handlers
from app.utils.logger import setup_logging, get_logger
from app.scheduler import start_scheduler

# ============================================================
# ИНИЦИАЛИЗАЦИЯ
# ============================================================
logger = get_logger(__name__)
app = Flask(__name__)

search_running = False
search_state = {}
TIMEZONE_OFFSET = 3

# ============================================================
# МАРКЕРЫ (ТОЛЬКО ОДИН - 000006)
# ============================================================
MARKERS = {
    42.86875000000006: ('under', 1.95, 'ТМ 2.5'),
}

TOP_LEAGUES = ['Premier League', 'La Liga', 'Bundesliga', 'Serie A', 'Ligue 1']

# ============================================================
# ДОМАШНЕЕ ПРЕИМУЩЕСТВО ПО ЛИГАМ
# ============================================================
HOME_ADVANTAGE = {
    'Premier League': 1.15,
    'La Liga': 1.12,
    'Bundesliga': 1.18,
    'Serie A': 1.10,
    'Ligue 1': 1.13,
    'Championship': 1.12,
    '2. Bundesliga': 1.15,
    'Eredivisie': 1.14,
    'Primeira Liga': 1.11,
    'Süper Lig': 1.16,
}

# ============================================================
# ЗАПАСНЫЕ ЗНАЧЕНИЯ XG
# ============================================================
FALLBACK_XG = {
    'Premier League': {'home': 1.6, 'away': 1.2},
    'La Liga': {'home': 1.5, 'away': 1.2},
    'Bundesliga': {'home': 1.7, 'away': 1.3},
    'Serie A': {'home': 1.5, 'away': 1.1},
    'Ligue 1': {'home': 1.5, 'away': 1.2},
    'Championship': {'home': 1.4, 'away': 1.1},
    'League One': {'home': 1.3, 'away': 1.0},
    'League Two': {'home': 1.2, 'away': 0.9},
    'La Liga 2': {'home': 1.3, 'away': 1.0},
    'Süper Lig': {'home': 1.5, 'away': 1.1},
    'Primeira Liga': {'home': 1.4, 'away': 1.1},
}

# ============================================================
# УЛУЧШЕНИЕ 1: МОНИТОРИНГ ПРОИЗВОДИТЕЛЬНОСТИ
# ============================================================
class PerformanceMonitor:
    def __init__(self):
        self.metrics = defaultdict(lambda: {
            'calls': 0,
            'total_time': 0,
            'max_time': 0,
            'min_time': float('inf'),
            'errors': 0
        })
        self.lock = Lock()
    
    def record(self, func_name, elapsed, error=False):
        with self.lock:
            metric = self.metrics[func_name]
            metric['calls'] += 1
            metric['total_time'] += elapsed
            metric['max_time'] = max(metric['max_time'], elapsed)
            metric['min_time'] = min(metric['min_time'], elapsed)
            if error:
                metric['errors'] += 1
    
    def get_report(self):
        report = []
        for func_name, metric in sorted(self.metrics.items()):
            avg_time = metric['total_time'] / metric['calls'] if metric['calls'] > 0 else 0
            error_rate = metric['errors'] / metric['calls'] * 100 if metric['calls'] > 0 else 0
            report.append({
                'function': func_name,
                'calls': metric['calls'],
                'avg_time': round(avg_time, 3),
                'max_time': round(metric['max_time'], 3),
                'min_time': round(metric['min_time'], 3),
                'error_rate': round(error_rate, 1),
                'total_time': round(metric['total_time'], 3)
            })
        return report
    
    def print_report(self):
        report = self.get_report()
        logger.info("=" * 60)
        logger.info("📊 ОТЧЕТ О ПРОИЗВОДИТЕЛЬНОСТИ")
        logger.info("=" * 60)
        for item in report:
            status = "✅" if item['error_rate'] < 5 else "⚠️" if item['error_rate'] < 20 else "❌"
            logger.info(
                f"{status} {item['function']}: "
                f"{item['calls']} вызовов, "
                f"среднее {item['avg_time']}с, "
                f"макс {item['max_time']}с, "
                f"ошибки {item['error_rate']}%"
            )

perf_monitor = PerformanceMonitor()

def timing_decorator(name=None):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            func_name = name or func.__name__
            start = time.time()
            error = False
            try:
                result = func(*args, **kwargs)
                return result
            except Exception as e:
                error = True
                raise
            finally:
                elapsed = time.time() - start
                perf_monitor.record(func_name, elapsed, error)
                if elapsed > 2.0:
                    logger.warning(f"⏱️ {func_name} выполняется {elapsed:.2f}с (медленно)")
                elif elapsed > 5.0:
                    logger.error(f"🐌 {func_name} ОЧЕНЬ МЕДЛЕННО: {elapsed:.2f}с")
        return wrapper
    return decorator

# ============================================================
# УЛУЧШЕНИЕ 2: УМНОЕ КЭШИРОВАНИЕ С TTL
# ============================================================
class SmartCache:
    def __init__(self, max_size=500):
        self.cache = {}
        self.cache_timestamps = {}
        self.hit_count = {}
        self.last_access = {}
        self.max_size = max_size
        self.default_ttl = 3600
        self.ttl_by_type = {
            'form': 300,
            'odds': 60,
            'statistics': 300,
            'standings': 1800,
            'matches': 3600,
            'h2h': 7200
        }
        
# ============================================================
# УЛУЧШЕНИЕ 5: КЭШИРОВАНИЕ ДАННЫХ (Пункт 5)
# ============================================================
class CacheManager:
    def __init__(self):
        self.standings_cache = {}
        self.form_cache = {}
        self.odds_cache = {}
        self.team_id_cache = {}
        self.timestamps = {}

    def get(self, cache_type, key, ttl_hours=6):
        now = time.time()
        if key in self.cache_store(cache_type) and self.timestamps.get(key, 0) + ttl_hours * 3600 > now:
            return self.cache_store(cache_type)[key]
        return None

    def set(self, cache_type, key, value):
        self.cache_store(cache_type)[key] = value
        self.timestamps[key] = time.time()

    def cache_store(self, cache_type):
        if cache_type == 'standings':
            return self.standings_cache
        elif cache_type == 'form':
            return self.form_cache
        elif cache_type == 'odds':
            return self.odds_cache
        elif cache_type == 'team_id':
            return self.team_id_cache
        return {}

cache_manager = CacheManager()
      class SmartCache:
    def __init__(self, max_size=500):
        self.cache = {}
        self.cache_timestamps = {}
        self.hit_count = {}
        self.last_access = {}
        self.max_size = max_size
        self.default_ttl = 3600
        self.ttl_by_type = {
            'form': 300,
            'odds': 60,
            'statistics': 300,
            'standings': 1800,
            'matches': 3600,
            'h2h': 7200
        }
    
    def get(self, key, data_type='default'):
        if key in self.cache:
            ttl = self.ttl_by_type.get(data_type, self.default_ttl)
            if time.time() - self.cache_timestamps.get(key, 0) < ttl:
                self.hit_count[key] = self.hit_count.get(key, 0) + 1
                self.last_access[key] = time.time()
                return self.cache[key]
            else:
                del self.cache[key]
                del self.cache_timestamps[key]
                if key in self.hit_count:
                    del self.hit_count[key]
                if key in self.last_access:
                    del self.last_access[key]
        return None
    
    def set(self, key, value):
        if len(self.cache) >= self.max_size:
            to_remove = min(
                self.cache.keys(),
                key=lambda k: (self.hit_count.get(k, 0), self.last_access.get(k, 0))
            )
            del self.cache[to_remove]
            if to_remove in self.hit_count:
                del self.hit_count[to_remove]
            if to_remove in self.last_access:
                del self.last_access[to_remove]
            if to_remove in self.cache_timestamps:
                del self.cache_timestamps[to_remove]
        self.cache[key] = value
        self.cache_timestamps[key] = time.time()
        self.hit_count[key] = 0
        self.last_access[key] = time.time()
    
    def clear(self):
        self.cache = {}
        self.cache_timestamps = {}
        self.hit_count = {}
        self.last_access = {}

# ============================================================
# УЛУЧШЕНИЕ 3: ОБРАБОТКА ОШИБОК И РЕТРАИ
# ============================================================
class APIError(Exception):
    pass

class APIErrorRetry(Exception):
    pass

class APIErrorFatal(Exception):
    pass

class APIRateLimiter:
    def __init__(self, max_requests=30, time_window=60):
        self.max_requests = max_requests
        self.time_window = time_window
        self.requests = []
        self.lock = Lock()
    
    def wait_if_needed(self):
        with self.lock:
            now = time.time()
            self.requests = [t for t in self.requests if now - t < self.time_window]
            if len(self.requests) >= self.max_requests:
                sleep_time = self.time_window - (now - self.requests[0])
                if sleep_time > 0:
                    logger.warning(f"⏳ Rate limit: ждем {sleep_time:.1f} сек")
                    time.sleep(sleep_time + 1)
                    return self.wait_if_needed()
            self.requests.append(now)

class RetryManager:
    def __init__(self, max_retries=3, base_delay=1, max_delay=10):
        self.max_retries = max_retries
        self.base_delay = base_delay
        self.max_delay = max_delay
    
    def retry(self, func, *args, **kwargs):
        for attempt in range(self.max_retries + 1):
            try:
                return func(*args, **kwargs)
            except APIErrorRetry as e:
                if attempt == self.max_retries:
                    raise APIErrorFatal(f"Превышено число попыток: {e}")
                delay = min(self.base_delay * (2 ** attempt), self.max_delay)
                jitter = random.uniform(0, delay * 0.1)
                total_delay = delay + jitter
                logger.warning(f"🔄 Попытка {attempt + 1}/{self.max_retries + 1} через {total_delay:.1f}с: {e}")
                time.sleep(total_delay)
            except APIErrorFatal:
                raise
            except Exception as e:
                raise APIErrorFatal(f"Неожиданная ошибка: {e}")

# ============================================================
# КЛАСС FOOTBALL_API С УЛУЧШЕНИЯМИ
# ============================================================
class FootballAPI:
    def __init__(self, api_key=None, base_url=None):
        self.api_key = api_key or Config.FOOTBALL_API_KEY
        self.base_url = base_url or "https://v3.football.api-sports.io"
        self.cache = SmartCache(max_size=500)
        self.last_request_time = 0
        self.min_request_interval = 0.2  # БЫЛО 1.5! Теперь 5 запросов в секунду
        self.rate_limiter = APIRateLimiter(max_requests=250, time_window=60) # БЫЛО 30! Под ваш тариф Pro
        self.retry_manager = RetryManager(max_retries=3, base_delay=1)
        self.error_stats = defaultdict(int)
        logger.info(f"🔑 API ключ загружен: {self.api_key[:8]}..." if self.api_key else "❌ API КЛЮЧ НЕ НАЙДЕН!")
    
    def _make_request(self, endpoint, params=None):
        try:
            self.rate_limiter.wait_if_needed()
            return self.retry_manager.retry(self._make_request_impl, endpoint, params)
        except APIErrorFatal as e:
            logger.error(f"❌ Фатальная ошибка API: {e}")
            self.error_stats['fatal'] += 1
            send_error_to_telegram(f"Фатальная ошибка API: {e}\nEndpoint: {endpoint}")
            if self.error_stats['fatal'] > 10:
                send_telegram("🚨 <b>КРИТИЧЕСКАЯ ОШИБКА</b>\nСлишком много ошибок API. Проверьте ключ!")
            return None
        except Exception as e:
            logger.error(f"❌ Неизвестная ошибка: {e}")
            self.error_stats['unknown'] += 1
            return None
    
    def _make_request_impl(self, endpoint, params=None):
        now = time.time()
        if now - self.last_request_time < self.min_request_interval:
            time.sleep(self.min_request_interval - (now - self.last_request_time))
        headers = {
            'x-rapidapi-key': self.api_key,
            'x-rapidapi-host': 'v3.football.api-sports.io'
        }
        url = f"{self.base_url}{endpoint}"
        logger.info(f"📡 Запрос: {endpoint}")
        logger.info(f"📡 Параметры: {params}")
        try:
            response = requests.get(url, headers=headers, params=params, timeout=15)
            self.last_request_time = time.time()
            logger.info(f"📡 Статус ответа: {response.status_code}")
            
            # Логирование остатка лимитов API
            try:
                rem_min = response.headers.get('x-ratelimit-remaining')
                lim_min = response.headers.get('x-ratelimit-limit')
                rem_day = response.headers.get('x-ratelimit-requests-remaining')
                lim_day = response.headers.get('x-ratelimit-requests-limit')
                logger.info(f"📊 Лимиты API: {rem_min}/{lim_min} в минуту | {rem_day}/{lim_day} в день")
            except:
                pass

            if response.status_code == 429:
                retry_after = int(response.headers.get('Retry-After', 60))
                logger.warning(f"⏳ Rate limit 429, ждем {retry_after}с")
                time.sleep(retry_after)
                raise APIErrorRetry("Rate limit превышен")
            if response.status_code == 403:
                logger.error("❌ API ключ недействителен или истек")
                raise APIErrorFatal("Недействительный API ключ")
            if response.status_code != 200:
                logger.error(f"❌ API ошибка {response.status_code}: {response.text[:200]}")
                if response.status_code >= 500:
                    raise APIErrorRetry(f"Серверная ошибка {response.status_code}")
                else:
                    raise APIErrorFatal(f"Ошибка API {response.status_code}")
            data = response.json()
            if data.get('errors'):
                error_msg = data['errors']
                if 'rate limit' in str(error_msg).lower():
                    raise APIErrorRetry("Rate limit в ответе")
                elif 'api key' in str(error_msg).lower():
                    raise APIErrorFatal(f"Ошибка ключа: {error_msg}")
                else:
                    logger.error(f"❌ API ошибка: {error_msg}")
                    return None
            if 'response' in data:
                logger.info(f"📡 Получено записей: {len(data['response'])}")
            return data
        except requests.exceptions.Timeout:
            raise APIErrorRetry("Таймаут запроса")
        except requests.exceptions.ConnectionError:
            raise APIErrorRetry("Ошибка соединения")
        except json.JSONDecodeError as e:
            raise APIErrorFatal(f"Ошибка парсинга JSON: {e}")
        except Exception as e:
            raise APIErrorFatal(f"Неизвестная ошибка: {e}")
    
    @timing_decorator()
    def get_matches(self, league_id, date):
        cache_key = f"matches_{league_id}_{date}"
        cached = self.cache.get(cache_key, data_type='matches')
        if cached is not None:
            return cached
        params = {
            'league': league_id,
            'season': datetime.now().year,
            'date': date
        }
        data = self._make_request('/fixtures', params)
        if data and 'response' in data:
            matches = data['response']
            self.cache.set(cache_key, matches)
            return matches
        return []
    
    @timing_decorator()
    def get_form(self, team_id):
        cache_key = f"form_{team_id}"
        cached = self.cache.get(cache_key, data_type='form')
        if cached is not None:
            return cached
        try:
            params = {
                'team': team_id,
                'last': 5,
                'status': 'FT'
            }
            data = self._make_request('/fixtures', params)
            if data and 'response' in data:
                matches = data['response']
                if matches:
                    goals_scored = []
                    goals_conceded = []
                    wins = 0
                    draws = 0
                    losses = 0
                    for match in matches:
                        goals = match.get('goals', {})
                        teams = match.get('teams', {})
                        if teams.get('home', {}).get('id') == team_id:
                            scored = goals.get('home', 0) or 0
                            conceded = goals.get('away', 0) or 0
                        else:
                            scored = goals.get('away', 0) or 0
                            conceded = goals.get('home', 0) or 0
                        goals_scored.append(scored)
                        goals_conceded.append(conceded)
                        if scored > conceded:
                            wins += 1
                        elif scored == conceded:
                            draws += 1
                        else:
                            losses += 1
                    if goals_scored:
                        result = {
                            'goals_avg': round(sum(goals_scored) / len(goals_scored), 2),
                            'conceded_avg': round(sum(goals_conceded) / len(goals_conceded), 2),
                            'wins': wins,
                            'draws': draws,
                            'losses': losses,
                            'matches': len(matches),
                            'form': self._calculate_form(matches, team_id)
                        }
                        self.cache.set(cache_key, result)
                        return result
        except Exception as e:
            logger.error(f"Ошибка получения формы команды {team_id}: {e}")
        return None
    
    def _calculate_form(self, matches, team_id):
        form = []
        for match in matches:
            teams = match.get('teams', {})
            goals = match.get('goals', {})
            home_score = goals.get('home', 0) or 0
            away_score = goals.get('away', 0) or 0
            if teams.get('home', {}).get('id') == team_id:
                if home_score > away_score:
                    form.append('W')
                elif home_score == away_score:
                    form.append('D')
                else:
                    form.append('L')
            else:
                if away_score > home_score:
                    form.append('W')
                elif away_score == home_score:
                    form.append('D')
                else:
                    form.append('L')
        return ''.join(form)
    
    @timing_decorator()
    def get_match_statistics(self, fixture_id):
        cache_key = f"stats_{fixture_id}"
        cached = self.cache.get(cache_key, data_type='statistics')
        if cached is not None:
            return cached
        try:
            params = {'fixture': fixture_id}
            data = self._make_request('/fixtures/statistics', params)
            if data and 'response' in data:
                statistics = {}
                for team_stats in data['response']:
                    team_name = team_stats.get('team', {}).get('name', 'Unknown')
                    stats = {}
                    for stat in team_stats.get('statistics', []):
                        key = stat.get('type', '')
                        value = stat.get('value', 0)
                        if value is None:
                            value = 0
                        elif isinstance(value, str):
                            if '%' in value:
                                try:
                                    value = float(value.replace('%', ''))
                                except:
                                    value = 0
                            else:
                                try:
                                    value = float(value)
                                except:
                                    value = 0
                        elif isinstance(value, (int, float)):
                            value = float(value)
                        else:
                            value = 0
                        stats[key] = value
                    statistics[team_name] = stats
                self.cache.set(cache_key, statistics)
                return statistics
            else:
                logger.warning(f"⚠️ API вернул пустой ответ для /fixtures/statistics")
                return None
        except Exception as e:
            logger.error(f"❌ Ошибка получения статистики матча {fixture_id}: {e}")
        return None
    
    @timing_decorator()
    def get_standings(self, league_id):
        cache_key = f"standings_{league_id}"
        cached = self.cache.get(cache_key, data_type='standings')
        if cached is not None:
            return cached
        try:
            params = {
                'league': league_id,
                'season': datetime.now().year
            }
            data = self._make_request('/standings', params)
            if data and 'response' in data:
                standings = {}
                for league in data['response']:
                    for standing in league.get('league', {}).get('standings', []):
                        for team in standing:
                            team_name = team.get('team', {}).get('name', '')
                            standings[team_name] = {
                                'position': team.get('rank', 0),
                                'points': team.get('points', 0),
                                'form': team.get('form', ''),
                                'goals_diff': team.get('goalsDiff', 0)
                            }
                self.cache.set(cache_key, standings)
                return standings
        except Exception as e:
            logger.error(f"Ошибка получения таблицы {league_id}: {e}")
        return None
    
    def get_injuries(self, team_id):
        cache_key = f"injuries_{team_id}"
        cached = self.cache.get(cache_key, data_type='default')
        if cached is not None:
            return cached
        try:
            params = {
                'team': team_id,
                'season': datetime.now().year
            }
            data = self._make_request('/injuries', params)
            if data and 'response' in data:
                injuries = data['response']
                self.cache.set(cache_key, injuries)
                return injuries
        except Exception as e:
            logger.error(f"Ошибка получения травм команды {team_id}: {e}")
        return []
    
    @timing_decorator()
    def get_match_result(self, fixture_id):
        cache_key = f"result_{fixture_id}"
        cached = self.cache.get(cache_key, data_type='default')
        if cached is not None:
            return cached
        try:
            params = {'id': fixture_id}
            data = self._make_request('/fixtures', params)
            if data and 'response' in data:
                fixtures = data['response']
                if fixtures:
                    fixture = fixtures[0]
                    goals = fixture.get('goals', {})
                    result = {
                        'goals': {
                            'home': goals.get('home'),
                            'away': goals.get('away')
                        },
                        'status': fixture.get('status', {}).get('short', 'FT')
                    }
                    self.cache.set(cache_key, result)
                    return result
        except Exception as e:
            logger.error(f"Ошибка получения результата {fixture_id}: {e}")
        return None
    
    @timing_decorator()
    def get_head_to_head(self, home_team, away_team):
        cache_key = f"h2h_{home_team}_{away_team}"
        cached = self.cache.get(cache_key, data_type='h2h')
        if cached is not None:
            return cached
        try:
            home_id = self.get_team_id(home_team)
            away_id = self.get_team_id(away_team)
            if home_id and away_id:
                params = {
                    'h2h': f"{home_id}-{away_id}",
                    'last': 5
                }
                data = self._make_request('/fixtures/headtohead', params)
                if data and 'response' in data:
                    fixtures = data['response']
                    if fixtures:
                        result = {
                            'matches': [],
                            'home_wins': 0,
                            'away_wins': 0,
                            'draws': 0,
                            'goals_scored': 0,
                            'goals_conceded': 0
                        }
                        for fixture in fixtures:
                            teams = fixture.get('teams', {})
                            goals = fixture.get('goals', {})
                            home_score = goals.get('home', 0) or 0
                            away_score = goals.get('away', 0) or 0
                            result['matches'].append({
                                'home': teams.get('home', {}).get('name', ''),
                                'away': teams.get('away', {}).get('name', ''),
                                'home_score': home_score,
                                'away_score': away_score
                            })
                            if home_score > away_score:
                                result['home_wins'] += 1
                            elif home_score < away_score:
                                result['away_wins'] += 1
                            else:
                                result['draws'] += 1
                            result['goals_scored'] += home_score
                            result['goals_conceded'] += away_score
                        if result['matches']:
                            total_matches = len(result['matches'])
                            result['avg_goals'] = round((result['goals_scored'] + result['goals_conceded']) / total_matches, 2)
                            result['home_win_rate'] = round((result['home_wins'] / total_matches) * 100, 1)
                            result['total_matches'] = total_matches
                            self.cache.set(cache_key, result)
                            return result
                    else:
                        logger.warning(f"⚠️ Нет данных H2H для {home_team} vs {away_team}")
            else:
                logger.warning(f"⚠️ Не найдены ID команд для H2H")
        except Exception as e:
            logger.error(f"❌ Ошибка получения H2H: {e}")
        return None
    
    def get_team_id(self, team_name):
        cache_key = f"team_id_{team_name}"
        cached = self.cache.get(cache_key, data_type='default')
        if cached is not None:
            return cached
        try:
            params = {'name': team_name}
            data = self._make_request('/teams', params)
            if data and 'response' in data:
                for team in data['response']:
                    team_data = team.get('team', {})
                    if team_data.get('name', '').lower() == team_name.lower():
                        team_id = team_data.get('id')
                        self.cache.set(cache_key, team_id)
                        return team_id
        except Exception as e:
            logger.error(f"Ошибка получения ID команды {team_name}: {e}")
        return None
    
    @timing_decorator()
    def get_match_odds(self, fixture_id):
        cache_key = f"odds_{fixture_id}"
        cached = self.cache.get(cache_key, data_type='odds')
        if cached is not None:
            return cached
        try:
            params = {'fixture': fixture_id}
            data = self._make_request('/fixtures/odds', params)
            if data and 'response' in data:
                odds_data = data['response']
                if odds_data:
                    result = self._extract_best_odds(odds_data)
                    if result.get('best_odds', 0) > 0:
                        logger.info(f"✅ Найдены кэфы для матча {fixture_id}")
                    self.cache.set(cache_key, result)
                    return result
                return None
            return None
        except Exception as e:
            logger.error(f"❌ Ошибка получения кэфов {fixture_id}: {e}")
            return None
    
    def _extract_best_odds(self, odds_data):
        result = {
            'best_odds': 0,
            'bookmaker': '—',
            'home_odds': 0,
            'draw_odds': 0,
            'away_odds': 0,
            'under_odds': 0,
            'over_odds': 0
        }
        for bookmaker in odds_data:
            bookmaker_name = bookmaker.get('bookmaker', {}).get('name', '—')
            bets = bookmaker.get('bets', [])
            for bet in bets:
                bet_name = bet.get('name', '').lower()
                values = bet.get('values', [])
                if not values:
                    continue
                if 'матч' in bet_name or 'match' in bet_name or 'побед' in bet_name:
                    for value in values:
                        value_name = value.get('value', '').lower()
                        odd = value.get('odd', 0)
                        if odd <= 0:
                            continue
                        if '1' in value_name or 'home' in value_name:
                            if odd > result['home_odds']:
                                result['home_odds'] = odd
                        elif '2' in value_name or 'away' in value_name:
                            if odd > result['away_odds']:
                                result['away_odds'] = odd
                        elif 'x' in value_name or 'draw' in value_name:
                            if odd > result['draw_odds']:
                                result['draw_odds'] = odd
                        if odd > result['best_odds']:
                            result['best_odds'] = odd
                            result['bookmaker'] = bookmaker_name
                if 'тотал' in bet_name or 'total' in bet_name:
                    is_2_5 = False
                    for value in values:
                        if '2.5' in value.get('value', ''):
                            is_2_5 = True
                            break
                    if not is_2_5:
                        continue
                    for value in values:
                        value_name = value.get('value', '').lower()
                        odd = value.get('odd', 0)
                        if odd <= 0:
                            continue
                        if 'меньше' in value_name or 'under' in value_name:
                            if odd > result['under_odds']:
                                result['under_odds'] = odd
                        elif 'больше' in value_name or 'over' in value_name:
                            if odd > result['over_odds']:
                                result['over_odds'] = odd
                        if odd > result['best_odds']:
                            result['best_odds'] = odd
                            result['bookmaker'] = bookmaker_name
        return result
    
    def clear_cache(self):
        self.cache.clear()
        logger.info("🧹 Кэш очищен")
    
    def find_fixture_by_teams(self, home_team, away_team):
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            params = {
                'date': today,
                'status': 'FT'
            }
            data = self._make_request('/fixtures', params)
            if data and 'response' in data:
                for fixture in data['response']:
                    teams = fixture.get('teams', {})
                    home = teams.get('home', {}).get('name', '')
                    away = teams.get('away', {}).get('name', '')
                    if home_team.lower() in home.lower() and away_team.lower() in away.lower():
                        return fixture.get('fixture', {}).get('id')
        except Exception as e:
            logger.error(f"Ошибка поиска матча {home_team} vs {away_team}: {e}")
        return None

# ============================================================
# СОЗДАЕМ ЭКЗЕМПЛЯР
# ============================================================
football_api = FootballAPI()

# ============================================================
# КЛАСС ODD_API
# ============================================================
class OddsAPIClient:
    def __init__(self, api_key=None):
        from app.config import Config
        self.api_key = api_key or Config.ODDS_API_KEY
        self.base_url = Config.ODDS_API_URL
        self.cache = SmartCache(max_size=200)
        self.last_request_time = 0
        self.min_request_interval = 0.5
        self.rate_limiter = APIRateLimiter(max_requests=50, time_window=60)
        self.retry_manager = RetryManager(max_retries=2, base_delay=0.5)
        logger.info(f"🎯 Odds API ключ загружен: {self.api_key[:8]}..." if self.api_key else "❌ Odds API КЛЮЧ НЕ НАЙДЕН!")
    
    def _make_request(self, endpoint, params=None):
        try:
            self.rate_limiter.wait_if_needed()
            return self.retry_manager.retry(self._make_request_impl, endpoint, params)
        except Exception as e:
            logger.error(f"❌ Ошибка запроса Odds API: {e}")
            return None
    
    def _make_request_impl(self, endpoint, params=None):
        now = time.time()
        if now - self.last_request_time < self.min_request_interval:
            time.sleep(self.min_request_interval - (now - self.last_request_time))
        url = f"{self.base_url}{endpoint}"
        params = params or {}
        params['apiKey'] = self.api_key
        logger.info(f"📡 Запрос Odds API: {url}")
        logger.info(f"📡 Параметры: {params}")
        response = requests.get(url, params=params, timeout=10)
        self.last_request_time = time.time()
        if response.status_code == 200:
            return response.json()
        elif response.status_code == 429:
            retry_after = int(response.headers.get('Retry-After', 60))
            logger.warning(f"⏳ Odds API rate limit, ждем {retry_after}с")
            time.sleep(retry_after)
            raise APIErrorRetry("Odds API rate limit")
        else:
            logger.error(f"❌ Odds API ошибка {response.status_code}: {response.text[:200]}")
            return None
    
    @timing_decorator()
    def get_odds_for_match(self, home_team, away_team, league):
        cache_key = f"odds_{home_team}_{away_team}_{league}"
        cached = self.cache.get(cache_key, data_type='odds')
        if cached is not None:
            return cached
        try:
            sport_map = {
                'Premier League': 'soccer_epl',
                'La Liga': 'soccer_spain_la_liga',
                'Bundesliga': 'soccer_germany_bundesliga',
                'Serie A': 'soccer_italy_serie_a',
                'Ligue 1': 'soccer_france_ligue_one',
                'Лига Чемпионов УЕФА': 'soccer_uefa_champs_league',
                'UEFA Champions League': 'soccer_uefa_champs_league',
                'Лига Европы УЕФА': 'soccer_uefa_europa_league',
                'UEFA Europa League': 'soccer_uefa_europa_league',
                'MLS': 'soccer_usa_mls',
                'Бразилия Серия А': 'soccer_brazil_campeonato',
                'Аргентина Примера': 'soccer_argentina_primera_division',
            }
            sport_key = sport_map.get(league, 'soccer_epl')
            region = 'eu'
            endpoint = f"/sports/{sport_key}/events"
            params = {
                'region': region,
                'markets': 'h2h,totals'
            }
            data = self._make_request(endpoint, params)
            if data:
                for event in data:
                    event_home = event.get('home_team', '').lower()
                    event_away = event.get('away_team', '').lower()
                    home_lower = home_team.lower()
                    away_lower = away_team.lower()
                    if (home_lower in event_home or event_home in home_lower) and \
                       (away_lower in event_away or event_away in away_lower):
                        result = self._extract_odds(event)
                        self.cache.set(cache_key, result)
                        return result
            return None
        except Exception as e:
            logger.error(f"❌ Ошибка получения коэффициентов: {e}")
            return None
    
    def _extract_odds(self, event):
        result = {
            'best_odds': 0,
            'bookmaker_name': '—',
            'home_odds': 0,
            'draw_odds': 0,
            'away_odds': 0,
            'under_odds': 0,
            'over_odds': 0
        }
        for bookmaker in event.get('bookmakers', []):
            bookmaker_key = bookmaker.get('key', '')
            for market in bookmaker.get('markets', []):
                market_key = market.get('key', '')
                for outcome in market.get('outcomes', []):
                    name = outcome.get('name', '')
                    price = outcome.get('price', 0)
                    if market_key == 'h2h':
                        if name == event.get('home_team'):
                            result['home_odds'] = max(result['home_odds'], price)
                        elif name == event.get('away_team'):
                            result['away_odds'] = max(result['away_odds'], price)
                        elif name == 'Draw':
                            result['draw_odds'] = max(result['draw_odds'], price)
                    elif market_key == 'totals' and '2.5' in name:
                        if 'Over' in name:
                            result['over_odds'] = max(result['over_odds'], price)
                        elif 'Under' in name:
                            result['under_odds'] = max(result['under_odds'], price)
                    if price > result['best_odds']:
                        result['best_odds'] = price
                        result['bookmaker_name'] = bookmaker_key
        return result

# ============================================================
# СОЗДАЕМ ЭКЗЕМПЛЯР
# ============================================================
odds_api = OddsAPIClient()

# ============================================================
# КЛАСС AUTOBET
# ============================================================
class AutoBet:
    def __init__(self):
        self.enabled = True
        self.bets_today = 0
        self.max_bets_per_day = 10
        
    def check_and_bet(self, match_data):
        if not self.enabled:
            return None
        bets = match_data.get('bets', [])
        if not bets:
            return None
        best_bet = max(bets, key=lambda x: x.get('ev', 0))
        if best_bet.get('ev', 0) <= 0:
            return None
        if best_bet.get('odds', 0) < 1.5:
            return None
        bank = storage.load_bank()
        stake = best_bet.get('stake', 0)
        max_stake = bank * 0.1
        if stake > max_stake:
            stake = max_stake
            best_bet['stake'] = stake
        self.bets_today += 1
        return {
            'match': f"{match_data.get('home', '')} vs {match_data.get('away', '')}",
            'match_time': match_data.get('match_time', ''),
            'bet': best_bet.get('label', ''),
            'odds': best_bet.get('odds', 0),
            'stake': stake,
            'ev': best_bet.get('ev', 0),
            'marker_stake': best_bet.get('marker_stake', 0),
            'xg_total': match_data.get('total_xg', 0),
            'prob': best_bet.get('prob', 0),
            'home_form': match_data.get('home_form', ''),
            'away_form': match_data.get('away_form', ''),
            'home_position': match_data.get('standings', {}).get('home_position', '?'),
            'away_position': match_data.get('standings', {}).get('away_position', '?'),
            'bookmaker': best_bet.get('bookmaker', '—'),
            'bet_type': best_bet.get('type', 'under'),
            'source': match_data.get('source', '70_percent')
        }

auto_bet = AutoBet()

# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================
def send_error_to_telegram(error_text: str):
    try:
        url = f"https://api.telegram.org/bot{Config.TELEGRAM_TOKEN}/sendMessage"
        if len(error_text) > 4000:
            error_text = error_text[:4000] + "...(обрезано)"
        data = {
            'chat_id': Config.ADMIN_CHAT_ID,
            'text': f"❌ <b>ОШИБКА БОТА</b>\n\n{error_text}",
            'parse_mode': 'HTML'
        }
        requests.post(url, json=data, timeout=5)
    except Exception as e:
        logger.error(f"Не удалось отправить ошибку в Telegram: {e}")

def send_telegram(text: str, parse_mode: str = 'HTML'):
    try:
        url = f"https://api.telegram.org/bot{Config.TELEGRAM_TOKEN}/sendMessage"
        data = {
            'chat_id': Config.ADMIN_CHAT_ID,
            'text': text,
            'parse_mode': parse_mode
        }
        response = requests.post(url, json=data, timeout=10)
        if response.status_code != 200:
            logger.error(f"❌ Ошибка отправки: {response.text}")
    except Exception as e:
        logger.error(f"❌ Send error: {e}")
        send_error_to_telegram(f"Ошибка отправки в Telegram: {e}")

def export_to_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    history = storage.load_history()
    if not history:
        return None, "📭 Нет данных для экспорта"
    wb = Workbook()
    ws = wb.active
    ws.title = "Ставки"
    headers = ["Дата", "Матч", "Счёт", "Ставка", "Коэф", "EV%", "Сумма", "Результат", "Прибыль", "Букмекер"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    total_profit = 0
    for bet in history:
        date = bet.get('date', '')
        home = bet.get('home', '')
        away = bet.get('away', '')
        home_goals = bet.get('home_goals', '')
        away_goals = bet.get('away_goals', '')
        score = f"{home_goals}-{away_goals}" if home_goals is not None and away_goals is not None else "-"
        bet_type = bet.get('bet', '')
        odds = bet.get('odds', 0)
        ev = bet.get('ev', 0)
        stake = bet.get('stake', 0)
        result = bet.get('result', 'pending')
        profit = bet.get('profit', 0)
        bookmaker = bet.get('bookmaker', '—')
        if result == 'win':
            profit = round(stake * (odds - 1), 2) if profit == 0 else profit
            total_profit += profit
        elif result == 'loss':
            profit = -round(stake, 2) if profit == 0 else profit
            total_profit += profit
        else:
            profit = 0
        ws.append([date, f"{home} vs {away}", score, bet_type, odds, ev, stake, result, profit, bookmaker])
    ws.append([])
    ws.append(["ИТОГО", "", "", "", "", "", "", "", round(total_profit, 2), ""])
    for col in range(1, len(headers) + 1):
        column_letter = chr(64 + col)
        ws.column_dimensions[column_letter].width = 15
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, f"✅ Экспорт завершен! Всего ставок: {len(history)}, Прибыль: ${round(total_profit, 2)}"

def get_profit_data(history):
    profits = []
    days = 7
    for i in range(days - 1, -1, -1):
        day_profit = 0
        day = datetime.now() - timedelta(days=i)
        for bet in history:
            try:
                bet_date = datetime.strptime(bet.get('date', '').split()[0], '%Y-%m-%d')
                if bet_date.date() == day.date():
                    stake = bet.get('stake', 0)
                    if isinstance(stake, str):
                        try:
                            stake = float(stake)
                        except:
                            stake = 0
                    odds = bet.get('odds', 1)
                    if isinstance(odds, str):
                        try:
                            odds = float(odds)
                        except:
                            odds = 1
                    if bet.get('result') == 'win':
                        day_profit += stake * (odds - 1)
                    elif bet.get('result') == 'loss':
                        day_profit -= stake
            except:
                pass
        profits.append(round(day_profit, 2))
    dates = [(datetime.now() - timedelta(days=i)).strftime('%d.%m') for i in range(days - 1, -1, -1)]
    return {'dates': dates, 'profits': profits}

# ============================================================
# ОСНОВНЫЕ ФУНКЦИИ АНАЛИЗА
# ============================================================
def get_motivation(position):
    if position <= 4:
        return 'champions_league'
    elif position <= 6:
        return 'europa_league'
    elif position <= 17:
        return 'mid_table'
    else:
        return 'relegation'

def analyze_form(form_string):
    if not form_string:
        return 'average'
    wins = form_string.count('W')
    if wins >= 4:
        return 'excellent'
    elif wins >= 3:
        return 'good'
    elif wins >= 2:
        return 'average'
    else:
        return 'poor'

def is_quality_match(home_xg, away_xg, home_position, away_position, home_form, away_form):
    try:
        total_xg = home_xg + away_xg
        conditions = [
            total_xg > 1.5,
            total_xg < 3.5,
            home_position < 20,
            away_position < 20,
            len(home_form) >= 3,
            len(away_form) >= 3,
        ]
        return all(conditions)
    except:
        return False

def calculate_poisson_probability(home_xg, away_xg):
    def poisson_prob(avg, goals):
        return (math.exp(-avg) * avg ** goals) / math.factorial(goals)
    home_goals_prob = [poisson_prob(home_xg, i) for i in range(6)]
    away_goals_prob = [poisson_prob(away_xg, i) for i in range(6)]
    prob_home_win = 0
    prob_away_win = 0
    prob_draw = 0
    prob_1X = 0
    prob_X2 = 0
    prob_over_2_5 = 0
    prob_under_2_5 = 0
    prob_btts = 0
    for h_g in range(6):
        for a_g in range(6):
            p = home_goals_prob[h_g] * away_goals_prob[a_g]
            total_goals = h_g + a_g
            if h_g > a_g:
                prob_home_win += p
            elif h_g < a_g:
                prob_away_win += p
            else:
                prob_draw += p
            if h_g >= a_g:
                prob_1X += p
            if a_g >= h_g:
                prob_X2 += p
            if total_goals > 2.5:
                prob_over_2_5 += p
            else:
                prob_under_2_5 += p
            if h_g > 0 and a_g > 0:
                prob_btts += p
    return {
        'home_win': prob_home_win,
        'away_win': prob_away_win,
        'draw': prob_draw,
        '1X': prob_1X,
        'X2': prob_X2,
        'over_2_5': prob_over_2_5,
        'under_2_5': prob_under_2_5,
        'btts': prob_btts
    }

def calculate_form_probability(home_form, away_form):
    home_form_quality = analyze_form(home_form)
    away_form_quality = analyze_form(away_form)
    prob = {
        'home_win': 0.35,
        'away_win': 0.30,
        'draw': 0.35,
        '1X': 0.70,
        'X2': 0.65,
        'over_2_5': 0.45,
        'under_2_5': 0.55,
        'btts': 0.45
    }
    if home_form_quality == 'excellent' and away_form_quality == 'poor':
        prob['home_win'] += 0.15
        prob['1X'] += 0.10
        prob['away_win'] -= 0.10
        prob['X2'] -= 0.10
    elif home_form_quality == 'poor' and away_form_quality == 'excellent':
        prob['away_win'] += 0.15
        prob['X2'] += 0.10
        prob['home_win'] -= 0.10
        prob['1X'] -= 0.10
    return prob

def calculate_h2h_probability(h2h_data):
    prob = {
        'home_win': 0.33,
        'away_win': 0.33,
        'draw': 0.34,
        '1X': 0.67,
        'X2': 0.67,
        'over_2_5': 0.50,
        'under_2_5': 0.50,
        'btts': 0.50
    }
    if h2h_data:
        total = h2h_data.get('total_matches', 0)
        if total > 0:
            home_wins = h2h_data.get('home_wins', 0) / total
            away_wins = h2h_data.get('away_wins', 0) / total
            draws = h2h_data.get('draws', 0) / total
            prob['home_win'] = home_wins * 0.5 + 0.25
            prob['away_win'] = away_wins * 0.5 + 0.25
            prob['draw'] = draws * 0.5 + 0.25
            prob['1X'] = prob['home_win'] + prob['draw']
            prob['X2'] = prob['away_win'] + prob['draw']
            avg_goals = h2h_data.get('avg_goals', 2.5)
            if avg_goals > 2.5:
                prob['over_2_5'] = 0.55
                prob['under_2_5'] = 0.45
            else:
                prob['over_2_5'] = 0.45
                prob['under_2_5'] = 0.55
    return prob

def ensemble_probability(home_xg, away_xg, home_form, away_form, h2h_data):
    poisson = calculate_poisson_probability(home_xg, away_xg)
    form_prob = calculate_form_probability(home_form, away_form)
    h2h_prob = calculate_h2h_probability(h2h_data)
    final_prob = {
        'home_win': poisson['home_win'] * 0.5 + form_prob['home_win'] * 0.3 + h2h_prob['home_win'] * 0.2,
        'away_win': poisson['away_win'] * 0.5 + form_prob['away_win'] * 0.3 + h2h_prob['away_win'] * 0.2,
        'draw': poisson['draw'] * 0.5 + form_prob['draw'] * 0.3 + h2h_prob['draw'] * 0.2,
        '1X': poisson['1X'] * 0.5 + form_prob['1X'] * 0.3 + h2h_prob['1X'] * 0.2,
        'X2': poisson['X2'] * 0.5 + form_prob['X2'] * 0.3 + h2h_prob['X2'] * 0.2,
        'over_2_5': poisson['over_2_5'] * 0.5 + form_prob['over_2_5'] * 0.3 + h2h_prob['over_2_5'] * 0.2,
        'under_2_5': poisson['under_2_5'] * 0.5 + form_prob['under_2_5'] * 0.3 + h2h_prob['under_2_5'] * 0.2,
        'btts': poisson['btts'] * 0.5 + form_prob['btts'] * 0.3 + h2h_prob['btts'] * 0.2
    }
    total_win_prob = final_prob['home_win'] + final_prob['draw'] + final_prob['away_win']
    if total_win_prob > 0:
        final_prob['home_win'] /= total_win_prob
        final_prob['away_win'] /= total_win_prob
        final_prob['draw'] /= total_win_prob
        final_prob['1X'] = final_prob['home_win'] + final_prob['draw']
        final_prob['X2'] = final_prob['away_win'] + final_prob['draw']
    return final_prob

def update_manual_result(match_name, score):
    try:
        home_goals = None
        away_goals = None
        if score and '-' in score:
            parts = score.split('-')
            try:
                home_goals = int(parts[0].strip())
                away_goals = int(parts[1].strip())
            except:
                return "❌ Неверный формат счета. Используй: 2-1"
        history = storage.load_history()
        found = False
        result = 'pending'
        for bet in history:
            if bet.get('result') == 'pending' or bet.get('result') is None:
                home = bet.get('home', '')
                away = bet.get('away', '')
                full_match = f"{home} vs {away}"
                if match_name.lower() in full_match.lower() or full_match.lower() in match_name.lower():
                    bet['home_goals'] = home_goals
                    bet['away_goals'] = away_goals
                    bet_type = bet.get('bet', '')
                    result = determine_bet_result(bet_type, home_goals, away_goals)
                    bet['result'] = result
                    if result == 'win':
                        bet['profit'] = round(bet['stake'] * (bet['odds'] - 1), 2)
                    elif result == 'loss':
                        bet['profit'] = -bet['stake']
                    else:
                        bet['profit'] = 0
                    found = True
                    break
        if found:
            storage.save_history(history)
            recalc_stats()
            msg = f"✅ Результат обновлен!\n"
            msg += f"🏟️ {match_name}\n"
            msg += f"⚽ Счет: {home_goals}-{away_goals}\n"
            msg += f"📊 Результат: {result}"
            return msg
        else:
            return f"❌ Матч '{match_name}' не найден в истории или уже завершен"
    except Exception as e:
        logger.error(f"Ошибка ручного обновления: {e}")
        return f"❌ Ошибка: {e}"

def analyze_match(match_name):
    try:
        cache = storage.load_cache()
        matches = cache.get('top_matches', [])
        for match in matches:
            home = match.get('home', '')
            away = match.get('away', '')
            full_match = f"{home} vs {away}"
            if match_name.lower() in full_match.lower() or full_match.lower() in match_name.lower():
                result = f"📊 <b>АНАЛИЗ МАТЧА</b>\n"
                result += f"🏟️ {full_match}\n"
                result += f"🏆 Лига: {match.get('league', 'Unknown')}\n"
                result += f"📅 Дата: {match.get('match_time', 'Unknown')}\n\n"
                best = match.get('best_bet', {})
                result += f"🎯 <b>ЛУЧШАЯ СТАВКА: {best.get('label', '—')}</b>\n"
                result += f"📈 EV: <b>{best.get('ev', 0)}%</b> | Вероятность: {best.get('prob', 0)}%\n"
                result += f"💰 Коэффициент: {best.get('odds', 0)}\n"
                if best.get('bookmaker'):
                    result += f"🏷️ Лучший коэффициент: {best.get('bookmaker')}\n"
                result += "\n"
                result += "📊 <b>ВСЕ СТАВКИ:</b>\n"
                bets = match.get('bets', [])
                for i, bet in enumerate(bets[:7], 1):
                    ev = bet.get('ev', 0)
                    emoji = "🟢" if ev > 10 else ("🟡" if ev > 5 else "🔴")
                    result += f"{emoji} {i}. {bet.get('label', '—')} | EV: {bet.get('ev', 0)}% | Prob: {bet.get('prob', 0)}% | КЭФ: {bet.get('odds', 0)}\n"
                result += f"\n⚽ XG: {match.get('total_xg', 0):.2f}"
                result += f" | Хозяева: {match.get('home_xg', 0):.2f}"
                result += f" | Гости: {match.get('away_xg', 0):.2f}\n"
                result += f"📈 Форма: {match.get('home_form', '—')} vs {match.get('away_form', '—')}\n"
                result += f"🏆 Позиция: #{match.get('standings', {}).get('home_position', '?')} vs #{match.get('standings', {}).get('away_position', '?')}\n\n"
                if best.get('ev', 0) > 10:
                    result += f"💡 <b>Рекомендация: {best.get('label', '—')}</b> (EV: {best.get('ev', 0)}%) ✅"
                elif best.get('ev', 0) > 5:
                    result += f"💡 <b>Рекомендация: {best.get('label', '—')}</b> (EV: {best.get('ev', 0)}%) ⚠️"
                else:
                    result += "⚠️ <b>Ставка с низким EV</b>. Рекомендуется пропустить."
                return result
        return f"❌ Матч '{match_name}' не найден в кэше. Запустите /update сначала."
    except Exception as e:
        logger.error(f"Ошибка анализа матча: {e}")
        return f"❌ Ошибка: {e}"

def update_odds_for_matches(matches):
    updated_matches = []
    for match_data in matches:
        try:
            home = match_data.get('home')
            away = match_data.get('away')
            league = match_data.get('league')
            fixture_id = match_data.get('fixture_id')
            best_bet = match_data.get('best_bet', {})
            bet_type = best_bet.get('type', 'under')
            new_odds = None
            bookmaker = '—'
            source = None
            odds_data = odds_api.get_odds_for_match(home, away, league)
            if odds_data and odds_data.get('best_odds', 0) > 0:
                if bet_type == 'under' and odds_data.get('under_odds', 0) > 0:
                    new_odds = odds_data['under_odds']
                elif bet_type == 'over' and odds_data.get('over_odds', 0) > 0:
                    new_odds = odds_data['over_odds']
                elif bet_type in ['1X', 'П1'] and odds_data.get('home_odds', 0) > 0:
                    new_odds = odds_data['home_odds']
                elif bet_type in ['X2', 'П2'] and odds_data.get('away_odds', 0) > 0:
                    new_odds = odds_data['away_odds']
                else:
                    new_odds = odds_data.get('best_odds', 0)
                if new_odds and new_odds > 0:
                    bookmaker = odds_data.get('bookmaker_name', 'Odds API')
                    source = 'Odds API'
                    logger.info(f"✅ Odds API: {home} vs {away} | {new_odds} ({bookmaker})")
            if not new_odds or new_odds <= 0:
                if fixture_id:
                    logger.info(f"📡 Odds API не нашел, пробуем Football API для {home} vs {away} (ID: {fixture_id})")
                    football_odds = football_api.get_match_odds(fixture_id)
                    if football_odds:
                        if bet_type == 'under' and football_odds.get('under_odds', 0) > 0:
                            new_odds = football_odds['under_odds']
                            bookmaker = football_odds.get('bookmaker', 'Football API')
                            source = 'Football API'
                        elif bet_type == 'over' and football_odds.get('over_odds', 0) > 0:
                            new_odds = football_odds['over_odds']
                            bookmaker = football_odds.get('bookmaker', 'Football API')
                            source = 'Football API'
                        elif bet_type in ['1X', 'П1'] and football_odds.get('home_odds', 0) > 0:
                            new_odds = football_odds['home_odds']
                            bookmaker = football_odds.get('bookmaker', 'Football API')
                            source = 'Football API'
                        elif bet_type in ['X2', 'П2'] and football_odds.get('away_odds', 0) > 0:
                            new_odds = football_odds['away_odds']
                            bookmaker = football_odds.get('bookmaker', 'Football API')
                            source = 'Football API'
                        elif bet_type == 'btts' and football_odds.get('btts_yes', 0) > 0:
                            new_odds = football_odds['btts_yes']
                            bookmaker = football_odds.get('bookmaker', 'Football API')
                            source = 'Football API'
                        else:
                            new_odds = football_odds.get('best_odds', 0)
                            if new_odds > 0:
                                bookmaker = football_odds.get('bookmaker', 'Football API')
                                source = 'Football API'
                        if new_odds and new_odds > 0:
                            logger.info(f"✅ Football API: {home} vs {away} | {new_odds} ({bookmaker})")
            if new_odds and new_odds > 0:
                prob = best_bet.get('prob', 0) / 100
                new_ev = (prob * new_odds) - 1
                best_bet['odds'] = round(new_odds, 2)
                best_bet['ev'] = round(new_ev * 100, 1)
                best_bet['bookmaker'] = bookmaker
                best_bet['odds_source'] = source
                match_data['best_bet'] = best_bet
                match_data['odds_updated'] = True
                logger.info(f"✅ ИТОГ: {home} vs {away} | {best_bet['label']} | КЭФ: {new_odds} | EV: {best_bet['ev']}% | Источник: {source}")
            else:
                logger.info(f"ℹ️ Кэфы не найдены для {home} vs {away}, оставляем 1.95")
            updated_matches.append(match_data)
        except Exception as e:
            logger.error(f"❌ Ошибка обновления коэффициентов {match_data.get('home')}: {e}")
            updated_matches.append(match_data)
    return updated_matches

# ============================================================
# ПОИСК МАТЧЕЙ
# ============================================================
def get_matches_with_factors():
    all_matches = []
    today = datetime.now().strftime('%Y-%m-%d')
    dates_to_search = [today]
    logger.info(f"🔍 Поиск матчей на: {today}")
    all_leagues = Config.LEAGUES + getattr(Config, 'CUP_LEAGUES', [])
    logger.info(f"📊 Всего соревнований: {len(all_leagues)}")
    for league_id in all_leagues:
        for search_date in dates_to_search:
            try:
                matches = football_api.get_matches(league_id, search_date)
                league_name = Config.LEAGUE_NAMES.get(league_id, str(league_id))
                send_telegram(f"✅ Лига <b>{league_name}</b> обработана. Всего найдено матчей: {len(all_matches)}")
                if not matches or not isinstance(matches, list):
                    logger.info(f"🔥 Нет матчей в {league_name} на {search_date}")
                    continue
                for match in matches:
                    if not isinstance(match, dict):
                        continue
                    fixture = match.get("fixture")
                    if not fixture or not isinstance(fixture, dict):
                        continue
                    status = fixture.get("status", {})
                    if not isinstance(status, dict):
                        continue
                    if status.get("short") == "NS":
                        match_id = fixture.get("id")
                        if not match_id:
                            continue
                        existing_ids = []
                        for m in all_matches:
                            if isinstance(m, dict):
                                existing_ids.append(m.get("fixture", {}).get("id"))
                        if match_id in existing_ids:
                            continue
                        teams = match.get("teams", {})
                        if not isinstance(teams, dict):
                            continue
                        home_team = teams.get("home", {})
                        away_team = teams.get("away", {})
                        if not isinstance(home_team, dict) or not isinstance(away_team, dict):
                            continue
                        home_id = home_team.get("id")
                        away_id = away_team.get("id")
                        if not home_id or not away_id:
                            continue
                        match["factors"] = {
                            "home_form": football_api.get_form(home_id) if home_id else None,
                            "away_form": football_api.get_form(away_id) if away_id else None,
                            "home_injuries_list": football_api.get_injuries(home_id) if home_id else [],
                            "away_injuries_list": football_api.get_injuries(away_id) if away_id else [],
                            "home_id": home_id,
                            "away_id": away_id,
                            "referee": fixture.get("referee")
                        }
                        match["weather"] = None
                        match["weather_reason"] = "🌤️ Погода отключена"
                        league_data = match.get("league", {})
                        if isinstance(league_data, dict):
                            league_data["name"] = league_name
                        all_matches.append(match)
            except Exception as e:
                error_msg = f"Ошибка {league_name} на {search_date}: {e}"
                logger.error(f"❌ {error_msg}")
                send_error_to_telegram(error_msg)
            time.sleep(0.1)
    logger.info(f"📊 ВСЕГО найдено матчей: {len(all_matches)}")
    return all_matches

# ============================================================
# ТОП МАТЧЕЙ - 70%+ (МАКСИМАЛЬНАЯ ТОЧНОСТЬ)
# ============================================================
@timing_decorator()
def find_top_matches(matches):
    bank = storage.load_bank()
    max_bets = Config.MAX_BETS_PER_RUN
    logger.info(f"🔍 Анализ {len(matches)} матчей из всех лиг с фильтрами для 70%+...")
    best_matches = []
    bet_type_count = {}
    league_count = {}
    
    for match in matches:
        if not match or not isinstance(match, dict):
            continue
        try:
            fixture = match.get("fixture")
            if not fixture or not isinstance(fixture, dict):
                continue
            fixture_id = fixture.get("id")
            if not fixture_id:
                continue
            
            teams = match.get("teams")
            if not teams or not isinstance(teams, dict):
                continue
            home_team = teams.get("home")
            away_team = teams.get("away")
            if not isinstance(home_team, dict) or not isinstance(away_team, dict):
                continue
            home = home_team.get("name", "Unknown")
            away = away_team.get("name", "Unknown")
            
            league_data = match.get("league")
            league_name = league_data.get('name', 'Unknown') if isinstance(league_data, dict) else "Unknown"
            if league_name == 'Unknown':
                league_id = league_data.get('id') if isinstance(league_data, dict) else None
                if league_id and league_id in Config.LEAGUE_NAMES:
                    league_name = Config.LEAGUE_NAMES[league_id]
            league_id = league_data.get('id') if isinstance(league_data, dict) else None
            match_time = fixture.get("date", "")
            if match_time:
                try:
                    dt = datetime.fromisoformat(match_time.replace("Z", "+00:00"))
                    dt = dt + timedelta(hours=TIMEZONE_OFFSET)
                    match_time = dt.strftime("%d.%m.%Y %H:%M")
                except:
                    match_time = "Время не указано"
            
            # ПУНКТ 4: Реальные голы из формы
            home_form_data = cache_manager.get('form', home_team.get("id"), ttl_hours=6) or football_api.get_form(home_team.get("id"))
            away_form_data = cache_manager.get('form', away_team.get("id"), ttl_hours=6) or football_api.get_form(away_team.get("id"))
            cache_manager.set('form', home_team.get("id"), home_form_data)
            cache_manager.set('form', away_team.get("id"), away_form_data)
            
            home_form = home_form_data.get('form', '') if home_form_data else ''
            away_form = away_form_data.get('form', '') if away_form_data else ''
            
            home_goals_avg = home_form_data.get('goals_avg', 1.2) if home_form_data else 1.2
            away_goals_avg = away_form_data.get('goals_avg', 1.0) if away_form_data else 1.0
            home_conceded_avg = home_form_data.get('conceded_avg', 1.0) if home_form_data else 1.0
            away_conceded_avg = away_form_data.get('conceded_avg', 1.2) if away_form_data else 1.2
            
            # ПУНКТ 4: Real XG (модель из форм)
            home_xg = (home_goals_avg + away_conceded_avg) / 2
            away_xg = (away_goals_avg + home_conceded_avg) / 2
            
            # ПУНКТ 2: Разница в голах (коэффициент)
            standings = cache_manager.get('standings', league_id, ttl_hours=12) or football_api.get_standings(league_id)
            cache_manager.set('standings', league_id, standings)
            home_position = 99
            away_position = 99
            home_gd = 0
            away_gd = 0
            if standings:
                if home in standings:
                    home_position = standings[home].get('position', 99)
                    home_gd = standings[home].get('goals_diff', 0)
                if away in standings:
                    away_position = standings[away].get('position', 99)
                    away_gd = standings[away].get('goals_diff', 0)
            
            home_motivation = get_motivation(home_position)
            away_motivation = get_motivation(away_position)
            if home_motivation == 'mid_table' and away_motivation == 'mid_table':
                logger.info(f"⏭️ Пропускаем (нет мотивации): {home} vs {away}")
                continue
            
            # ПУНКТ 3: Травмы
            h_inj = len(match['factors'].get('home_injuries_list', []))
            a_inj = len(match['factors'].get('away_injuries_list', []))
            if h_inj > 3:
                home_xg *= 0.8
            if a_inj > 3:
                away_xg *= 0.8
            
            # ПУНКТ 2: Разница в позициях (разница в klasse)
            if home_position < away_position - 10:
                home_win_prob += 0.10
                away_win_prob -= 0.10
            if away_position < home_position - 10:
                away_win_prob += 0.10
                home_win_prob -= 0.10
            
            # ПУНКТ 1: Смарт кэфы (микро-оптимизация)
            home_win_prob = 0.55 + (home_xg - away_xg) * 0.2 - (h_inj - a_inj) * 0.02
            draw_prob = 0.25
            away_win_prob = 0.20 - (home_xg - away_xg) * 0.2
            home_win_prob = 0.55 + (home_xg - away_xg) * 0.2 - (h_inj - a_inj) * 0.02
            draw_prob = 0.25
            away_win_prob = 0.20 - (home_xg - away_xg) * 0.2
            
            # ПУНКТ 2: Разница в позициях (разница в klasse)
            if home_position < away_position - 10:
                home_win_prob += 0.10
                away_win_prob -= 0.10
            if away_position < home_position - 10:
                away_win_prob += 0.10
                home_win_prob -= 0.10
            
            # ПУНКТ 1: Смарт кэфы (микро-оптимизация)
            if home_xg > away_xg + 0.5:
                home_win_prob += 0.08
                away_win_prob -= 0.08
            elif home_xg < away_xg - 0.5:
                away_win_prob += 0.08
                home_win_prob -= 0.08
            
            # Нормализация (sum = 1)
            total_prob = home_win_prob + draw_prob + away_win_prob
            if total_prob > 0:
                home_win_prob /= total_prob
                draw_prob /= total_prob
                away_win_prob /= total_prob
            
            prob_1X = home_win_prob + draw_prob
            prob_X2 = away_win_prob + draw_prob
            prob_over_2_5 = 1 - (home_win_prob * away_win_prob)
            prob_under_2_5 = 1 - prob_over_2_5
            
            # ПУНКТ 1: Real Fair Odds
            odds = {
                '1X': 1.85,
                'X2': 1.85,
                'П1': 2.10,
                'П2': 2.10,
                'ТМ 2.5': 1.95,
                'ТБ 2.5': 1.95,
                'ОБЗ': 1.90,
            }
            odds['1X'] = 1.85 if prob_1X > 0.70 else 1.75
            odds['X2'] = 1.85 if prob_X2 > 0.70 else 1.75
            odds['П1'] = 2.10 if prob_home_win > 0.60 else 2.10
            odds['П2'] = 2.10 if prob_away_win > 0.60 else 2.10
            odds['ТМ 2.5'] = 1.95 if prob_under_2_5 > 0.60 else 1.95
            odds['ТБ 2.5'] = 1.95 if prob_over_2_5 > 0.60 else 1.95
            
            # BETS
            bets = []
            ev_1x = (prob_1X * odds['1X']) - 1
            bets.append({
                'type': '1X', 'label': '1X', 'prob': round(prob_1X * 100, 1),
                'ev': round(ev_1x * 100, 1), 'odds': odds['1X'], 'stake': round(42.86875, 2)
            })
            
            ev_x2 = (prob_X2 * odds['X2']) - 1
            bets.append({
                'type': 'X2', 'label': 'X2', 'prob': round(prob_X2 * 100, 1),
                'ev': round(ev_x2 * 100, 1), 'odds': odds['X2'], 'stake': round(42.86875, 2)
            })
            
            ev_p1 = (prob_home_win * odds['П1']) - 1
            bets.append({
                'type': 'П1', 'label': 'П1', 'prob': round(prob_home_win * 100, 1),
                'ev': round(ev_p1 * 100, 1), 'odds': odds['П1'], 'stake': round(42.86875, 2)
            })
            
            ev_p2 = (prob_away_win * odds['П2']) - 1
            bets.append({
                'type': 'П2', 'label': 'П2', 'prob': round(prob_away_win * 100, 1),
                'ev': round(ev_p2 * 100, 1), 'odds': odds['П2'], 'stake': round(42.86875, 2)
            })
            
            ev_under = (prob_under_2_5 * odds['ТМ 2.5']) - 1
            bets.append({
                'type': 'under', 'label': 'ТМ 2.5', 'prob': round(prob_under_2_5 * 100, 1),
                'ev': round(ev_under * 100, 1), 'odds': odds['ТМ 2.5'], 'stake': round(42.86875, 2)
            })
            
            ev_over = (prob_over_2_5 * odds['ТБ 2.5']) - 1
            bets.append({
                'type': 'over', 'label': 'ТБ 2.5', 'prob': round(prob_over_2_5 * 100, 1),
                'ev': round(ev_over * 100, 1), 'odds': odds['ТБ 2.5'], 'stake': round(42.86875, 2)
            })
            
            bets.sort(key=lambda x: x['ev'], reverse=True)
            best_bet = bets[0]
            
            if best_bet['ev'] < ev_min:
                logger.info(f"⏭️ Пропускаем (EV < {ev_min}%): {home} vs {away} | EV: {best_bet['ev']}%")
                continue
            if best_bet['prob'] < prob_min:
                logger.info(f"⏭️ Пропускаем (Prob < {prob_min}%): {home} vs {away} | Prob: {best_bet['prob']}%")
                continue
            
            bet_type = best_bet['type']
            bet_type_count[bet_type] = bet_type_count.get(bet_type, 0) + 1
            if bet_type_count[bet_type] > 3:
                logger.info(f"⏭️ Пропускаем (лимит типа {bet_type}): {home} vs {away}")
                continue
            
            league_count[league_name] = league_count.get(league_name, 0) + 1
            if league_count[league_name] > 2:
                logger.info(f"⏭️ Пропускаем (лимит лиги {league_name}): {home} vs {away}")
                continue
            
            match_data = {
                "home": home, "away": away, "league": league_name, "fixture_id": fixture_id,
                "match_time": match_time, "home_xg": round(home_xg, 2), "away_xg": round(away_xg, 2),
                "total_xg": round(total_xg, 2), "home_form": home_form, "away_form": away_form,
                "home_form_quality": analyze_form(home_form), "away_form_quality": analyze_form(away_form),
                "home_goals_avg": home_goals_avg, "away_goals_avg": away_goals_avg,
                "home_conceded_avg": home_conceded_avg, "away_conceded_avg": away_conceded_avg,
                "standings": {"home_position": home_position, "away_position": away_position,
                              "home_motivation": home_motivation, "away_motivation": away_motivation},
                "bets": bets, "best_bet": best_bet, "weather_reason": "🌤️", "factors": {},
                "source": "70_percent"
            }
            best_matches.append(match_data)
            logger.info(f"✅ КАНДИДАТ (70%+): {home} vs {away} | ЛУЧШАЯ СТАВКА: {best_bet['label']} | EV: {best_bet['ev']}% | Prob: {best_bet['prob']}%")
        except Exception as e:
            logger.error(f"❌ Ошибка: {e}")
            continue
    
    best_matches.sort(key=lambda x: x['best_bet']['ev'], reverse=True)
    top_matches = best_matches[:max_bets]
    logger.info(f"📊 Найдено {len(best_matches)} кандидатов (70%+), выбрано {len(top_matches)} лучших")
    return top_matches

# ============================================================
# ПОИСК ТМ 2.5
# ============================================================
@timing_decorator()
def find_tm25_matches(matches):
    tm25_candidates = []
    from app.config import Config
    MAX_TM25_BETS = getattr(Config, 'MAX_TM25_BETS', 5)
    PREMIUM_MIN_EV = getattr(Config, 'PREMIUM_MIN_EV', 30) / 100
    PREMIUM_MIN_PROB = getattr(Config, 'PREMIUM_MIN_PROB', 60) / 100
    PREMIUM_XG_MIN = getattr(Config, 'PREMIUM_XG_MIN', 1.0)
    PREMIUM_XG_MAX = getattr(Config, 'PREMIUM_XG_MAX', 2.8)
    STANDARD_MIN_EV = getattr(Config, 'STANDARD_MIN_EV', 15) / 100
    STANDARD_MIN_PROB = getattr(Config, 'STANDARD_MIN_PROB', 50) / 100
    STANDARD_XG_MIN = getattr(Config, 'TM25_XG_MIN', 0.8)
    STANDARD_XG_MAX = getattr(Config, 'TM25_XG_MAX', 3.0)
    TOP_LEAGUES = getattr(Config, 'TOP_LEAGUES', [])
    TM25_TOP_LEAGUE_EV = getattr(Config, 'TM25_TOP_LEAGUE_EV', 35) / 100
    logger.info("🔍 Специальный поиск ТМ 2.5 (двухуровневый)...")
    logger.info(f"📊 УРОВЕНЬ 1 (PREMIUM): EV>{PREMIUM_MIN_EV*100}%, Prob>{PREMIUM_MIN_PROB*100}%, XG {PREMIUM_XG_MIN}-{PREMIUM_XG_MAX}")
    logger.info(f"📊 УРОВЕНЬ 2 (STANDARD): EV>{STANDARD_MIN_EV*100}%, Prob>{STANDARD_MIN_PROB*100}%, XG {STANDARD_XG_MIN}-{STANDARD_XG_MAX}")
    logger.info(f"📊 Всего матчей для анализа: {len(matches)}")
    stats = {'total': 0, 'premium_found': 0, 'standard_found': 0}
    logger.info("🎯 ПОИСК УРОВНЯ 1 (PREMIUM EV>30%)...")
    for match in matches:
        if not match or not isinstance(match, dict):
            continue
        stats['total'] += 1
        try:
            fixture = match.get("fixture")
            if not fixture or not isinstance(fixture, dict):
                continue
            fixture_id = fixture.get("id")
            if not fixture_id:
                continue
            teams = match.get("teams")
            if not teams or not isinstance(teams, dict):
                continue
            home_team = teams.get("home")
            away_team = teams.get("away")
            if not isinstance(home_team, dict) or not isinstance(away_team, dict):
                continue
            home = home_team.get("name", "Unknown")
            away = away_team.get("name", "Unknown")
            league_data = match.get("league")
            league_name = league_data.get("name", "Unknown") if isinstance(league_data, dict) else "Unknown"
            league_id = league_data.get("id") if isinstance(league_data, dict) else None
            match_time = fixture.get("date", "")
            if match_time:
                try:
                    dt = datetime.fromisoformat(match_time.replace("Z", "+00:00"))
                    dt = dt + timedelta(hours=TIMEZONE_OFFSET)
                    match_time = dt.strftime("%d.%m.%Y %H:%M")
                except:
                    match_time = "Время не указано"
            statistics = football_api.get_match_statistics(fixture_id)
            home_xg = 1.2
            away_xg = 1.0
            if statistics:
                for team_name, stats_dict in statistics.items():
                    if home.lower() in team_name.lower() or team_name.lower() in home.lower():
                        xg_val = stats_dict.get('xG')
                        if xg_val is not None and xg_val > 0:
                            home_xg = float(xg_val)
                    elif away.lower() in team_name.lower() or team_name.lower() in away.lower():
                        xg_val = stats_dict.get('xG')
                        if xg_val is not None and xg_val > 0:
                            away_xg = float(xg_val)
            if home_xg == 1.2 and away_xg == 1.0:
                if league_name in FALLBACK_XG:
                    home_xg = FALLBACK_XG[league_name]['home']
                    away_xg = FALLBACK_XG[league_name]['away']
                else:
                    home_xg = 1.3
                    away_xg = 1.0
                random.seed(fixture_id)
                home_xg *= (1 + random.uniform(-0.1, 0.1))
                away_xg *= (1 + random.uniform(-0.1, 0.1))
            home_adv = HOME_ADVANTAGE.get(league_name, 1.10)
            home_xg *= home_adv
            away_xg /= home_adv
            total_xg = home_xg + away_xg
            if total_xg < PREMIUM_XG_MIN or total_xg > PREMIUM_XG_MAX:
                continue
            home_form_data = football_api.get_form(home_team.get("id"))
            away_form_data = football_api.get_form(away_team.get("id"))
            home_form = home_form_data.get('form', '') if home_form_data else ''
            away_form = away_form_data.get('form', '') if away_form_data else ''
            standings = football_api.get_standings(league_id) if league_id else None
            home_position = 99
            away_position = 99
            if standings:
                if home in standings:
                    home_position = standings[home].get('position', 99)
                if away in standings:
                    away_position = standings[away].get('position', 99)
            h2h_data = football_api.get_head_to_head(home, away)
            probs = ensemble_probability(home_xg, away_xg, home_form, away_form, h2h_data)
            prob_under_2_5 = probs['under_2_5']
            odds_tm25 = 1.95
            ev_under = (prob_under_2_5 * odds_tm25) - 1
            if ev_under >= PREMIUM_MIN_EV and prob_under_2_5 >= PREMIUM_MIN_PROB:
                if league_name in TOP_LEAGUES and ev_under < 0.35:
                    continue
                best_bet = {
                    'type': 'under',
                    'label': 'ТМ 2.5 🔥',
                    'prob': round(prob_under_2_5 * 100, 1),
                    'ev': round(ev_under * 100, 1),
                    'odds': odds_tm25,
                    'stake': round(42.86875, 2),
                    'level': 'PREMIUM'
                }
                match_data = {
                    "home": home,
                    "away": away,
                    "league": league_name,
                    "fixture_id": fixture_id,
                    "match_time": match_time,
                    "home_xg": round(home_xg, 2),
                    "away_xg": round(away_xg, 2),
                    "total_xg": round(total_xg, 2),
                    "home_form": home_form,
                    "away_form": away_form,
                    "standings": {
                        "home_position": home_position,
                        "away_position": away_position,
                    },
                    "bets": [best_bet],
                    "best_bet": best_bet,
                    "source": "tm25_premium",
                    "weather_reason": "🌤️",
                }
                tm25_candidates.append(match_data)
                stats['premium_found'] += 1
                logger.info(f"🔥 PREMIUM ТМ2.5: {home} vs {away} | EV: {ev_under*100:.1f}% | Prob: {prob_under_2_5*100:.1f}% | XG: {total_xg:.2f}")
                if len(tm25_candidates) >= MAX_TM25_BETS:
                    logger.info(f"⏹️ Достигнут лимит PREMIUM ({MAX_TM25_BETS}), остановка поиска")
                    break
        except Exception as e:
            logger.error(f"❌ Ошибка PREMIUM: {e}")
            continue
    if len(tm25_candidates) < MAX_TM25_BETS:
        logger.info(f"🎯 PREMIUM найдено: {len(tm25_candidates)}, ищем STANDARD (EV>15%)...")
        for match in matches:
            if not match or not isinstance(match, dict):
                continue
            existing_keys = [f"{m['home']}_{m['away']}" for m in tm25_candidates]
            if f"{match.get('teams', {}).get('home', {}).get('name', '')}_{match.get('teams', {}).get('away', {}).get('name', '')}" in existing_keys:
                continue
            try:
                fixture = match.get("fixture")
                if not fixture or not isinstance(fixture, dict):
                    continue
                fixture_id = fixture.get("id")
                if not fixture_id:
                    continue
                teams = match.get("teams")
                if not teams or not isinstance(teams, dict):
                    continue
                home_team = teams.get("home")
                away_team = teams.get("away")
                if not isinstance(home_team, dict) or not isinstance(away_team, dict):
                    continue
                home = home_team.get("name", "Unknown")
                away = away_team.get("name", "Unknown")
                league_data = match.get("league")
                league_name = league_data.get("name", "Unknown") if isinstance(league_data, dict) else "Unknown"
                league_id = league_data.get("id") if isinstance(league_data, dict) else None
                match_time = fixture.get("date", "")
                if match_time:
                    try:
                        dt = datetime.fromisoformat(match_time.replace("Z", "+00:00"))
                        dt = dt + timedelta(hours=TIMEZONE_OFFSET)
                        match_time = dt.strftime("%d.%m.%Y %H:%M")
                    except:
                        match_time = "Время не указано"
                statistics = football_api.get_match_statistics(fixture_id)
                home_xg = 1.2
                away_xg = 1.0
                if statistics:
                    for team_name, stats_dict in statistics.items():
                        if home.lower() in team_name.lower() or team_name.lower() in home.lower():
                            xg_val = stats_dict.get('xG')
                            if xg_val is not None and xg_val > 0:
                                home_xg = float(xg_val)
                        elif away.lower() in team_name.lower() or team_name.lower() in away.lower():
                            xg_val = stats_dict.get('xG')
                            if xg_val is not None and xg_val > 0:
                                away_xg = float(xg_val)
                if home_xg == 1.2 and away_xg == 1.0:
                    if league_name in FALLBACK_XG:
                        home_xg = FALLBACK_XG[league_name]['home']
                        away_xg = FALLBACK_XG[league_name]['away']
                    else:
                        home_xg = 1.3
                        away_xg = 1.0
                    random.seed(fixture_id)
                    home_xg *= (1 + random.uniform(-0.1, 0.1))
                    away_xg *= (1 + random.uniform(-0.1, 0.1))
                home_adv = HOME_ADVANTAGE.get(league_name, 1.10)
                home_xg *= home_adv
                away_xg /= home_adv
                total_xg = home_xg + away_xg
                if total_xg < STANDARD_XG_MIN or total_xg > STANDARD_XG_MAX:
                    continue
                home_form_data = football_api.get_form(home_team.get("id"))
                away_form_data = football_api.get_form(away_team.get("id"))
                home_form = home_form_data.get('form', '') if home_form_data else ''
                away_form = away_form_data.get('form', '') if away_form_data else ''
                standings = football_api.get_standings(league_id) if league_id else None
                home_position = 99
                away_position = 99
                if standings:
                    if home in standings:
                        home_position = standings[home].get('position', 99)
                    if away in standings:
                        away_position = standings[away].get('position', 99)
                h2h_data = football_api.get_head_to_head(home, away)
                probs = ensemble_probability(home_xg, away_xg, home_form, away_form, h2h_data)
                prob_under_2_5 = probs['under_2_5']
                odds_tm25 = 1.95
                ev_under = (prob_under_2_5 * odds_tm25) - 1
                if ev_under >= STANDARD_MIN_EV and prob_under_2_5 >= STANDARD_MIN_PROB:
                    if league_name in TOP_LEAGUES and ev_under < 0.20:
                        continue
                    best_bet = {
                        'type': 'under',
                        'label': 'ТМ 2.5',
                        'prob': round(prob_under_2_5 * 100, 1),
                        'ev': round(ev_under * 100, 1),
                        'odds': odds_tm25,
                        'stake': round(42.86875, 2),
                        'level': 'STANDARD'
                    }
                    match_data = {
                        "home": home,
                        "away": away,
                        "league": league_name,
                        "fixture_id": fixture_id,
                        "match_time": match_time,
                        "home_xg": round(home_xg, 2),
                        "away_xg": round(away_xg, 2),
                        "total_xg": round(total_xg, 2),
                        "home_form": home_form,
                        "away_form": away_form,
                        "standings": {
                            "home_position": home_position,
                            "away_position": away_position,
                        },
                        "bets": [best_bet],
                        "best_bet": best_bet,
                        "source": "tm25_standard",
                        "weather_reason": "🌤️",
                    }
                    tm25_candidates.append(match_data)
                    stats['standard_found'] += 1
                    logger.info(f"✅ STANDARD ТМ2.5: {home} vs {away} | EV: {ev_under*100:.1f}% | Prob: {prob_under_2_5*100:.1f}% | XG: {total_xg:.2f}")
                    if len(tm25_candidates) >= MAX_TM25_BETS:
                        logger.info(f"⏹️ Достигнут лимит STANDARD ({MAX_TM25_BETS}), остановка поиска")
                        break
            except Exception as e:
                logger.error(f"❌ Ошибка STANDARD: {e}")
                continue
    logger.info(f"📊 СТАТИСТИКА ТМ2.5: PREMIUM: {stats['premium_found']}, STANDARD: {stats['standard_found']}")
    tm25_candidates.sort(key=lambda x: x['best_bet']['ev'], reverse=True)
    logger.info(f"📊 Найдено ТМ2.5 кандидатов: {len(tm25_candidates)} (PREMIUM: {stats['premium_found']}, STANDARD: {stats['standard_found']})")
    return tm25_candidates

# ============================================================
# ОБЪЕДИНЕННЫЙ ПОИСК
# ============================================================
@timing_decorator()
def find_top_matches_with_tm25(matches):
    logger.info("=" * 50)
    logger.info("📊 ПОТОК 1: Поиск 70%+ матчей")
    logger.info("=" * 50)
    top_matches_70 = find_top_matches(matches)
    logger.info("=" * 50)
    logger.info("📊 ПОТОК 2: Поиск ТМ 2.5 матчей")
    logger.info("=" * 50)
    tm25_matches = find_tm25_matches(matches)
    combined_matches = []
    match_keys = set()
    for m in top_matches_70:
        key = f"{m['home']}_{m['away']}"
        if key not in match_keys:
            combined_matches.append(m)
            match_keys.add(key)
    for m in tm25_matches:
        key = f"{m['home']}_{m['away']}"
        if key not in match_keys:
            combined_matches.append(m)
            match_keys.add(key)
            logger.info(f"🔄 Добавлен ТМ2.5 матч (уникальный): {m['home']} vs {m['away']}")
    combined_matches.sort(key=lambda x: x['best_bet']['ev'], reverse=True)
    if combined_matches:
        logger.info(f"📡 Запрос реальных коэффициентов для {len(combined_matches)} матчей...")
        combined_matches = update_odds_for_matches(combined_matches)
    max_total = Config.MAX_BETS_PER_RUN + getattr(Config, 'MAX_TM25_BETS', 5)
    combined_matches = combined_matches[:max_total]
    cache = storage.load_cache()
    cache['top_matches'] = combined_matches
    storage.save_cache(cache)
    history = storage.load_history()
    for match_data in combined_matches:
        best_bet = match_data.get('best_bet', {})
        bet_record = {
            'home': match_data.get('home', 'Unknown'),
            'away': match_data.get('away', 'Unknown'),
            'league': match_data.get('league', 'Unknown'),
            'bet': best_bet.get('label', '—'),
            'odds': best_bet.get('odds', 0),
            'stake': best_bet.get('stake', 42.87),
            'ev': best_bet.get('ev', 0),
            'result': 'pending',
            'profit': 0,
            'date': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'fixture_id': match_data.get('fixture_id'),
            'bookmaker': best_bet.get('bookmaker', '—')
        }
        history.append(bet_record)
    storage.save_history(history)
    logger.info(f"💾 Сохранено {len(combined_matches)} ставок в историю")
    logger.info("=" * 50)
    logger.info(f"📊 ИТОГО: {len(combined_matches)} матчей (70%+: {len(top_matches_70)}, ТМ2.5: {len(tm25_matches)})")
    logger.info("=" * 50)
    return combined_matches

# ============================================================
# ОБНОВЛЕНИЕ РЕЗУЛЬТАТОВ
# ============================================================
def determine_bet_result(bet_type, home_goals, away_goals):
    total = home_goals + away_goals
    bet_type_lower = bet_type.lower()
    if 'п1' in bet_type_lower:
        return 'win' if home_goals > away_goals else ('push' if home_goals == away_goals else 'loss')
    elif 'п2' in bet_type_lower:
        return 'win' if away_goals > home_goals else ('push' if home_goals == away_goals else 'loss')
    elif '1x' in bet_type_lower:
        return 'win' if home_goals >= away_goals else 'loss'
    elif 'x2' in bet_type_lower:
        return 'win' if away_goals >= home_goals else 'loss'
    elif 'обз' in bet_type_lower or 'btts' in bet_type_lower:
        return 'win' if home_goals > 0 and away_goals > 0 else 'loss'
    elif 'тм 2.5' in bet_type_lower or 'under' in bet_type_lower:
        return 'win' if total < 2.5 else 'loss'
    elif 'тб 2.5' in bet_type_lower or 'over' in bet_type_lower:
        return 'win' if total > 2.5 else 'loss'
    return 'pending'

@timing_decorator()
def update_pending_bets():
    history = storage.load_history()
    updated = 0
    for bet in history:
        if bet.get('result') == 'pending' or bet.get('result') is None:
            fixture_id = bet.get('fixture_id')
            if not fixture_id:
                home = bet.get('home', '')
                away = bet.get('away', '')
                if home and away:
                    fixture_id = football_api.find_fixture_by_teams(home, away)
                    if fixture_id:
                        bet['fixture_id'] = fixture_id
            if fixture_id:
                match_data = football_api.get_match_result(fixture_id)
                if match_data:
                    home_goals = match_data['goals']['home']
                    away_goals = match_data['goals']['away']
                    if home_goals is not None and away_goals is not None:
                        bet_type = bet.get('bet', '')
                        result = determine_bet_result(bet_type, home_goals, away_goals)
                        if result != 'pending':
                            bet['result'] = result
                            bet['home_goals'] = home_goals
                            bet['away_goals'] = away_goals
                            if result == 'win':
                                bet['profit'] = round(bet['stake'] * (bet['odds'] - 1), 2)
                            elif result == 'loss':
                                bet['profit'] = -bet['stake']
                            else:
                                bet['profit'] = 0
                            updated += 1
                            logger.info(f"✅ Обновлена ставка: {bet['home']} vs {bet['away']} → {result} ({home_goals}-{away_goals})")
    if updated > 0:
        storage.save_history(history)
        recalc_stats()
    return updated

def recalc_stats():
    history = storage.load_history()
    stats = storage.load_stats()
    total = len(history)
    wins = sum(1 for b in history if b.get('result') == 'win')
    losses = sum(1 for b in history if b.get('result') == 'loss')
    pushes = sum(1 for b in history if b.get('result') == 'push')
    total_profit = sum(b.get('profit', 0) for b in history)
    total_stake = sum(b.get('stake', 0) for b in history)
    stats['total'] = total
    stats['wins'] = wins
    stats['losses'] = losses
    stats['pushes'] = pushes
    stats['total_profit'] = round(total_profit, 2)
    stats['winrate'] = round(wins / (wins + losses) * 100, 1) if (wins + losses) > 0 else 0
    stats['roi'] = round((total_profit / total_stake * 100), 1) if total_stake > 0 else 0
    storage.save_stats(stats)
    logger.info(f"📊 Статистика пересчитана: {stats}")

def find_fixture_by_teams(self, home_team, away_team):
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        params = {
            'date': today,
            'status': 'FT'
        }
        data = self._make_request('/fixtures', params)
        if data and 'response' in data:
            for fixture in data['response']:
                teams = fixture.get('teams', {})
                home = teams.get('home', {}).get('name', '')
                away = teams.get('away', {}).get('name', '')
                if home_team.lower() in home.lower() and away_team.lower() in away.lower():
                    return fixture.get('fixture', {}).get('id')
    except Exception as e:
        logger.error(f"Ошибка поиска матча {home_team} vs {away_team}: {e}")
    return None

# Добавляем метод в FootballAPI
FootballAPI.find_fixture_by_teams = find_fixture_by_teams

# ============================================================
# АВТОМАТИЧЕСКОЕ ОБНОВЛЕНИЕ РЕЗУЛЬТАТОВ
# ============================================================
def schedule_updates():
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        func=auto_update_results,
        trigger='interval',
        hours=6,
        id='auto_update',
        replace_existing=True
    )
    scheduler.start()
    logger.info("⏰ Авто-обновление результатов запущено (каждые 6 часов)")

def auto_update_results():
    try:
        logger.info("🔄 Авто-обновление: проверка результатов...")
        updated = update_pending_bets()
        if updated > 0:
            logger.info(f"✅ Авто-обновление: обновлено {updated} результатов")
            send_telegram(f"🔄 <b>АВТО-ОБНОВЛЕНИЕ</b>\n✅ Обновлено {updated} результатов матчей!")
        else:
            logger.info("📭 Авто-обновление: нет новых результатов")
    except Exception as e:
        logger.error(f"❌ Ошибка авто-обновления: {e}")
        send_error_to_telegram(f"Ошибка авто-обновления: {e}")

def schedule_notifications():
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        func=notification_system.run_all_checks,
        trigger='interval',
        hours=1,
        id='notifications',
        replace_existing=True
    )
    scheduler.start()
    logger.info("⏰ Уведомления запущены (каждые 6 часов)")

def schedule_performance_report():
    def report():
        perf_monitor.print_report()
        report_data = perf_monitor.get_report()
        slow_functions = [f for f in report_data if f['avg_time'] > 3.0]
        if slow_functions:
            msg = "⚠️ <b>МЕДЛЕННЫЕ ФУНКЦИИ</b>\n\n"
            for f in slow_functions:
                msg += f"• {f['function']}: {f['avg_time']}с\n"
            send_telegram(msg)
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        func=report,
        trigger='interval',
        hours=6,
        id='perf_report',
        replace_existing=True
    )
    scheduler.start()
    logger.info("⏰ Отчет производительности запущен (каждые 6 часов)")

# ============================================================
# УЛУЧШЕНИЕ 4: ВЕРИФИКАЦИЯ СТАВОК
# ============================================================
class BetVerificationSystem:
    def __init__(self):
        self.thresholds = {
            'min_odds': 1.50,
            'max_odds': 3.00,
            'min_ev': 15,
            'min_prob': 50,
            'max_stake_percent': 10,
            'min_samples': 10,
        }
        self.warnings = []
    
    def verify(self, bet_data):
        self.warnings = []
        self._check_odds(bet_data)
        self._check_ev_prob(bet_data)
        self._check_stake(bet_data)
        self._check_historical_accuracy(bet_data)
        self._check_xg_correlation(bet_data)
        self._check_league(bet_data)
        self._check_match_time(bet_data)
        self._check_form(bet_data)
        if len(self.warnings) == 0:
            return {'status': '✅', 'message': 'Ставка прошла все проверки'}
        elif len(self.warnings) <= 2:
            return {'status': '⚠️', 'message': f'{len(self.warnings)} предупреждений', 'warnings': self.warnings}
        else:
            return {'status': '❌', 'message': 'Ставка отклонена', 'warnings': self.warnings}
    
    def _check_odds(self, bet_data):
        odds = bet_data.get('odds', 0)
        if odds < self.thresholds['min_odds']:
            self.warnings.append(f"Низкий коэффициент: {odds} (мин {self.thresholds['min_odds']})")
        if odds > self.thresholds['max_odds']:
            self.warnings.append(f"Высокий коэффициент: {odds} (макс {self.thresholds['max_odds']})")
        if odds < 1.10:
            self.warnings.append(f"Подозрительный коэффициент: {odds}")
    
    def _check_ev_prob(self, bet_data):
        ev = bet_data.get('ev', 0)
        prob = bet_data.get('prob', 0)
        if ev < self.thresholds['min_ev']:
            self.warnings.append(f"Низкий EV: {ev}% (мин {self.thresholds['min_ev']}%)")
        if prob < self.thresholds['min_prob']:
            self.warnings.append(f"Низкая вероятность: {prob}% (мин {self.thresholds['min_prob']}%)")
    
    def _check_stake(self, bet_data):
        stake = bet_data.get('stake', 0)
        bank = storage.load_bank()
        if bank > 0:
            stake_percent = (stake / bank) * 100
            if stake_percent > self.thresholds['max_stake_percent']:
                self.warnings.append(f"Ставка {stake_percent:.1f}% от банка (макс {self.thresholds['max_stake_percent']}%)")
        if stake < 1:
            self.warnings.append(f"Минимальная ставка: {stake}")
    
    def _check_historical_accuracy(self, bet_data):
        bet_type = bet_data.get('type', '')
        history = storage.load_history()
        similar_bets = [b for b in history if b.get('bet') == bet_data.get('label', '') and b.get('result') != 'pending']
        if len(similar_bets) >= self.thresholds['min_samples']:
            wins = sum(1 for b in similar_bets if b.get('result') == 'win')
            winrate = wins / len(similar_bets)
            if winrate < 0.4:
                self.warnings.append(f"Исторический винрейт {winrate*100:.1f}% ({wins}/{len(similar_bets)})")
            profit = sum(b.get('profit', 0) for b in similar_bets)
            if profit < 0:
                self.warnings.append(f"Исторический убыток: ${profit:.2f}")
        else:
            if len(similar_bets) == 0:
                self.warnings.append(f"Нет исторических данных для типа {bet_type}")
    
    def _check_xg_correlation(self, bet_data):
        bet_type = bet_data.get('type', '')
        total_xg = bet_data.get('total_xg', 0)
        if bet_type == 'under':
            if total_xg > 3.0:
                self.warnings.append(f"XG {total_xg:.2f} > 3.0 для ТМ 2.5")
            elif total_xg > 2.5:
                self.warnings.append(f"XG {total_xg:.2f} > 2.5 (рисковано для ТМ 2.5)")
        if bet_type == 'over':
            if total_xg < 2.0:
                self.warnings.append(f"XG {total_xg:.2f} < 2.0 для ТБ 2.5")
        home_xg = bet_data.get('home_xg', 0)
        away_xg = bet_data.get('away_xg', 0)
        if abs(home_xg - away_xg) > 1.0:
            self.warnings.append(f"Большая разница XG: {home_xg:.2f} vs {away_xg:.2f}")
    
    def _check_league(self, bet_data):
        league = bet_data.get('league', '')
        top_leagues = ['Premier League', 'La Liga', 'Bundesliga', 'Serie A', 'Ligue 1']
        if league in top_leagues:
            ev = bet_data.get('ev', 0)
            if ev < 25:
                self.warnings.append(f"Топ-лига {league}, EV {ev}% ниже рекомендованного 25%")
    
    def _check_match_time(self, bet_data):
        match_time = bet_data.get('match_time', '')
        if match_time:
            try:
                match_dt = datetime.strptime(match_time, '%d.%m.%Y %H:%M')
                now = datetime.now()
                if match_dt < now:
                    self.warnings.append("Матч уже начался или прошел")
                time_diff = (match_dt - now).total_seconds() / 3600
                if 0 < time_diff < 1:
                    self.warnings.append(f"Матч через {time_diff:.1f} часов (мало времени)")
                if time_diff > 48:
                    self.warnings.append(f"Матч через {time_diff:.1f} часов (коэффициенты могут измениться)")
            except:
                pass
    
    def _check_form(self, bet_data):
        home_form = bet_data.get('home_form', '')
        away_form = bet_data.get('away_form', '')
        if home_form.endswith('LLL'):
            self.warnings.append(f"Хозяева в серии поражений: {home_form[-3:]}")
        if away_form.endswith('LLL'):
            self.warnings.append(f"Гости в серии поражений: {away_form[-3:]}")
        home_position = bet_data.get('standings', {}).get('home_position', 99)
        away_position = bet_data.get('standings', {}).get('away_position', 99)
        if home_position > 18 and away_position > 18:
            self.warnings.append("Обе команды в зоне вылета")
    
    def get_verification_report(self, bet_data):
        result = self.verify(bet_data)
        report = f"📋 <b>ВЕРИФИКАЦИЯ СТАВКИ</b>\n"
        report += f"🏟️ {bet_data.get('home', '')} vs {bet_data.get('away', '')}\n"
        report += f"🎯 {bet_data.get('label', '')}\n\n"
        report += f"Статус: {result['status']} {result['message']}\n"
        if result.get('warnings'):
            report += f"\n⚠️ <b>Предупреждения ({len(result['warnings'])}):</b>\n"
            for warn in result['warnings']:
                report += f"• {warn}\n"
        return report

# ============================================================
# УЛУЧШЕНИЕ 5: УВЕДОМЛЕНИЯ О ВАЖНЫХ СОБЫТИЯХ
# ============================================================
class NotificationSystem:
    def __init__(self):
        self.last_notification = {}
        self.min_interval = 3600
        self.sent_events = set()
    
    def send_if_needed(self, event_type, message, force=False):
        current_time = time.time()
        if not force:
            last = self.last_notification.get(event_type, 0)
            if current_time - last < self.min_interval:
                return
        send_telegram(message)
        self.last_notification[event_type] = current_time
        self.sent_events.add(f"{event_type}_{datetime.now().strftime('%Y-%m-%d')}")
    
    def check_bank_status(self):
        bank = storage.load_bank()
        stats = storage.load_stats()
        if stats.get('total_profit', 0) < 0:
            drawdown = abs(stats['total_profit'])
            bank_percent = (drawdown / bank) * 100 if bank > 0 else 0
            if bank_percent > 20:
                self.send_if_needed('bank_drawdown', f"🔴 <b>КРИТИЧЕСКАЯ ПРОСАДКА БАНКА</b>\nПотеряно: ${drawdown:.2f} ({bank_percent:.1f}%)\nТекущий банк: ${bank:.2f}")
            elif bank_percent > 10:
                self.send_if_needed('bank_drawdown_mid', f"⚠️ <b>ПРОСАДКА БАНКА</b>\nПотеряно: ${drawdown:.2f} ({bank_percent:.1f}%)\nТекущий банк: ${bank:.2f}")
        if stats.get('total_profit', 0) > bank * 0.1:
            self.send_if_needed('bank_growth', f"🟢 <b>РОСТ БАНКА</b>\nПрибыль: ${stats['total_profit']:.2f}\nТекущий банк: ${bank:.2f}\nВинрейт: {stats.get('winrate', 0)}%")
    
    def check_streaks(self):
        history = storage.load_history()
        if len(history) < 5:
            return
        recent = history[-10:]
        current_streak = 0
        streak_type = None
        profit_streak = 0
        for bet in reversed(recent):
            result = bet.get('result')
            if result == 'win':
                if streak_type == 'win' or streak_type is None:
                    current_streak += 1
                    streak_type = 'win'
                    profit_streak += bet.get('profit', 0)
                else:
                    break
            elif result == 'loss':
                if streak_type == 'loss' or streak_type is None:
                    current_streak += 1
                    streak_type = 'loss'
                    profit_streak += bet.get('profit', 0)
                else:
                    break
            elif result == 'push':
                continue
            else:
                break
        if current_streak >= 5:
            emoji = "🟢" if streak_type == 'win' else "🔴"
            title = "ПОБЕД" if streak_type == 'win' else "ПОРАЖЕНИЙ"
            self.send_if_needed(f'streak_{streak_type}_{current_streak}', f"{emoji} <b>СЕРИЯ {title}</b>\n{current_streak} подряд\nПрибыль: ${profit_streak:.2f}")
    
    def check_roi(self):
        stats = storage.load_stats()
        roi = stats.get('roi', 0)
        if roi > 20:
            self.send_if_needed('roi_high', f"📈 <b>ВЫСОКИЙ ROI</b>\nROI: {roi}%\nВсего ставок: {stats.get('total', 0)}\nПрибыль: ${stats.get('total_profit', 0):.2f}")
        elif roi < -10:
            self.send_if_needed('roi_low', f"📉 <b>НИЗКИЙ ROI</b>\nROI: {roi}%\nВсего ставок: {stats.get('total', 0)}\nУбыток: ${stats.get('total_profit', 0):.2f}")
    
    def run_all_checks(self):
        try:
            self.check_bank_status()
            self.check_streaks()
            self.check_roi()
        except Exception as e:
            logger.error(f"Ошибка в проверках: {e}")

notification_system = NotificationSystem()

# ============================================================
# УЛУЧШЕНИЕ 6: A/B ТЕСТИРОВАНИЕ
# ============================================================
class StrategyTester:
    def __init__(self):
        self.strategies = {
            '70_percent': {'name': '70%+ матчи', 'bets': [], 'profit': 0, 'wins': 0, 'losses': 0, 'total_stake': 0, 'active': True, 'start_date': datetime.now().isoformat()},
            'tm25_premium': {'name': 'ТМ 2.5 PREMIUM', 'bets': [], 'profit': 0, 'wins': 0, 'losses': 0, 'total_stake': 0, 'active': True, 'start_date': datetime.now().isoformat()},
            'tm25_standard': {'name': 'ТМ 2.5 STANDARD', 'bets': [], 'profit': 0, 'wins': 0, 'losses': 0, 'total_stake': 0, 'active': True, 'start_date': datetime.now().isoformat()}
        }
        self.current_test = None
    
    def add_bet(self, source, bet_data):
        if source in self.strategies and self.strategies[source]['active']:
            self.strategies[source]['bets'].append(bet_data)
            if bet_data.get('result') == 'win':
                self.strategies[source]['wins'] += 1
                self.strategies[source]['profit'] += bet_data.get('profit', 0)
            elif bet_data.get('result') == 'loss':
                self.strategies[source]['losses'] += 1
                self.strategies[source]['profit'] += bet_data.get('profit', 0)
            self.strategies[source]['total_stake'] += bet_data.get('stake', 0)
    
    def get_strategy_stats(self, strategy_name):
        if strategy_name not in self.strategies:
            return {}
        strategy = self.strategies[strategy_name]
        total_bets = len(strategy['bets'])
        total_wins = strategy['wins']
        total_losses = strategy['losses']
        winrate = (total_wins / total_bets * 100) if total_bets > 0 else 0
        roi = (strategy['profit'] / strategy['total_stake'] * 100) if strategy['total_stake'] > 0 else 0
        return {'total_bets': total_bets, 'wins': total_wins, 'losses': total_losses, 'winrate': round(winrate, 1), 'profit': round(strategy['profit'], 2), 'roi': round(roi, 1), 'total_stake': round(strategy['total_stake'], 2)}
    
    def get_comparison_report(self):
        report = "📊 <b>СРАВНЕНИЕ СТРАТЕГИЙ</b>\n\n"
        sorted_strategies = sorted(self.strategies.items(), key=lambda x: x[1]['profit'], reverse=True)
        for i, (name, data) in enumerate(sorted_strategies, 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else "  "
            total_bets = len(data['bets'])
            wins = data['wins']
            losses = data['losses']
            winrate = (wins / total_bets * 100) if total_bets > 0 else 0
            report += f"{emoji} <b>{data['name']}</b>\n"
            report += f"   Ставок: {total_bets}\n"
            report += f"   Винрейт: {winrate:.1f}%\n"
            report += f"   Прибыль: ${data['profit']:.2f}\n"
            report += f"   ROI: {(data['profit'] / data['total_stake'] * 100) if data['total_stake'] > 0 else 0:.1f}%\n\n"
        return report

strategy_tester = StrategyTester()

# ============================================================
# УЛУЧШЕНИЕ 7: СОСТОЯНИЕ БОТА
# ============================================================
class BotState:
    def __init__(self):
        self.state_file = 'bot_state.json'
        self.backup_dir = 'state_backups'
        self.state = self.load_state()
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)
    
    def load_state(self):
        default_state = {
            'start_time': datetime.now().isoformat(),
            'search_running': False,
            'bets_today': 0,
            'last_update': None,
            'last_full_search': None,
            'stats': {'total_processed': 0, 'total_found': 0, 'total_bets': 0, 'last_bet_time': None},
            'errors': [],
            'warnings': [],
            'version': '1.0.0'
        }
        try:
            if os.path.exists(self.state_file):
                with open(self.state_file, 'r', encoding='utf-8') as f:
                    state = json.load(f)
                for key, value in default_state.items():
                    if key not in state:
                        state[key] = value
                return state
            return default_state
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки состояния: {e}")
            return default_state
    
    def save_state(self):
        try:
            self.state['last_save'] = datetime.now().isoformat()
            self.create_backup()
            with open(self.state_file, 'w', encoding='utf-8') as f:
                json.dump(self.state, f, indent=2, default=str)
            logger.info("💾 Состояние сохранено")
        except Exception as e:
            logger.error(f"❌ Ошибка сохранения состояния: {e}")
    
    def create_backup(self):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = f"{self.backup_dir}/state_{timestamp}.json"
        if os.path.exists(self.state_file):
            import shutil
            shutil.copy2(self.state_file, backup_file)
            self.clean_old_backups()
    
    def clean_old_backups(self, keep=10):
        try:
            backups = sorted([f for f in os.listdir(self.backup_dir) if f.startswith('state_')])
            if len(backups) > keep:
                for old_backup in backups[:-keep]:
                    os.remove(os.path.join(self.backup_dir, old_backup))
        except Exception as e:
            logger.error(f"Ошибка очистки бэкапов: {e}")
    
    def update(self, **kwargs):
        for key, value in kwargs.items():
            if key in self.state:
                self.state[key] = value
        self.state['last_update'] = datetime.now().isoformat()
        self.save_state()
    
    def add_error(self, error_message):
        error_entry = {'time': datetime.now().isoformat(), 'message': error_message[:200]}
        self.state['errors'].append(error_entry)
        if len(self.state['errors']) > 100:
            self.state['errors'] = self.state['errors'][-100:]
        self.save_state()
    
    def get_status_report(self):
        report = f"🤖 <b>СТАТУС БОТА</b>\n\n"
        report += f"🕐 Время работы: {self.get_uptime()}\n"
        report += f"📊 Обработано матчей: {self.state['stats']['total_processed']}\n"
        report += f"🎯 Найдено ставок: {self.state['stats']['total_found']}\n"
        report += f"💰 Всего ставок: {self.state['stats']['total_bets']}\n"
        report += f"📅 Последний поиск: {self.state.get('last_full_search', 'Никогда')}\n"
        report += f"🔍 Поиск активен: {'Да' if self.state.get('search_running') else 'Нет'}\n"
        bank = storage.load_bank()
        report += f"💰 Банк: ${bank:.2f}\n"
        return report
    
    def get_uptime(self):
        start_time = datetime.fromisoformat(self.state['start_time'])
        uptime = datetime.now() - start_time
        hours = uptime.total_seconds() / 3600
        if hours < 24:
            return f"{hours:.1f} часов"
        else:
            days = hours // 24
            hours_remain = hours % 24
            return f"{int(days)} дней {hours_remain:.1f} часов"

bot_state = BotState()

# ============================================================
# УЛУЧШЕНИЕ 8: СРАВНЕНИЕ КОМАНД (НОВОЕ!)
# ============================================================
def get_team_comparison(home_team, away_team, league_id, fixture_id):
    """Сравнение команд по 12 параметрам (НЕ ПОКАЗЫВАЕТСЯ ПОЛЬЗОВАТЕЛЮ)"""
    try:
        home_name = home_team.get('name') if isinstance(home_team, dict) else str(home_team)
        away_name = away_team.get('name') if isinstance(away_team, dict) else str(away_team)
        home_id = home_team.get('id') if isinstance(home_team, dict) else None
        away_id = away_team.get('id') if isinstance(away_team, dict) else None
        
        standings = football_api.get_standings(league_id) if league_id else None
        home_stats = football_api.get_form(home_id) if home_id else None
        away_stats = football_api.get_form(away_id) if away_id else None
        h2h = football_api.get_head_to_head(home_name, away_name)
        injuries_home = football_api.get_injuries(home_id) if home_id else []
        injuries_away = football_api.get_injuries(away_id) if away_id else []
        
        home_xg = 1.2
        away_xg = 1.0
        if fixture_id:
            stats = football_api.get_match_statistics(fixture_id)
            if stats:
                for team_name, team_stats in stats.items():
                    if home_name.lower() in team_name.lower():
                        home_xg = team_stats.get('xG', 1.2)
                    elif away_name.lower() in team_name.lower():
                        away_xg = team_stats.get('xG', 1.0)
        
        home_position = 99
        away_position = 99
        home_points = 0
        away_points = 0
        home_goals_for = 0
        home_goals_against = 0
        away_goals_for = 0
        away_goals_against = 0
        
        if standings:
            if home_name in standings:
                home_position = standings[home_name].get('position', 99)
                home_points = standings[home_name].get('points', 0)
                home_goals_for = standings[home_name].get('goals_for', 0)
                home_goals_against = standings[home_name].get('goals_against', 0)
            if away_name in standings:
                away_position = standings[away_name].get('position', 99)
                away_points = standings[away_name].get('points', 0)
                away_goals_for = standings[away_name].get('goals_for', 0)
                away_goals_against = standings[away_name].get('goals_against', 0)
        
        home_home_wins = 0
        home_home_losses = 0
        away_away_wins = 0
        away_away_losses = 0
        
        if home_stats:
            home_home_wins = home_stats.get('home_wins', 0)
            home_home_losses = home_stats.get('home_losses', 0)
        if away_stats:
            away_away_wins = away_stats.get('away_wins', 0)
            away_away_losses = away_stats.get('away_losses', 0)
        
        home_form = home_stats.get('form', '') if home_stats else ''
        away_form = away_stats.get('form', '') if away_stats else ''
        home_goals_avg = home_stats.get('goals_avg', 1.2) if home_stats else 1.2
        away_goals_avg = away_stats.get('goals_avg', 1.0) if away_stats else 1.0
        home_conceded_avg = home_stats.get('conceded_avg', 1.0) if home_stats else 1.0
        away_conceded_avg = away_stats.get('conceded_avg', 1.2) if away_stats else 1.2
        
        home_motivation = get_motivation(home_position)
        away_motivation = get_motivation(away_position)
        
        h2h_wins = 0
        h2h_losses = 0
        h2h_draws = 0
        if h2h and h2h.get('total_matches', 0) > 0:
            h2h_wins = h2h.get('home_wins', 0)
            h2h_losses = h2h.get('away_wins', 0)
            h2h_draws = h2h.get('draws', 0)
        
        home_injuries_count = len(injuries_home) if injuries_home else 0
        away_injuries_count = len(injuries_away) if injuries_away else 0
        
        home_win_streak = 0
        home_loss_streak = 0
        away_win_streak = 0
        away_loss_streak = 0
        
        if home_form:
            for char in home_form:
                if char == 'W':
                    home_win_streak += 1
                elif char == 'L':
                    home_loss_streak += 1
                else:
                    break
        if away_form:
            for char in away_form:
                if char == 'W':
                    away_win_streak += 1
                elif char == 'L':
                    away_loss_streak += 1
                else:
                    break
        
        home_matches_in_week = 0
        away_matches_in_week = 0
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            home_fixtures = football_api.get_matches(league_id, today) if league_id else []
            if home_fixtures:
                home_matches_in_week = len([m for m in home_fixtures if home_name in str(m)])
        except:
            pass
        
        home_odds = 0
        draw_odds = 0
        away_odds = 0
        if fixture_id:
            odds_data = football_api.get_match_odds(fixture_id)
            if odds_data:
                home_odds = odds_data.get('home_odds', 0)
                draw_odds = odds_data.get('draw_odds', 0)
                away_odds = odds_data.get('away_odds', 0)
        
        adjustment = 0
        home_advantage = 0
        
        if home_xg > away_xg + 0.5:
            adjustment += 0.05
            home_advantage += 0.03
        elif home_xg < away_xg - 0.5:
            adjustment -= 0.05
            home_advantage -= 0.03
        
        home_form_wins = home_form.count('W') if home_form else 0
        away_form_wins = away_form.count('W') if away_form else 0
        
        if home_form_wins >= 4:
            adjustment += 0.05
            home_advantage += 0.03
        elif home_form_wins <= 1:
            adjustment -= 0.03
        if away_form_wins >= 4:
            adjustment -= 0.05
            home_advantage -= 0.03
        elif away_form_wins <= 1:
            adjustment += 0.03
            home_advantage += 0.02
        
        if home_motivation == 'champions_league':
            adjustment += 0.05
            home_advantage += 0.05
        elif home_motivation == 'relegation':
            adjustment += 0.03
            home_advantage += 0.03
        elif home_motivation == 'mid_table':
            adjustment -= 0.02
        if away_motivation == 'champions_league':
            adjustment -= 0.05
            home_advantage -= 0.05
        elif away_motivation == 'relegation':
            adjustment -= 0.03
            home_advantage -= 0.03
        elif away_motivation == 'mid_table':
            adjustment += 0.02
        
        if h2h_wins > h2h_losses + 1:
            adjustment += 0.03
            home_advantage += 0.03
        elif h2h_wins < h2h_losses - 1:
            adjustment -= 0.03
            home_advantage -= 0.03
        
        if home_home_wins > home_home_losses:
            adjustment += 0.03
            home_advantage += 0.03
        if away_away_wins > away_away_losses:
            adjustment -= 0.03
            home_advantage -= 0.03
        
        if home_win_streak >= 3:
            adjustment += 0.03
            home_advantage += 0.03
        if home_loss_streak >= 3:
            adjustment -= 0.03
            home_advantage -= 0.03
        if away_win_streak >= 3:
            adjustment -= 0.03
            home_advantage -= 0.03
        if away_loss_streak >= 3:
            adjustment += 0.03
            home_advantage += 0.03
        
        if home_injuries_count >= 3:
            adjustment -= 0.05
            home_advantage -= 0.05
        if away_injuries_count >= 3:
            adjustment += 0.05
            home_advantage += 0.05
        
        if home_position < away_position - 5:
            adjustment += 0.05
            home_advantage += 0.05
        elif home_position > away_position + 5:
            adjustment -= 0.05
            home_advantage -= 0.05
        
        home_gd = home_goals_for - home_goals_against
        away_gd = away_goals_for - away_goals_against
        if home_gd > away_gd + 5:
            adjustment += 0.03
            home_advantage += 0.03
        elif home_gd < away_gd - 5:
            adjustment -= 0.03
            home_advantage -= 0.03
        
        if home_matches_in_week >= 3:
            adjustment -= 0.03
            home_advantage -= 0.03
        if away_matches_in_week >= 3:
            adjustment += 0.03
            home_advantage += 0.03
        
        if home_odds > 0 and away_odds > 0:
            if home_odds < away_odds:
                adjustment += 0.03
                home_advantage += 0.03
            elif home_odds > away_odds:
                adjustment -= 0.03
                home_advantage -= 0.03
        
        adjustment = max(-0.20, min(0.20, adjustment))
        home_advantage = max(-0.15, min(0.15, home_advantage))
        
        return {
            'home_name': home_name,
            'away_name': away_name,
            'home_xg': home_xg,
            'away_xg': away_xg,
            'home_position': home_position,
            'away_position': away_position,
            'home_points': home_points,
            'away_points': away_points,
            'home_form': home_form,
            'away_form': away_form,
            'home_goals_avg': home_goals_avg,
            'away_goals_avg': away_goals_avg,
            'home_conceded_avg': home_conceded_avg,
            'away_conceded_avg': away_conceded_avg,
            'home_motivation': home_motivation,
            'away_motivation': away_motivation,
            'h2h_wins': h2h_wins,
            'h2h_losses': h2h_losses,
            'h2h_draws': h2h_draws,
            'home_injuries_count': home_injuries_count,
            'away_injuries_count': away_injuries_count,
            'home_win_streak': home_win_streak,
            'home_loss_streak': home_loss_streak,
            'away_win_streak': away_win_streak,
            'away_loss_streak': away_loss_streak,
            'home_goals_for': home_goals_for,
            'home_goals_against': home_goals_against,
            'away_goals_for': away_goals_for,
            'away_goals_against': away_goals_against,
            'home_matches_in_week': home_matches_in_week,
            'away_matches_in_week': away_matches_in_week,
            'home_odds': home_odds,
            'draw_odds': draw_odds,
            'away_odds': away_odds,
            'home_home_wins': home_home_wins,
            'home_home_losses': home_home_losses,
            'away_away_wins': away_away_wins,
            'away_away_losses': away_away_losses,
            'adjustment': adjustment,
            'home_advantage': home_advantage,
            'total_adjustment': adjustment + home_advantage
        }
    except Exception as e:
        logger.error(f"Ошибка сравнения команд: {e}")
        return None

# ============================================================
# ЗАГРУЗКА НАСТРОЕК
# ============================================================
def load_bot_settings():
    try:
        settings_file = 'bot_settings.json'
        if os.path.exists(settings_file):
            with open(settings_file, 'r') as f:
                settings = json.load(f)
            Config.EV_MIN_70 = settings.get('ev_min_70', getattr(Config, 'EV_MIN_70', 20))
            Config.PROB_MIN_70 = settings.get('prob_min_70', getattr(Config, 'PROB_MIN_70', 60))
            Config.XG_MIN_70 = settings.get('xg_min_70', getattr(Config, 'XG_MIN_70', 1.8))
            Config.XG_MAX_70 = settings.get('xg_max_70', getattr(Config, 'XG_MAX_70', 3.0))
            Config.POSITION_MAX_70 = settings.get('position_max_70', getattr(Config, 'POSITION_MAX_70', 15))
            Config.PREMIUM_MIN_EV = settings.get('premium_ev', getattr(Config, 'PREMIUM_MIN_EV', 30))
            Config.STANDARD_MIN_EV = settings.get('standard_ev', getattr(Config, 'STANDARD_MIN_EV', 15))
            Config.TM25_XG_MIN = settings.get('xg_min_tm25', getattr(Config, 'TM25_XG_MIN', 1.0))
            Config.TM25_XG_MAX = settings.get('xg_max_tm25', getattr(Config, 'TM25_XG_MAX', 3.0))
            Config.MAX_TM25_BETS = settings.get('max_tm25_bets', getattr(Config, 'MAX_TM25_BETS', 5))
            Config.TM25_TOP_LEAGUE_EV = settings.get('top_league_ev', getattr(Config, 'TM25_TOP_LEAGUE_EV', 35))
            logger.info("✅ Настройки загружены из bot_settings.json")
            return True
        else:
            logger.info("ℹ️ Файл bot_settings.json не найден, используются стандартные настройки")
            return False
    except Exception as e:
        logger.error(f"❌ Ошибка загрузки настроек: {e}")
        return False

# ============================================================
# FLASK WEBHOOK (ГЛАВНЫЙ)
# ============================================================
@app.route('/webhook', methods=['POST'])
def webhook():
    global search_running, search_state
    
    try:
        data = request.get_json()
        if not data:
            return "ok", 200
        
        logger.info("=" * 50)
        logger.info(f"📨 ПОЛУЧЕН ЗАПРОС ОТ TELEGRAM")
        logger.info("=" * 50)
        
        if 'callback_query' in data:
            callback = data['callback_query']
            callback_data = callback.get('data', '')
            logger.info(f"📨 Нажата кнопка: {callback_data}")
            answer_url = f"https://api.telegram.org/bot{Config.TELEGRAM_TOKEN}/answerCallbackQuery"
            try:
                requests.post(answer_url, json={
                    "callback_query_id": callback.get('id', ''),
                    "text": "✅ Результат сохранён!"
                })
            except Exception as e:
                logger.error(f"Ошибка ответа: {e}")
            return "ok", 200
        
        if 'message' not in data:
            return "ok", 200
        
        message = data['message']
        text = message.get('text', '')
        chat_id = message.get('chat', {}).get('id')
        
        if str(chat_id) != str(Config.ADMIN_CHAT_ID):
            logger.warning(f"⛔ ДОСТУП ЗАПРЕЩЕН для {chat_id}")
            send_telegram("⛔ Нет доступа")
            return "ok", 200
        
        if text == '/start':
            send_telegram(handlers.handle_start())
        elif text == '/help':
            send_telegram(handlers.handle_help())
        elif text == '/update':
            if search_running:
                if 'start_time' in search_state:
                    elapsed = (datetime.now() - search_state['start_time']).seconds
                    if elapsed > 900:  # Таймаут увеличен до 15 минут
                        search_running = False
                        search_state = {}
                        send_telegram("⏰ Поиск был принудительно сброшен (таймаут 15 мин)")
                    else:
                        send_telegram(f"⚠️ Поиск уже запущен! Идет {elapsed} секунд.")
                        return "ok", 200
                else:
                    send_telegram("⚠️ Поиск уже запущен!")
                    return "ok", 200
            else:
                search_running = True
                search_state = {'start_time': datetime.now()}
                
                # МГНОВЕННЫЙ ОТВЕТ (чтобы Telegram не отключил бота)
                send_telegram("🔎 Запущен полный анализ ВСЕХ лиг. Это может занять 10-20 минут. Я пришлю результат, когда закончу.")

                def run_full_search():
                    global search_running, search_state
                    try:
                        start_time = datetime.now()
                        matches = get_matches_with_factors()
                        
                        if matches:
                            send_telegram(f"📊 Найдено {len(matches)} матчей. Анализирую каждую игру...")
                            top_matches = find_top_matches_with_tm25(matches)
                            
                            if top_matches:
                                elapsed = (datetime.now() - start_time).seconds
                                minutes = elapsed // 60
                                seconds = elapsed % 60
                                
                                # Формируем сообщение
                                matches_text = ""
                                for i, m in enumerate(top_matches[:10], 1):
                                    best = m['best_bet']
                                    matches_text += f"{i}. <b>{m['home']} vs {m['away']}</b>\n"
                                    matches_text += f"   🏆 {m['league']}\n"
                                    matches_text += f"   🎯 {best['label']} | КЭФ: {best['odds']}\n"
                                    matches_text += f"   📈 EV: <b>{best['ev']}%</b> | Prob: {best['prob']}%\n"
                                    matches_text += f"   ⚽ XG: {m['total_xg']:.2f}\n\n"
                                
                                send_telegram(
                                    f"✅ <b>ПОИСК ЗАВЕРШЕН!</b>\n"
                                    f"🎯 Кандидатов: {len(top_matches)}\n"
                                    f"⏱️ Время: {minutes} мин {seconds} сек.\n\n"
                                    f"📋 <b>СПИСОК СТАВОК:</b>\n\n{matches_text}"
                                )
                            else:
                                send_telegram("❌ Ставок не найдено (70%+ и ТМ2.5).")
                        else:
                            send_telegram("❌ Матчей не найдено на сегодня.")
                    except Exception as e:
                        logger.error(f"❌ Ошибка фонового поиска: {e}", exc_info=True)
                        send_error_to_telegram(f"Ошибка фонового поиска: {e}")
                    finally:
                        search_running = False
                        search_state = {}
                
                # Запускаем в фоновом потоке
                t = Thread(target=run_full_search)
                t.daemon = True
                t.start()
                
            
        elif text == '/reset_search':
            search_running = False
            search_state = {}
            send_telegram("✅ Поиск сброшен! Теперь можно запускать заново.")
        elif text == '/stats':
            stats = storage.load_stats()
            bank = storage.load_bank()
            msg = f"📊 <b>СТАТИСТИКА</b>\n\n"
            msg += f"💰 Банк: ${bank:.2f}\n"
            msg += f"📊 Всего ставок: {stats.get('total', 0)}\n"
            msg += f"✅ Побед: {stats.get('wins', 0)}\n"
            msg += f"❌ Поражений: {stats.get('losses', 0)}\n"
            msg += f"🤝 Возвратов: {stats.get('pushes', 0)}\n"
            msg += f"📈 Винрейт: {stats.get('winrate', 0)}%\n"
            msg += f"💰 Прибыль: ${stats.get('total_profit', 0):.2f}\n"
            msg += f"📊 ROI: {stats.get('roi', 0)}%\n"
            if strategy_tester:
                msg += f"\n{strategy_tester.get_comparison_report()}"
            send_telegram(msg)
        elif text == '/bank':
            bank = storage.load_bank()
            send_telegram(f"💰 <b>ТЕКУЩИЙ БАНК</b>\n\n${bank:.2f}")
        elif text == '/export':
            file, message = export_to_excel()
            if file:
                send_telegram(message)
                url = f"https://api.telegram.org/bot{Config.TELEGRAM_TOKEN}/sendDocument"
                files = {'document': ('history.xlsx', file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {'chat_id': Config.ADMIN_CHAT_ID, 'caption': '📊 История ставок'}
                try:
                    requests.post(url, files=files, data=data, timeout=30)
                except Exception as e:
                    logger.error(f"Ошибка отправки файла: {e}")
            else:
                send_telegram(message)
        elif text == '/autobet':
            auto_bet.enabled = not auto_bet.enabled
            send_telegram(f"🤖 AutoBet: {'ВКЛЮЧЕН' if auto_bet.enabled else 'ВЫКЛЮЧЕН'}")
        elif text == '/update_results':
            send_telegram("🔄 Проверка результатов матчей...")
            updated = update_pending_bets()
            if updated > 0:
                send_telegram(f"✅ Обновлено {updated} результатов!")
            else:
                send_telegram("📭 Нет завершённых матчей для обновления")
        elif text == '/status':
            send_telegram(bot_state.get_status_report())
        elif text == '/strategies':
            send_telegram(strategy_tester.get_comparison_report())
        elif text.startswith('/result'):
            parts = text.replace('/result', '').strip()
            if ' vs ' in parts:
                match_part = parts.split(' vs ')
                if len(match_part) == 2:
                    match_and_score = match_part[1].split(' ')
                    if len(match_and_score) >= 2:
                        away = match_and_score[0]
                        score = match_and_score[1] if len(match_and_score) > 1 else ''
                        home = match_part[0].strip()
                        match = f"{home} vs {away}"
                        send_telegram(f"🔄 Обновление результата: {match} {score}")
                        result = update_manual_result(match, score)
                        send_telegram(result)
                    else:
                        send_telegram("⚠️ Используй: /result Aris Thessalonikis vs OFI 2-1")
                else:
                    send_telegram("⚠️ Используй: /result Aris Thessalonikis vs OFI 2-1")
            else:
                send_telegram("⚠️ Используй: /result Aris Thessalonikis vs OFI 2-1")
        elif text.startswith('/analyze'):
            match_name = text.replace('/analyze', '').strip()
            if match_name:
                send_telegram("🔍 Анализирую матч...")
                result = analyze_match(match_name)
                send_telegram(result)
            else:
                send_telegram("⚠️ Используй: /analyze Aris Thessalonikis vs OFI")
        elif text == '/stop':
            search_running = False
            search_state = {}
            send_telegram("⏹️ Поиск остановлен")
        else:
            send_telegram("❌ Неизвестная команда. /help")
        
        return "ok", 200
    except Exception as e:
        error_msg = f"Webhook error: {e}"
        logger.error(f"❌ {error_msg}")
        send_error_to_telegram(error_msg)
        return "ok", 200

# ============================================================
# API ЭНДПОИНТЫ
# ============================================================
@app.route('/api/stats', methods=['GET'])
def api_stats():
    stats = storage.load_stats()
    bank = storage.load_bank()
    return jsonify({'bank': bank, **stats})

@app.route('/api/history', methods=['GET'])
def api_history():
    history = storage.load_history()
    return jsonify(history)

@app.route('/api/matches', methods=['GET'])
def api_matches():
    cache = storage.load_cache()
    return jsonify(cache.get('top_matches', []))

@app.route('/api/all_data', methods=['GET'])
def all_data():
    try:
        logger.info("📡 Запрос всех данных для веб-приложения")
        stats = storage.load_stats()
        bank = storage.load_bank()
        history = storage.load_history()
        cache = storage.load_cache()
        profit_data = get_profit_data(history)
        result = {
            'stats': {
                'bank': bank,
                'total_bets': stats.get('total', 0),
                'wins': stats.get('wins', 0),
                'losses': stats.get('losses', 0),
                'profit': stats.get('total_profit', 0),
                'winrate': stats.get('winrate', 0),
                'roi': stats.get('roi', 0),
                'avg_stake': stats.get('avg_stake', 0)
            },
            'history': history,
            'profit_data': profit_data,
            'matches': cache.get('top_matches', [])
        }
        logger.info(f"✅ Данные отправлены: {len(history)} ставок, {len(result['matches'])} матчей")
        return jsonify(result)
    except Exception as e:
        logger.error(f"❌ Ошибка в /api/all_data: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/import_excel', methods=['POST'])
def import_excel():
    try:
        data = request.json
        excel_data = data.get('data', [])
        if not excel_data:
            return jsonify({'error': 'Нет данных'}), 400
        history = storage.load_history()
        imported = 0
        for row in excel_data:
            match = row.get('Матч', '') or row.get('Match', '')
            home = ''
            away = ''
            if ' vs ' in match:
                parts = match.split(' vs ')
                home = parts[0].strip()
                away = parts[1].strip()
            elif ' - ' in match:
                parts = match.split(' - ')
                home = parts[0].strip()
                away = parts[1].strip()
            score = row.get('Счёт', '') or row.get('Score', '')
            home_goals = None
            away_goals = None
            if score and '-' in str(score):
                parts = str(score).split('-')
                try:
                    home_goals = int(parts[0].strip())
                    away_goals = int(parts[1].strip())
                except:
                    pass
            bet = row.get('Ставка', '') or row.get('Bet', '')
            odds = float(row.get('Коэф', 1.85))
            stake = float(row.get('Сумма', 0))
            ev = float(row.get('EV%', 0))
            result = row.get('Результат', 'pending')
            profit = float(row.get('Прибыль', 0))
            date = row.get('Дата', '') or datetime.now().strftime('%Y-%m-%d %H:%M')
            bookmaker = row.get('Букмекер', '—')
            bet_record = {
                'home': home or 'Unknown',
                'away': away or 'Unknown',
                'league': 'Импорт из Excel',
                'bet': bet,
                'odds': odds,
                'stake': stake,
                'ev': ev,
                'result': result,
                'profit': profit,
                'date': date,
                'home_goals': home_goals,
                'away_goals': away_goals,
                'bookmaker': bookmaker
            }
            history.append(bet_record)
            imported += 1
        storage.save_history(history)
        recalc_stats()
        return jsonify({'success': True, 'count': imported})
    except Exception as e:
        logger.error(f"Ошибка импорта Excel: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/import_project', methods=['POST'])
def import_project():
    try:
        data = request.json
        history = data.get('history', [])
        stats = data.get('stats', {})
        if not history:
            return jsonify({'error': 'Нет данных для импорта'}), 400
        current_history = storage.load_history()
        existing_keys = set()
        for bet in current_history:
            key = f"{bet.get('date', '')}_{bet.get('home', '')}_{bet.get('away', '')}"
            existing_keys.add(key)
        imported = 0
        for bet in history:
            key = f"{bet.get('date', '')}_{bet.get('home', '')}_{bet.get('away', '')}"
            if key not in existing_keys:
                current_history.append(bet)
                imported += 1
                existing_keys.add(key)
        if stats and 'bank' in stats:
            storage.save_bank(stats['bank'])
        storage.save_history(current_history)
        recalc_stats()
        return jsonify({'success': True, 'count': imported})
    except Exception as e:
        logger.error(f"Ошибка импорта проекта: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/edit_bet', methods=['POST'])
def edit_bet():
    try:
        data = request.json
        index = data.get('index')
        history = storage.load_history()
        if index >= len(history):
            return jsonify({'error': 'Ставка не найдена'}), 404
        history[index]['home'] = data.get('home', history[index]['home'])
        history[index]['away'] = data.get('away', history[index]['away'])
        history[index]['home_goals'] = data.get('home_goals')
        history[index]['away_goals'] = data.get('away_goals')
        history[index]['bet'] = data.get('bet', history[index]['bet'])
        history[index]['odds'] = data.get('odds', history[index]['odds'])
        history[index]['stake'] = data.get('stake', history[index]['stake'])
        history[index]['ev'] = data.get('ev', history[index]['ev'])
        history[index]['result'] = data.get('result', history[index]['result'])
        history[index]['bookmaker'] = data.get('bookmaker', history[index].get('bookmaker', '—'))
        if history[index]['result'] == 'win':
            history[index]['profit'] = round(history[index]['stake'] * (history[index]['odds'] - 1), 2)
        elif history[index]['result'] == 'loss':
            history[index]['profit'] = -history[index]['stake']
        else:
            history[index]['profit'] = 0
        storage.save_history(history)
        recalc_stats()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete_bet', methods=['POST'])
def delete_bet():
    try:
        data = request.json
        index = data.get('index')
        history = storage.load_history()
        if index >= len(history):
            return jsonify({'error': 'Ставка не найдена'}), 404
        history.pop(index)
        storage.save_history(history)
        recalc_stats()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/bank', methods=['POST'])
def update_bank():
    try:
        data = request.json
        if 'bank' in data:
            storage.save_bank(data['bank'])
            return jsonify({'success': True, 'bank': data['bank']})
        return jsonify({'error': 'No bank value'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/simulate', methods=['POST'])
def simulate():
    try:
        data = request.json
        count = data.get('count', 1000)
        history = storage.load_history()
        if len(history) < 5:
            return jsonify({'error': 'Нужно минимум 5 ставок для симуляции'}), 400
        wins = sum(1 for b in history if b.get('result') == 'win')
        total = len(history)
        winrate = wins / total if total > 0 else 0
        avg_stake = sum(float(b.get('stake', 0)) for b in history) / total if total > 0 else 10
        results = []
        profit_history = []
        total_profit = 0
        for i in range(count):
            if random.random() < winrate:
                profit = avg_stake * random.uniform(0.5, 1.5)
                total_profit += profit
                results.append('win')
            else:
                profit = -avg_stake
                total_profit += profit
                results.append('loss')
            profit_history.append(round(total_profit, 2))
        wins_sim = results.count('win')
        losses_sim = results.count('loss')
        max_profit = max(profit_history) if profit_history else 0
        min_profit = min(profit_history) if profit_history else 0
        return jsonify({
            'total': count,
            'wins': wins_sim,
            'losses': losses_sim,
            'profit': round(total_profit, 2),
            'winrate': round(wins_sim / count * 100, 1),
            'roi': round((total_profit / (avg_stake * count)) * 100, 2) if avg_stake > 0 else 0,
            'risk': round((abs(min_profit) / (avg_stake * count)) * 100, 2) if avg_stake > 0 else 0,
            'max_profit': round(max_profit, 2),
            'min_profit': round(min_profit, 2),
            'avg_stake': round(avg_stake, 2),
            'history': profit_history[:100],
            'labels': list(range(1, min(count, 100) + 1))
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/add_manual_match', methods=['POST'])
def add_manual_match():
    try:
        data = request.json
        match_name = data.get('match', '')
        score = data.get('score', '-')
        result = data.get('result', 'win')
        stake = data.get('stake', 0)
        bet_type = data.get('bet', '')
        odds = data.get('odds', 1.85)
        bookmaker = data.get('bookmaker', 'Ручное добавление')
        if not match_name:
            return jsonify({'error': 'Название матча обязательно'}), 400
        home_goals = None
        away_goals = None
        if score and '-' in score:
            parts = score.split('-')
            try:
                home_goals = int(parts[0].strip())
                away_goals = int(parts[1].strip())
            except:
                pass
        home = 'Unknown'
        away = 'Unknown'
        if ' vs ' in match_name:
            parts = match_name.split(' vs ')
            home = parts[0].strip()
            away = parts[1].strip()
        elif ' - ' in match_name:
            parts = match_name.split(' - ')
            home = parts[0].strip()
            away = parts[1].strip()
        if result == 'win':
            profit = round(stake * (odds - 1), 2)
        elif result == 'loss':
            profit = -stake
        else:
            profit = 0
        history = storage.load_history()
        bet_record = {
            'home': home or 'Unknown',
            'away': away or 'Unknown',
            'league': 'Ручное добавление',
            'bet': bet_type,
            'odds': odds,
            'stake': stake,
            'ev': 0,
            'result': result,
            'profit': profit,
            'date': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'home_goals': home_goals,
            'away_goals': away_goals,
            'manual': True,
            'bookmaker': bookmaker
        }
        history.append(bet_record)
        storage.save_history(history)
        recalc_stats()
        return jsonify({'success': True, 'count': 1})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/update_settings', methods=['POST'])
def update_settings():
    try:
        data = request.json
        settings_file = 'bot_settings.json'
        with open(settings_file, 'w') as f:
            json.dump(data, f, indent=2)
        Config.EV_MIN_70 = data.get('ev_min_70', getattr(Config, 'EV_MIN_70', 20))
        Config.PROB_MIN_70 = data.get('prob_min_70', getattr(Config, 'PROB_MIN_70', 60))
        Config.XG_MIN_70 = data.get('xg_min_70', getattr(Config, 'XG_MIN_70', 1.8))
        Config.XG_MAX_70 = data.get('xg_max_70', getattr(Config, 'XG_MAX_70', 3.0))
        Config.POSITION_MAX_70 = data.get('position_max_70', getattr(Config, 'POSITION_MAX_70', 15))
        Config.PREMIUM_MIN_EV = data.get('premium_ev', getattr(Config, 'PREMIUM_MIN_EV', 30))
        Config.STANDARD_MIN_EV = data.get('standard_ev', getattr(Config, 'STANDARD_MIN_EV', 15))
        Config.TM25_XG_MIN = data.get('xg_min_tm25', getattr(Config, 'TM25_XG_MIN', 1.0))
        Config.TM25_XG_MAX = data.get('xg_max_tm25', getattr(Config, 'TM25_XG_MAX', 3.0))
        Config.MAX_TM25_BETS = data.get('max_tm25_bets', getattr(Config, 'MAX_TM25_BETS', 5))
        Config.TM25_TOP_LEAGUE_EV = data.get('top_league_ev', getattr(Config, 'TM25_TOP_LEAGUE_EV', 35))
        logger.info("✅ Настройки обновлены через API")
        return jsonify({'success': True, 'message': 'Настройки обновлены'})
    except Exception as e:
        logger.error(f"Ошибка обновления настроек: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/strategies', methods=['GET'])
def api_strategies():
    return jsonify({
        'strategies': strategy_tester.strategies,
        'current_test': strategy_tester.current_test
    })

@app.route('/health', methods=['GET'])
def health():
    return {"status": "ok", "time": datetime.now().isoformat()}

@app.route('/', methods=['GET'])
def index():
    return f"🤖 Quantum Bot PRO (70%+ Target + ТМ 2.5 Special) | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

# ============================================================
# ЗАПУСК
# ============================================================
if __name__ == "__main__":
    setup_logging()
    load_bot_settings()
    start_scheduler()
    schedule_updates()
    schedule_notifications()
    schedule_performance_report()
    
    port = int(os.environ.get("PORT", 10000))
    logger.info("🚀 БОТ ЗАПУЩЕН (70%+ TARGET + ТМ 2.5 SPECIAL)!")
    logger.info("📊 Сканируется {} лиг".format(len(Config.LEAGUES)))
    logger.info("🤖 Максимум ставок: {}".format(Config.MAX_BETS_PER_RUN))
    logger.info("🎯 ФИЛЬТРЫ ДЛЯ 70%+:")
    logger.info("   - EV > {}%".format(getattr(Config, 'EV_MIN_70', 20)))
    logger.info("   - Prob > {}%".format(getattr(Config, 'PROB_MIN_70', 60)))
    logger.info("   - XG {}-{}".format(getattr(Config, 'XG_MIN_70', 1.8), getattr(Config, 'XG_MAX_70', 3.0)))
    logger.info("   - Форма excellent/good")
    logger.info("   - Мотивация (не середняки)")
    logger.info("   - Лимит 3 ставки на тип")
    logger.info("   - Лимит 2 ставки на лигу")
    logger.info("🎯 ФИЛЬТРЫ ДЛЯ ТМ 2.5 (ДВУХУРОВНЕВЫЙ):")
    logger.info("   PREMIUM: EV > {}%".format(getattr(Config, 'PREMIUM_MIN_EV', 30)))
    logger.info("   STANDARD: EV > {}%".format(getattr(Config, 'STANDARD_MIN_EV', 15)))
    logger.info("   - Лимит {} ставки".format(getattr(Config, 'MAX_TM25_BETS', 5)))
    logger.info("🎯 КОЭФФИЦИЕНТЫ (ДВУХЭТАПНЫЙ ПОИСК):")
    logger.info("   1. Odds API (топ-лиги)")
    logger.info("   2. Football API (все лиги)")
    logger.info("   3. Заглушка 1.95 (если не найдены)")
    logger.info("✅ Команды: /update_results, /result, /analyze, /status, /strategies")
    logger.info("✅ Кэш матчей сохраняется")
    logger.info("⏰ Авто-обновление результатов: каждые 6 часов")
    logger.info("📊 A/B тестирование активно")
    logger.info("🔔 Уведомления активны")
    logger.info("📈 Мониторинг производительности активен")
    app.run(host='0.0.0.0', port=port)
