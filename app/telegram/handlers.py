# app/telegram/handlers.py
import requests
from datetime import datetime, timedelta
from app.database.storage import storage
from app.config import Config

class Handlers:
    def __init__(self):
        self.token = Config.TELEGRAM_TOKEN
        self.admin_chat_id = Config.ADMIN_CHAT_ID
    
    def send_message(self, text, parse_mode='HTML'):
        """Отправить сообщение в Telegram"""
        try:
            url = f"https://api.telegram.org/bot{self.token}/sendMessage"
            data = {
                'chat_id': self.admin_chat_id,
                'text': text,
                'parse_mode': parse_mode
            }
            response = requests.post(url, json=data, timeout=10)
            return response.json()
        except Exception as e:
            print(f"❌ Ошибка отправки: {e}")
            return None
    
    def handle_start(self):
        """/start - Приветствие и помощь"""
        return """🤖 <b>Quantum Bet Bot PRO v12</b>

Доступные команды:

/start - Приветствие и помощь
/update - Поиск матчей на сегодня 
/today - ТОП-5 матчей из кэша
/bank - Текущий банк
/stats - Общая статистика
/team <название> - Статистика по команде
/bettypes - Статистика по типам ставок
/timestats - Статистика по времени ставок
/mlstats - Статистика машинного обучения
/report - Еженедельный отчёт
/export - Экспорт истории в Excel
/autobet - Включить/выключить авто-ставки
/train - Обучить нейросеть
/arb - Поиск букмекерских вилок
/anomalies - Поиск аномалий
/security - Статистика безопасности
/unblock <IP> - Разблокировать IP
/result <команда1> <команда2> <счёт> - Ввести результат
/stop - Остановить поиск
/help - Помощь"""
    
    def handle_bank(self):
        """/bank - Текущий банк"""
        bank = storage.load_bank()
        return f"💰 <b>Текущий банк:</b> ${bank:.2f}"
    
    def handle_stats(self):
        """/stats - Общая статистика"""
        stats = storage.load_stats()
        history = storage.load_history()
        
        total = len(history)
        wins = stats.get('wins', 0)
        losses = stats.get('losses', 0)
        pushes = stats.get('pushes', 0)
        profit = stats.get('total_profit', 0)
        winrate = stats.get('winrate', 0)
        roi = stats.get('roi', 0)
        
        total_stake = sum(bet.get('stake', 0) for bet in history)
        avg_stake = round(total_stake / total, 2) if total > 0 else 0
        
        msg = f"""📊 <b>ОБЩАЯ СТАТИСТИКА</b>

📈 Всего ставок: {total}
✅ Выигрыши: {wins}
❌ Проигрыши: {losses}
🔄 Возвраты: {pushes}
💰 Прибыль: ${profit:.2f}
🎯 Проходимость: {winrate}%
📈 ROI: {roi}%
📅 Средняя ставка: ${avg_stake}"""
        return msg
    
    def handle_today(self):
        """/today - ТОП-5 матчей из кэша"""
        cache = storage.load_cache()
        matches = cache.get('top_matches', [])
        
        if not matches:
            return "📭 Нет матчей в кэше. Запусти /update"
        
        msg = "📊 <b>ТОП-5 МАТЧЕЙ</b>\n\n"
        for i, match in enumerate(matches[:5], 1):
            msg += f"{i}. 🏟️ {match.get('home')} vs {match.get('away')}\n"
            msg += f"   📊 Лига: {match.get('league')}\n"
            if match.get('bets'):
                best = match['bets'][0]
                msg += f"   🎯 {best.get('label')} | КЭФ: {best.get('odds')} | EV: {best.get('ev')}%\n"
            msg += "\n"
        return msg
    
    def handle_team(self in, team_name):
        """/team <название> - Статистика по команде"""
        if not team_name:
            return "⚠️ Укажи команду: /team Real Madrid"
        
        b.get history = storage.load_history()
        team_bets = [
            b for b in history 
            if team_name.lower() in b.get('home', '').lower() 
            or team_name.lower()('away', '').lower()
        ]
        
        if not team_bets:
            return f"📭 Нет ставок на {team_name}"
        
        wins = sum(1 for b in team_bets if b.get('result') == 'win')
        losses = sum(1 for b in team_bets if b.get('result') == 'loss')
        pushes = sum(1 for b in team_bets if b.get('result') == 'push')
        profit = sum(b.get('profit', 0) for b in team_bets)
        total_stake = sum(b.get('stake', 0) for b in team_bets)
        
        msg = f"""📊 <b>Статистика по {team_name}</b>

📈 Всего: {len(team_bets)}
✅ Выигрыши: {wins}
❌ Проигрыши: {losses}
🔄 Возвраты: {pushes}
💰 Прибыль: ${profit:.2f}
🎯 Проходимость: {round(wins / (wins + losses) * 100, 1) if (wins + losses) > 0 else 0}%
📈 ROI: {round(profit / total_stake * 100, 1) if total_stake > 0 else 0}%"""
        return msg
    
    def handle_bettypes(self):
        """/bettypes - Статистика по типам ставок"""
        history = storage.load_history()
        bet_stats = {}
        
        for bet in history:
            btype = bet.get('bet', 'Unknown')
            if btype not in bet_stats:
                bet_stats[btype] = {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}
            bet_stats[btype]['total'] += 1
            if bet.get('result') == 'win':
                bet_stats[btype]['wins'] += 1
            elif bet.get('result') == 'loss':
                bet_stats[btype]['losses'] += 1
            bet_stats[btype]['profit'] += bet.get('profit', 0)
        
        if not bet_stats:
            return "📭 Нет данных"
        
        msg = "📊 <b>Статистика по типам ставок</b>\n\n"
        for btype, stats in sorted(bet_stats.items(), key=lambda x: x[1]['profit'], reverse=True)[:10]:
            total = stats['total']
            wins = stats['wins']
            winrate = round(wins / total * 100, 1) if total > 0 else 0
            msg += f"🎯 {btype}: {winrate}% ({wins}/{total}) | ${stats['profit']:.2f}\n"
        return msg
    
    def handle_timestats(self):
        """/timestats - Статистика по времени ставок"""
        history = storage.load_history()
        hour_stats = {}
        
        for bet in history:
            try:
                bet_date = datetime.strptime(bet.get('date', ''), '%Y-%m-%d %H:%M')
                hour = bet_date.hour
                if hour not in hour_stats:
                    hour_stats[hour] = {'total': 0, 'wins': 0, 'profit': 0}
                hour_stats[hour]['total'] += 1
                if bet.get('result') == 'win':
                    hour_stats[hour]['wins'] += 1
                hour_stats[hour]['profit'] += bet.get('profit', 0)
            except:
                pass
        
        if not hour_stats:
            return "📭 Нет данных"
        
        msg = "📊 <b>Статистика по времени ставок</b>\n\n"
        for hour, stats in sorted(hour_stats.items()):
            total = stats['total']
            wins = stats['wins']
            winrate = round(wins / total * 100, 1) if total > 0 else 0
            msg += f"🕐 {hour:02d}:00 - {winrate}% ({wins}/{total}) | ${stats['profit']:.2f}\n"
        return msg
    
    def handle_mlstats(self):
        """/mlstats - Статистика машинного обучения"""
        try:
            from app.ml.predictor import ml_predictor
            stats = ml_predictor.get_stats()
            
            msg = f"""🤖 <b>Статистика ML</b>

📊 Всего предсказаний: {stats.get('total', 0)}
✅ Точных: {stats.get('correct', 0)}
❌ Неверных: {stats.get('wrong', 0)}
🎯 Точность: {stats.get('accuracy', 0)}%
📈 Последнее обновление: {stats.get('last_update', 'Нет')}"""
            return msg
        except:
            return "⚠️ ML модуль отключен или не настроен"
    
    def handle_report(self):
        """/report - Еженедельный отчёт"""
        history = storage.load_history()
        week_ago = datetime.now() - timedelta(days=7)
        week_bets = []
        
        for bet in history:
            try:
                bet_date = datetime.strptime(bet.get('date', ''), '%Y-%m-%d %H:%M')
                if bet_date >= week_ago:
                    week_bets.append(bet)
            except:
                pass
        
        if not week_bets:
            return "📭 Нет ставок за последнюю неделю"
        
        wins = sum(1 for b in week_bets if b.get('result') == 'win')
        losses = sum(1 for b in week_bets if b.get('result') == 'loss')
        profit = sum(b.get('profit', 0) for b in week_bets)
        total_stake = sum(b.get('stake', 0) for b in week_bets)
        
        msg = f"""📊 <b>ЕЖЕНЕДЕЛЬНЫЙ ОТЧЁТ</b>
📅 За последние 7 дней

📈 Ставок: {len(week_bets)}
✅ Выигрышей: {wins}
❌ Проигрышей: {losses}
💰 Прибыль: ${profit:.2f}
📈 ROI: {round(profit / total_stake * 100, 1) if total_stake > 0 else 0}%"""
        return msg
    
    def handle_export(self):
        """/export - Экспорт истории в Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        
        history = storage.load_history()
        
        if not history:
            return "📭 Нет данных для экспорта"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ставки"
        
        headers = ["Дата", "Матч", "Счёт", "Ставка", "Коэф", "EV%", "Сумма", "Результат", "Прибыль"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for bet in history:
            date = bet.get('date', '')
            home = bet.get('home', '')
            away = bet.get('away', '')
            home_goals = bet.get('home_goals', '')
            away_goals = bet.get('away_goals', '')
            score = f"{home_goals}-{away_goals}" if home_goals is not None and away_goals is not None else "-"
            bet_type = bet.get('bet', '')
            odds = bet.get('odds', 0)
            ev = bet.get('ev', 0)
            stake = bet.get('stake', 0)
            result = bet.get('result', 'pending')
            profit = bet.get('profit', 0)
            
            ws.append([date, f"{home} vs {away}", score, bet_type, odds, ev, stake, result, profit])
        
        for col in range(1, len(headers) + 1):
            column_letter = chr(64 + col)
            ws.column_dimensions[column_letter].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Отправляем файл
        url = f"https://api.telegram.org/bot{self.token}/sendDocument"
        files = {'document': ('history.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'chat_id': self.admin_chat_id, 'caption': '📊 История ставок'}
        
        try:
            response = requests.post(url, files=files, data=data, timeout=30)
            return "✅ Экспорт отправлен!"
        except Exception as e:
            return f"❌ Ошибка отправки: {e}"
    
    def handle_autobet(self, enabled):
        """/autobet - Включить/выключить авто-ставки"""
        status = "✅ ВКЛЮЧЕНЫ" if enabled else "❌ ВЫКЛЮЧЕНЫ"
        return f"🤖 Авто-ставки {status}!"
    
    def handle_stop(self):
        """/stop - Остановить поиск"""
        return "🛑 ПОИСК ОСТАНОВЛЕН!"
    
    def handle_help(self):
        """/help - Помощь"""
        return self.handle_start()
    
    def handle_train(self):
        """/train - Обучить нейросеть"""
        try:
            from app.ml.predictor import ml_predictor
            result = ml_predictor.train()
            return f"🧠 {result}"
        except Exception as e:
            return f"❌ Ошибка обучения: {e}"
    
    def handle_arb(self):
        """/arb - Поиск букмекерских вилок"""
        try:
            from app.analytics.arbitrage import arbitrage_analyzer
            arbs = arbitrage_analyzer.find_arbitrage()
            if arbs:
                msg = "🔄 <b>БУКМЕКЕРСКИЕ ВИЛКИ</b>\n\n"
                for arb in arbs[:5]:
                    msg += f"🏟️ {arb.get('home')} vs {arb.get('away')}\n"
                    msg += f"   🎯 {arb.get('type')} | Прибыль: {arb.get('profit')}%\n\n"
                return msg
            return "📭 Вилок не найдено"
        except:
            return "⚠️ Модуль поиска вилок отключен"
    
    def handle_anomalies(self):
        """/anomalies - Поиск аномалий в коэффициентах"""
        try:
            from app.analytics.anomalies import anomaly_detector
            anomalies = anomaly_detector.get_anomalies()
            if anomalies:
                msg = "⚠️ <b>ОБНАРУЖЕНЫ АНОМАЛИИ</b>\n\n"
                for a in anomalies[:5]:
                    msg += f"🏟️ {a.get('match', 'Unknown')}\n"
                    msg += f"   📊 {a.get('type')} - {a.get('reason')}\n\n"
                return msg
            return "✅ Аномалий не обнаружено"
        except:
            return "⚠️ Модуль обнаружения аномалий отключен"
    
    def handle_security(self):
        """/security - Статистика безопасности"""
        try:
            from app.analytics.anomalies import anomaly_detector
            anomalies = anomaly_detector.get_anomalies() if hasattr(anomaly_detector, 'get_anomalies') else []
            
            msg = f"""🔒 <b>СТАТИСТИКА БЕЗОПАСНОСТИ</b>

🛡️ Блокировок IP: {len(anomalies) if anomalies else 0}
⚠️ Аномалий обнаружено: {len(anomalies) if anomalies else 0}"""
            
            if anomalies:
                msg += "\n\n📋 Последние:\n"
                for a in anomalies[-5:]:
                    msg += f"• {a.get('ip', 'Unknown')} - {a.get('reason', '')}\n"
            return msg
        except:
            return "🔒 Статистика безопасности недоступна"
    
    def handle_unblock(self, ip):
        """/unblock <IP> - Разблокировать IP"""
        if not ip:
            return "⚠️ Укажи IP: /unblock 192.168.1.1"
        
        try:
            from app.analytics.anomalies import anomaly_detector
            result = anomaly_detector.unblock_ip(ip) if hasattr(anomaly_detector, 'unblock_ip') else None
            
            if result:
                return f"✅ IP {ip} разблокирован!"
            return f"❌ IP {ip} не найден в блок-листе"
        except:
            return "⚠️ Не удалось разблокировать IP"
    
    def handle_result(self, match, score):
        """/result <команда1> <команда2> <счёт> - Ввести результат вручную"""
        if not match or not score:
            return "⚠️ Используй: /result Fulham Chelsea 2-1"
        
        try:
            parts = match.split(' vs ')
            if len(parts) != 2:
                return "⚠️ Формат: /result Fulham vs Chelsea 2-1"
            
            home, away = parts[0].strip(), parts[1].strip()
            score_parts = score.split('-')
            home_goals = int(score_parts[0].strip())
            away_goals = int(score_parts[1].strip())
            
            # Находим матч в истории
            history = storage.load_history()
            updated = 0
            
            for bet in history:
                if bet.get('result') == 'pending' and bet.get('home', '').lower() == home.lower() and bet.get('away', '').lower() == away.lower():
                    bet['home_goals'] = home_goals
                    bet['away_goals'] = away_goals
                    
                    # Определяем результат
                    bet_type = bet.get('bet', '').lower()
                    if 'п1' in bet_type or 'победа хозяев' in bet_type:
                        if home_goals > away_goals:
                            bet['result'] = 'win'
                            bet['profit'] = round(bet.get('stake', 0) * (bet.get('odds', 1) - 1), 2)
                        elif home_goals == away_goals:
                            bet['result'] = 'push'
                            bet['profit'] = 0
                        else:
                            bet['result'] = 'loss'
                            bet['profit'] = -bet.get('stake', 0)
                    elif 'п2' in bet_type or 'победа гостей' in bet_type:
                        if away_goals > home_goals:
                            bet['result'] = 'win'
                            bet['profit'] = round(bet.get('stake', 0) * (bet.get('odds', 1) - 1), 2)
                        elif home_goals == away_goals:
                            bet['result'] = 'push'
                            bet['profit'] = 0
                        else:
                            bet['result'] = 'loss'
                            bet['profit'] = -bet.get('stake', 0)
                    elif '1x' in bet_type:
                        if home_goals >= away_goals:
                            bet['result'] = 'win'
                            bet['profit'] = round(bet.get('stake', 0) * (bet.get('odds', 1) - 1), 2)
                        else:
                            bet['result'] = 'loss'
                            bet['profit'] = -bet.get('stake', 0)
                    elif 'x2' in bet_type:
                        if away_goals >= home_goals:
                            bet['result'] = 'win'
                            bet['profit'] = round(bet.get('stake', 0) * (bet.get('odds', 1) - 1), 2)
                        else:
                            bet['result'] = 'loss'
                            bet['profit'] = -bet.get('stake', 0)
                    else:
                        bet['result'] = 'pending'
                    
                    updated += 1
            
            if updated > 0:
                storage.save_history(history)
                return f"✅ Обновлен результат для {home} vs {away} ({home_goals}-{away_goals})\nОбновлено ставок: {updated}"
            else:
                return f"📭 Не найдено ожидающих ставок на {home} vs {away}"
                
        except Exception as e:
            return f"❌ Ошибка: {e}"


# Создаем экземпляр для использования
handlers = Handlers()
