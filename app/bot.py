import sys
import os
import time
import json

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from flask import Flask, request, jsonify
from datetime import datetime, timedelta

from app.config import Config
from app.database.storage import storage
from app.api.football import football_api
from app.api.weather import weather_api
from app.analytics.xg import xg_analyzer
from app.analytics.probability import (
    calculate_probabilities,
    calculate_ev,
    get_bet_types
)
from app.analytics.arbitrage import arbitrage_analyzer
from app.analytics.anomalies import anomaly_detector
from app.telegram.handlers import handlers
from app.utils.logger import setup_logging, get_logger
from app.scheduler import start_scheduler


# ============================================================
# ОСНОВНЫЕ НАСТРОЙКИ
# ============================================================

logger = get_logger(__name__)
app = Flask(__name__)

search_running = False

# UTC +3
TIMEZONE_OFFSET = 3


# ============================================================
# AUTOBET
# ============================================================

auto_bet = None


def get_auto_bet():
    """
    Загружает AutoBet только при первом использовании.
    Это предотвращает циклические импорты.
    """
    global auto_bet

    if auto_bet is None:
        try:
            from app.betting.auto_bet import AutoBet

            auto_bet = AutoBet()

            logger.info("✅ AutoBet загружен")

        except Exception as e:
            logger.exception("❌ Не удалось загрузить AutoBet")

            send_error_to_telegram(
                f"Не удалось загрузить AutoBet:\n{e}"
            )

            auto_bet = None
            return None

    return auto_bet


# ============================================================
# TELEGRAM
# ============================================================

def send_error_to_telegram(error_text: str):
    """
    Отправляет ошибку администратору Telegram.
    """
    try:
        import requests

        url = (
            f"https://api.telegram.org/"
            f"bot{Config.TELEGRAM_TOKEN}/sendMessage"
        )

        if len(error_text) > 4000:
            error_text = error_text[:4000] + "...(обрезано)"

        data = {
            "chat_id": Config.ADMIN_CHAT_ID,
            "text": f"❌ <b>ОШИБКА БОТА</b>\n\n{error_text}",
            "parse_mode": "HTML"
        }

        response = requests.post(
            url,
            json=data,
            timeout=5
        )

        if response.status_code != 200:
            logger.error(
                f"❌ Telegram error response: {response.text}"
            )

    except Exception as e:
        logger.error(
            f"❌ Не удалось отправить ошибку в Telegram: {e}"
        )


def send_telegram(
    text: str,
    parse_mode: str = "HTML"
):
    """
    Отправляет сообщение администратору Telegram.
    """
    try:
        import requests

        url = (
            f"https://api.telegram.org/"
            f"bot{Config.TELEGRAM_TOKEN}/sendMessage"
        )

        data = {
            "chat_id": Config.ADMIN_CHAT_ID,
            "text": text,
            "parse_mode": parse_mode
        }

        response = requests.post(
            url,
            json=data,
            timeout=10
        )

        if response.status_code != 200:
            logger.error(
                f"❌ Ошибка отправки Telegram: {response.text}"
            )

    except Exception as e:
        logger.error(
            f"❌ Send Telegram error: {e}"
        )


# ============================================================
# ЭКСПОРТ В EXCEL
# ============================================================

def export_to_excel():

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    try:
        history = storage.load_history()

        if not history:
            return None, "📭 Нет данных для экспорта"

        wb = Workbook()
        ws = wb.active
        ws.title = "Ставки"

        headers = [
            "Дата",
            "Матч",
            "Счёт",
            "Ставка",
            "Коэф",
            "EV%",
            "Сумма",
            "Результат",
            "Прибыль"
        ]

        ws.append(headers)

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )

        for col in range(1, len(headers) + 1):

            cell = ws.cell(
                row=1,
                column=col
            )

            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center"
            )

        total_profit = 0

        for bet in history:

            date = bet.get("date", "")

            home = bet.get("home", "")
            away = bet.get("away", "")

            home_goals = bet.get(
                "home_goals",
                None
            )

            away_goals = bet.get(
                "away_goals",
                None
            )

            if (
                home_goals is not None
                and away_goals is not None
            ):
                score = f"{home_goals}-{away_goals}"
            else:
                score = "-"

            bet_type = bet.get(
                "bet",
                ""
            )

            try:
                odds = float(
                    bet.get("odds", 0)
                )
            except (
                ValueError,
                TypeError
            ):
                odds = 0

            try:
                ev = float(
                    bet.get("ev", 0)
                )
            except (
                ValueError,
                TypeError
            ):
                ev = 0

            try:
                stake = float(
                    bet.get("stake", 0)
                )
            except (
                ValueError,
                TypeError
            ):
                stake = 0

            result = bet.get(
                "result",
                "pending"
            )

            try:
                profit = float(
                    bet.get("profit", 0)
                )
            except (
                ValueError,
                TypeError
            ):
                profit = 0

            if result == "win":

                if profit == 0:
                    profit = round(
                        stake * (odds - 1),
                        2
                    )

                total_profit += profit

            elif result == "loss":

                if profit == 0:
                    profit = -round(
                        stake,
                        2
                    )

                total_profit += profit

            else:
                profit = 0

            ws.append([
                date,
                f"{home} vs {away}",
                score,
                bet_type,
                odds,
                ev,
                stake,
                result,
                profit
            ])

        ws.append([])

        ws.append([
            "ИТОГО",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            round(total_profit, 2)
        ])

        widths = {
            "A": 20,
            "B": 35,
            "C": 10,
            "D": 20,
            "E": 10,
            "F": 10,
            "G": 12,
            "H": 12,
            "I": 12
        }

        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        output = io.BytesIO()

        wb.save(output)

        output.seek(0)

        return (
            output,
            f"✅ Экспорт завершен! "
            f"Всего ставок: {len(history)}, "
            f"Прибыль: ${round(total_profit, 2)}"
        )

    except Exception as e:

        logger.exception(
            "❌ Ошибка экспорта"
        )

        return (
            None,
            f"❌ Ошибка экспорта: {e}"
        )


# ============================================================
# ПОИСК МАТЧЕЙ
# ============================================================

def get_matches_with_factors():

    all_matches = []

    today = datetime.now().strftime(
        "%Y-%m-%d"
    )

    dates_to_search = [today]

    logger.info(
        f"🔍 Поиск матчей на: {today}"
    )

    all_leagues = (
        list(Config.LEAGUES)
        + list(
            getattr(
                Config,
                "CUP_LEAGUES",
                []
            )
        )
    )

    logger.info(
        f"📊 Всего соревнований: "
        f"{len(all_leagues)}"
    )

    for league_id in all_leagues:

        for search_date in dates_to_search:

            league_name = Config.LEAGUE_NAMES.get(
                league_id,
                str(league_id)
            )

            try:

                matches = football_api.get_matches(
                    league_id,
                    search_date
                )

                if (
                    not matches
                    or not isinstance(matches, list)
                ):

                    logger.info(
                        f"🔥 Нет матчей в "
                        f"{league_name} "
                        f"на {search_date}"
                    )

                    continue

                for match in matches:

                    if not isinstance(match, dict):
                        continue

                    fixture = match.get(
                        "fixture"
                    )

                    if not isinstance(
                        fixture,
                        dict
                    ):
                        continue

                    status = fixture.get(
                        "status",
                        {}
                    )

                    if not isinstance(
                        status,
                        dict
                    ):
                        continue

                    # Только матчи, которые ещё не начались
                    if status.get("short") != "NS":
                        continue

                    match_id = fixture.get("id")

                    if not match_id:
                        continue

                    # Проверяем дубликаты
                    existing_ids = set()

                    for m in all_matches:

                        if not isinstance(
                            m,
                            dict
                        ):
                            continue

                        existing_fixture = m.get(
                            "fixture",
                            {}
                        )

                        if isinstance(
                            existing_fixture,
                            dict
                        ):

                            existing_id = (
                                existing_fixture.get("id")
                            )

                            if existing_id:
                                existing_ids.add(
                                    existing_id
                                )

                    if match_id in existing_ids:
                        continue

                    teams = match.get(
                        "teams",
                        {}
                    )

                    if not isinstance(
                        teams,
                        dict
                    ):
                        continue

                    home_team = teams.get(
                        "home",
                        {}
                    )

                    away_team = teams.get(
                        "away",
                        {}
                    )

                    if not isinstance(
                        home_team,
                        dict
                    ):
                        continue

                    if not isinstance(
                        away_team,
                        dict
                    ):
                        continue

                    home_id = home_team.get("id")
                    away_id = away_team.get("id")

                    if not home_id or not away_id:
                        continue

                    # Получаем факторы
                    home_form = []
                    away_form = []

                    home_injuries = []
                    away_injuries = []

                    try:
                        home_form = football_api.get_form(
                            home_id
                        )
                    except Exception as e:
                        logger.warning(
                            f"⚠️ Ошибка формы хозяев "
                            f"{home_id}: {e}"
                        )

                    try:
                        away_form = football_api.get_form(
                            away_id
                        )
                    except Exception as e:
                        logger.warning(
                            f"⚠️ Ошибка формы гостей "
                            f"{away_id}: {e}"
                        )

                    try:
                        home_injuries = (
                            football_api.get_injuries(
                                home_id
                            )
                        )
                    except Exception as e:
                        logger.warning(
                            f"⚠️ Ошибка травм хозяев: {e}"
                        )

                    try:
                        away_injuries = (
                            football_api.get_injuries(
                                away_id
                            )
                        )
                    except Exception as e:
                        logger.warning(
                            f"⚠️ Ошибка травм гостей: {e}"
                        )

                    match["factors"] = {

                        "home_form": home_form,

                        "away_form": away_form,

                        "home_injuries_list":
                            home_injuries
                            if isinstance(
                                home_injuries,
                                list
                            )
                            else [],

                        "away_injuries_list":
                            away_injuries
                            if isinstance(
                                away_injuries,
                                list
                            )
                            else [],

                        "home_id": home_id,

                        "away_id": away_id,

                        "referee": fixture.get(
                            "referee"
                        )
                    }

                    # Погода отключена
                    match["weather"] = None

                    match["weather_reason"] = (
                        "🌤️ Погода отключена"
                    )

                    league_data = match.get(
                        "league"
                    )

                    if isinstance(
                        league_data,
                        dict
                    ):
                        league_data["name"] = league_name

                    all_matches.append(match)

            except Exception as e:

                error_msg = (
                    f"Ошибка {league_name} "
                    f"на {search_date}: {e}"
                )

                logger.exception(
                    f"❌ {error_msg}"
                )

                send_error_to_telegram(
                    error_msg
                )

            time.sleep(0.1)

    logger.info(
        f"📊 ВСЕГО найдено матчей: "
        f"{len(all_matches)}"
    )

    return all_matches


# ============================================================
# ПАРСИНГ КОЭФФИЦИЕНТОВ
# ============================================================

def parse_odds(odds_data):
    """
    Безопасно превращает коэффициенты API
    в float.

    Исправляет ошибку:

    can't multiply sequence by non-int
    of type 'float'
    """

    if not odds_data:
        return None

    if not isinstance(
        odds_data,
        dict
    ):
        logger.warning(
            "⚠️ odds_data не является dict"
        )
        return None

    bookmakers = odds_data.get(
        "bookmakers",
        []
    )

    if not isinstance(
        bookmakers,
        list
    ):
        return None

    if not bookmakers:
        return None

    odds_dict = {}

    # Перебираем всех букмекеров,
    # а не только первого
    for bookmaker in bookmakers:

        if not isinstance(
            bookmaker,
            dict
        ):
            continue

        bets = bookmaker.get(
            "bets",
            []
        )

        if not isinstance(
            bets,
            list
        ):
            continue

        for bet in bets:

            if not isinstance(
                bet,
                dict
            ):
                continue

            values = bet.get(
                "values",
                []
            )

            if not isinstance(
                values,
                list
            ):
                continue

            for value in values:

                if not isinstance(
                    value,
                    dict
                ):
                    continue

                bet_type = value.get(
                    "value",
                    ""
                )

                odd = value.get(
                    "odd"
                )

                if not bet_type:
                    continue

                if odd is None:
                    continue

                # ====================================================
                # ГЛАВНОЕ ИСПРАВЛЕНИЕ
                # API может вернуть "1.85" вместо 1.85
                # ====================================================

                try:

                    if isinstance(
                        odd,
                        str
                    ):
                        odd = odd.strip()

                        # На случай 1,85
                        odd = odd.replace(
                            ",",
                            "."
                        )

                    odd = float(odd)

                except (
                    ValueError,
                    TypeError
                ):

                    logger.warning(
                        f"⚠️ Некорректный коэффициент: "
                        f"{bet_type} = {odd}"
                    )

                    continue

                if odd <= 1:
                    continue

                # Сохраняем только первый найденный
                # нормальный коэффициент
                if str(bet_type) not in odds_dict:

                    odds_dict[
                        str(bet_type)
                    ] = odd

    if not odds_dict:

        logger.warning(
            "⚠️ Коэффициенты не найдены"
        )

        return None

    return odds_dict


# ============================================================
# НОРМАЛИЗАЦИЯ ФОРМЫ
# ============================================================

def normalize_form(form):

    if form is None:
        return ""

    if isinstance(
        form,
        (list, tuple)
    ):

        result = ""

        for item in form:

            if isinstance(
                item,
                dict
            ):

                value = (
                    item.get("result")
                    or item.get("form")
                    or item.get("outcome")
                    or ""
                )

                result += str(value)

            else:

                result += str(item)

        return result.upper()

    return str(form).upper()


# ============================================================
# РАСЧЕТ XG
# ============================================================

def calculate_adjusted_xg(
    home_id,
    away_id,
    factors
):

    home_xg = 1.2
    away_xg = 1.0

    factors = (
        factors
        if isinstance(
            factors,
            dict
        )
        else {}
    )

    # --------------------------------------------------------
    # Форма
    # --------------------------------------------------------

    home_form = normalize_form(
        factors.get(
            "home_form",
            ""
        )
    )

    away_form = normalize_form(
        factors.get(
            "away_form",
            ""
        )
    )

    # Убираем всё кроме W/D/L
    home_form = "".join(
        x for x in home_form
        if x in "WDL"
    )

    away_form = "".join(
        x for x in away_form
        if x in "WDL"
    )

    # --------------------------------------------------------
    # Форма хозяев
    # --------------------------------------------------------

    if home_form:

        home_form_points = sum(
            3 if x == "W"
            else 1 if x == "D"
            else 0
            for x in home_form
        )

        denominator = (
            len(home_form) * 3
        )

        home_form_ratio = (
            home_form_points / denominator
            if denominator > 0
            else 0.5
        )

        form_multiplier = (
            0.8
            + home_form_ratio * 0.4
        )

        home_xg *= form_multiplier

        logger.info(
            f"   📊 Форма хозяев: "
            f"{home_form} "
            f"(коэф: {form_multiplier:.2f})"
        )

    # --------------------------------------------------------
    # Форма гостей
    # --------------------------------------------------------

    if away_form:

        away_form_points = sum(
            3 if x == "W"
            else 1 if x == "D"
            else 0
            for x in away_form
        )

        denominator = (
            len(away_form) * 3
        )

        away_form_ratio = (
            away_form_points / denominator
            if denominator > 0
            else 0.5
        )

        form_multiplier = (
            0.8
            + away_form_ratio * 0.4
        )

        away_xg *= form_multiplier

        logger.info(
            f"   📊 Форма гостей: "
            f"{away_form} "
            f"(коэф: {form_multiplier:.2f})"
        )

    # --------------------------------------------------------
    # ТРАВМЫ
    # --------------------------------------------------------

    home_injuries = factors.get(
        "home_injuries_list",
        []
    )

    away_injuries = factors.get(
        "away_injuries_list",
        []
    )

    if not isinstance(
        home_injuries,
        list
    ):
        home_injuries = []

    if not isinstance(
        away_injuries,
        list
    ):
        away_injuries = []

    if home_injuries:

        injury_penalty = min(
            len(home_injuries) * 0.05,
            0.30
        )

        home_xg *= (
            1 - injury_penalty
        )

        logger.info(
            f"   🏥 Травмы хозяев: "
            f"{len(home_injuries)} "
            f"(пенальти: "
            f"{injury_penalty * 100:.0f}%)"
        )

    if away_injuries:

        injury_penalty = min(
            len(away_injuries) * 0.05,
            0.30
        )

        away_xg *= (
            1 - injury_penalty
        )

        logger.info(
            f"   🏥 Травмы гостей: "
            f"{len(away_injuries)} "
            f"(пенальти: "
            f"{injury_penalty * 100:.0f}%)"
        )

    # --------------------------------------------------------
    # ДОМАШНЕЕ ПРЕИМУЩЕСТВО
    # --------------------------------------------------------

    home_xg *= 1.10
    away_xg *= 0.90

    logger.info(
        "   🏠 Домашнее преимущество: "
        "+10% / -10%"
    )

    return (
        float(home_xg),
        float(away_xg)
    )


# ============================================================
# БЕЗОПАСНЫЙ РАСЧЕТ EV
# ============================================================

def safe_calculate_ev(
    probability,
    odds
):
    """
    Дополнительная защита от неправильных типов.
    """

    try:

        probability = float(
            probability
        )

        odds = float(
            odds
        )

        if odds <= 1:
            return None

        if probability < 0:
            probability = 0

        if probability > 1:
            probability = 1

        ev = calculate_ev(
            probability,
            odds
        )

        return float(ev)

    except (
        ValueError,
        TypeError,
        ArithmeticError
    ) as e:

        logger.error(
            f"❌ Ошибка расчета EV: "
            f"probability={probability}, "
            f"odds={odds}, "
            f"error={e}"
        )

        return None


# ============================================================
# ТОП МАТЧЕЙ
# ============================================================

def find_top_matches(matches):

    all_matches_data = []

    bets_placed = 0

    max_bets = int(
        getattr(
            Config,
            "MAX_BETS_PER_RUN",
            5
        )
    )

    logger.info(
        f"🔍 Анализ {len(matches)} матчей..."
    )

    BET_TYPES = [

        {
            "type": "under",
            "label": "ТМ 2.5",
            "marker": 42.86875,
            "keys": [
                "Under 2.5",
                "Under",
                "U 2.5"
            ]
        },

        {
            "type": "btts",
            "label": "ОБЗ",
            "marker": 40.7253125,
            "keys": [
                "Both Team Score",
                "BTTS",
                "Both Teams to Score"
            ]
        },

        {
            "type": "1X",
            "label": "1X",
            "marker": 45.125,
            "keys": [
                "Home/Draw",
                "1X"
            ]
        },

        {
            "type": "over",
            "label": "ТБ 2.5",
            "marker": 41.375,
            "keys": [
                "Over 2.5",
                "Over",
                "O 2.5"
            ]
        },

        {
            "type": "X2",
            "label": "X2",
            "marker": 43.1875,
            "keys": [
                "Away/Draw",
                "X2"
            ]
        }
    ]

    if not isinstance(
        matches,
        list
    ):
        return []

    for match in matches:

        if bets_placed >= max_bets:

            logger.info(
                f"⚠️ Достигнут лимит ставок: "
                f"{max_bets}"
            )

            break

        if not isinstance(
            match,
            dict
        ):
            continue

        try:

            # ----------------------------------------------------
            # FIXTURE
            # ----------------------------------------------------

            fixture = match.get(
                "fixture"
            )

            if not isinstance(
                fixture,
                dict
            ):
                continue

            fixture_id = fixture.get(
                "id"
            )

            if not fixture_id:
                continue

            # ----------------------------------------------------
            # TEAMS
            # ----------------------------------------------------

            teams = match.get(
                "teams"
            )

            if not isinstance(
                teams,
                dict
            ):
                continue

            home_team = teams.get(
                "home"
            )

            away_team = teams.get(
                "away"
            )

            if not isinstance(
                home_team,
                dict
            ):
                continue

            if not isinstance(
                away_team,
                dict
            ):
                continue

            home = home_team.get(
                "name",
                "Unknown"
            )

            away = away_team.get(
                "name",
                "Unknown"
            )

            logger.info(
                f"📊 Анализ: "
                f"{home} vs {away} "
                f"(ID: {fixture_id})"
            )

            # ----------------------------------------------------
            # КОЭФФИЦИЕНТЫ
            # ----------------------------------------------------

            odds_data = football_api.get_match_odds(
                fixture_id
            )

            if not odds_data:

                logger.warning(
                    f"⚠️ Нет коэффициентов "
                    f"для {home} vs {away}"
                )

                continue

            odds_dict = parse_odds(
                odds_data
            )

            if not odds_dict:

                logger.warning(
                    f"⚠️ Не удалось распарсить "
                    f"коэффициенты "
                    f"{home} vs {away}"
                )

                continue

            logger.info(
                f"   💹 Получено коэффициентов: "
                f"{len(odds_dict)}"
            )

            # ----------------------------------------------------
            # ФАКТОРЫ
            # ----------------------------------------------------

            factors = match.get(
                "factors",
                {}
            )

            if not isinstance(
                factors,
                dict
            ):
                factors = {}

            home_id = factors.get(
                "home_id"
            )

            away_id = factors.get(
                "away_id"
            )

            # ----------------------------------------------------
            # XG
            # ----------------------------------------------------

            home_xg, away_xg = (
                calculate_adjusted_xg(
                    home_id,
                    away_id,
                    factors
                )
            )

            home_xg = float(
                home_xg
            )

            away_xg = float(
                away_xg
            )

            logger.info(
                f"   📈 Итоговый XG: "
                f"{home} {home_xg:.2f} - "
                f"{away_xg:.2f} {away}"
            )

            # ----------------------------------------------------
            # PROBABILITIES
            # ----------------------------------------------------

            try:

                probs = calculate_probabilities(
                    home_xg,
                    away_xg
                )

            except Exception as e:

                logger.exception(
                    "❌ Ошибка расчета вероятностей"
                )

                send_error_to_telegram(
                    f"Ошибка расчета вероятностей "
                    f"{home} vs {away}:\n{e}"
                )

                continue

            if not isinstance(
                probs,
                dict
            ):

                logger.warning(
                    "⚠️ calculate_probabilities "
                    "вернул не dict"
                )

                continue

            # ----------------------------------------------------
            # ЛИГА
            # ----------------------------------------------------

            league_data = match.get(
                "league"
            )

            if isinstance(
                league_data,
                dict
            ):

                league = league_data.get(
                    "name",
                    "Unknown"
                )

            else:

                league = "Unknown"

            # ----------------------------------------------------
            # ВРЕМЯ
            # ----------------------------------------------------

            match_time = fixture.get(
                "date",
                ""
            )

            if match_time:

                try:

                    dt = datetime.fromisoformat(
                        match_time.replace(
                            "Z",
                            "+00:00"
                        )
                    )

                    dt = (
                        dt
                        + timedelta(
                            hours=TIMEZONE_OFFSET
                        )
                    )

                    match_time = dt.strftime(
                        "%d.%m.%Y %H:%M"
                    )

                except Exception:

                    match_time = (
                        "Время не указано"
                    )

            else:

                match_time = (
                    "Время не указано"
                )

            # ----------------------------------------------------
            # ДАННЫЕ МАТЧА
            # ----------------------------------------------------

            match_data = {

                "home": home,

                "away": away,

                "league": league,

                "fixture_id": fixture_id,

                "match_time": match_time,

                "home_xg": round(
                    home_xg,
                    2
                ),

                "away_xg": round(
                    away_xg,
                    2
                ),

                "weather_reason": "🌤️",

                "factors": factors,

                "intuition": [],

                "bets": []
            }

            # ----------------------------------------------------
            # СТАВКИ
            # ----------------------------------------------------

            for bet_config in BET_TYPES:

                bet_type = bet_config[
                    "type"
                ]

                label = bet_config[
                    "label"
                ]

                marker = float(
                    bet_config[
                        "marker"
                    ]
                )

                keys = bet_config[
                    "keys"
                ]

                # --------------------------------------------
                # Ищем коэффициент
                # --------------------------------------------

                odds = None

                for key in keys:

                    if key in odds_dict:

                        odds = odds_dict[key]

                        break

                if odds is None:
                    continue

                # --------------------------------------------
                # ГЛАВНОЕ ИСПРАВЛЕНИЕ
                # --------------------------------------------

                try:

                    odds = float(
                        odds
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    logger.warning(
                        f"⚠️ Некорректный odds: "
                        f"{odds}"
                    )

                    continue

                if odds <= 1:

                    logger.warning(
                        f"⚠️ Коэффициент <= 1: "
                        f"{odds}"
                    )

                    continue

                # --------------------------------------------
                # Вероятность
                # --------------------------------------------

                try:

                    prob = float(
                        probs.get(
                            bet_type,
                            0.33
                        )
                    )

                except (
                    ValueError,
                    TypeError
                ):

                    prob = 0.33

                # Защита
                prob = max(
                    0.0,
                    min(
                        prob,
                        1.0
                    )
                )

                # --------------------------------------------
                # EV
                # --------------------------------------------

                ev = safe_calculate_ev(
                    prob,
                    odds
                )

                if ev is None:

                    logger.warning(
                        f"⚠️ EV не рассчитан: "
                        f"{home} vs {away} | "
                        f"{label}"
                    )

                    continue

                logger.info(
                    f"   📊 {label}: "
                    f"prob={prob:.4f}, "
                    f"odds={odds:.2f}, "
                    f"EV={ev:.2f}%"
                )

                # --------------------------------------------
                # Минимальный EV
                # --------------------------------------------

                if ev < 5:

                    continue

                # --------------------------------------------
                # Добавляем ставку
                # --------------------------------------------

                bet_data = {

                    "bet_type": bet_type,

                    "label": label,

                    "odds": float(
                        odds
                    ),

                    "prob": round(
                        prob * 100,
                        1
                    ),

                    "ev": round(
                        ev,
                        1
                    ),

                    "stake": round(
                        marker,
                        2
                    ),

                    "marker_stake": float(
                        marker
                    )
                }

                match_data[
                    "bets"
                ].append(
                    bet_data
                )

                logger.info(
                    f"   ✅ ДОБАВЛЕНА СТАВКА: "
                    f"{label} | "
                    f"КЭФ: {odds:.2f} | "
                    f"EV: {ev:.1f}%"
                )

            # ----------------------------------------------------
            # ЕСЛИ НАЙДЕНЫ СТАВКИ
            # ----------------------------------------------------

            if not match_data["bets"]:
                continue

            all_matches_data.append(
                match_data
            )

            # ----------------------------------------------------
            # AUTOBET
            # ----------------------------------------------------

            try:

                auto = get_auto_bet()

                if auto is None:

                    logger.error(
                        "❌ AutoBet не загружен"
                    )

                    continue

                # Проверяем включен ли AutoBet
                if not getattr(
                    auto,
                    "enabled",
                    True
                ):

                    logger.info(
                        "⏸️ AutoBet отключен"
                    )

                    continue

                bet_result = (
                    auto.check_and_bet(
                        match_data
                    )
                )

                if bet_result:

                    bets_placed += 1

                    msg = (
                        f"🤖 "
                        f"<b>АВТО-СТАВКА "
                        f"#{bets_placed}</b>\n"
                    )

                    msg += (
                        f"🏟️ "
                        f"{bet_result.get('match', "
                        f"home + ' vs ' + away)}\n"
                    )

                    if bet_result.get(
                        "match_time"
                    ):

                        msg += (
                            f"📅 "
                            f"{bet_result['match_time']}\n"
                        )

                    msg += (
                        f"📊 "
                        f"{bet_result.get('bet', label)} "
                        f"| КЭФ: "
                        f"{bet_result.get('odds', odds)}\n"
                    )

                    msg += (
                        f"💰 Сумма: "
                        f"${bet_result.get('stake', marker)}\n"
                    )

                    msg += (
                        f"📈 EV: "
                        f"{bet_result.get('ev', ev)}%"
                    )

                    if bet_result.get(
                        "marker_stake"
                    ):

                        msg += (
                            f"\n🎯 Маркер: "
                            f"${bet_result['marker_stake']}"
                        )

                    send_telegram(
                        msg
                    )

                    logger.info(
                        f"✅ АВТО-СТАВКА "
                        f"#{bets_placed}"
                    )

            except Exception as e:

                logger.exception(
                    "❌ Ошибка авто-ставки"
                )

                send_error_to_telegram(
                    f"Ошибка авто-ставки "
                    f"{home} vs {away}:\n{e}"
                )

        except Exception as e:

            logger.exception(
                f"❌ Ошибка анализа "
                f"{match.get('home', '')} "
                f"vs "
                f"{match.get('away', '')}"
            )

            # ВАЖНО:
            # ошибка одного матча не останавливает
            # весь поиск
            continue

    logger.info(
        f"📊 Найдено "
        f"{len(all_matches_data)} матчей, "
        f"сделано "
        f"{bets_placed} ставок"
    )

    # ------------------------------------------------------------
    # CACHE
    # ------------------------------------------------------------

    try:

        cache = storage.load_cache()

        if not isinstance(
            cache,
            dict
        ):
            cache = {}

        cache[
            "top_matches"
        ] = all_matches_data

        storage.save_cache(
            cache
        )

    except Exception as e:

        logger.error(
            f"❌ Ошибка сохранения cache: {e}"
        )

    return all_matches_data[:20]


# ============================================================
# ОПРЕДЕЛЕНИЕ РЕЗУЛЬТАТА СТАВКИ
# ============================================================

def determine_bet_result(
    bet_type,
    home_goals,
    away_goals
):

    try:

        home_goals = int(
            home_goals
        )

        away_goals = int(
            away_goals
        )

    except (
        ValueError,
        TypeError
    ):

        return "pending"

    total = (
        home_goals
        + away_goals
    )

    bet_type_lower = str(
        bet_type or ""
    ).lower().strip()

    # --------------------------------------------------------
    # BTTS
    # --------------------------------------------------------

    if (
        "оз - да"
        in bet_type_lower
        or "обз"
        in bet_type_lower
        or "btts"
        in bet_type_lower
        or "both teams"
        in bet_type_lower
    ):

        return (
            "win"
            if (
                home_goals > 0
                and away_goals > 0
            )
            else "loss"
        )

    # --------------------------------------------------------
    # UNDER 2.5
    # --------------------------------------------------------

    if (
        "тм 2.5"
        in bet_type_lower
        or "under 2.5"
        in bet_type_lower
        or "under"
        in bet_type_lower
    ):

        return (
            "win"
            if total < 2.5
            else "loss"
        )

    # --------------------------------------------------------
    # OVER 2.5
    # --------------------------------------------------------

    if (
        "тб 2.5"
        in bet_type_lower
        or "over 2.5"
        in bet_type_lower
        or "over"
        in bet_type_lower
    ):

        return (
            "win"
            if total > 2.5
            else "loss"
        )

    # --------------------------------------------------------
    # 1X
    # --------------------------------------------------------

    if (
        bet_type_lower == "1x"
        or "1x" in bet_type_lower
        or "1х" in bet_type_lower
    ):

        return (
            "win"
            if home_goals >= away_goals
            else "loss"
        )

    # --------------------------------------------------------
    # X2
    # --------------------------------------------------------

    if (
        bet_type_lower == "x2"
        or "x2" in bet_type_lower
        or "х2" in bet_type_lower
    ):

        return (
            "win"
            if away_goals >= home_goals
            else "loss"
        )

    # --------------------------------------------------------
    # ПОБЕДА ХОЗЯЕВ
    # --------------------------------------------------------

    if (
        "п1" in bet_type_lower
        or "победа хозяев"
        in bet_type_lower
    ):

        if home_goals > away_goals:
            return "win"

        if home_goals == away_goals:
            return "push"

        return "loss"

    # --------------------------------------------------------
    # ПОБЕДА ГОСТЕЙ
    # --------------------------------------------------------

    if (
        "п2" in bet_type_lower
        or "победа гостей"
        in bet_type_lower
    ):

        if away_goals > home_goals:
            return "win"

        if home_goals == away_goals:
            return "push"

        return "loss"

    return "pending"


# ============================================================
# ОБНОВЛЕНИЕ РЕЗУЛЬТАТОВ
# ============================================================

def update_pending_bets():

    history = storage.load_history()

    if not isinstance(
        history,
        list
    ):
        return 0

    updated = 0

    for bet in history:

        if not isinstance(
            bet,
            dict
        ):
            continue

        current_result = bet.get(
            "result"
        )

        if (
            current_result != "pending"
            and current_result is not None
        ):
            continue

        fixture_id = bet.get(
            "fixture_id"
        )

        # ----------------------------------------------------
        # Если fixture_id отсутствует,
        # ищем матч по названиям
        # ----------------------------------------------------

        if not fixture_id:

            home = bet.get(
                "home",
                ""
            )

            away = bet.get(
                "away",
                ""
            )

            if (
                home
                and away
                and home != "Unknown"
                and away != "Unknown"
            ):

                try:

                    fixture_id = (
                        football_api.find_fixture_by_teams(
                            home,
                            away
                        )
                    )

                    if fixture_id:

                        bet[
                            "fixture_id"
                        ] = fixture_id

                except Exception as e:

                    logger.warning(
                        f"⚠️ Не удалось найти "
                        f"fixture_id для "
                        f"{home} vs {away}: "
                        f"{e}"
                    )

        if not fixture_id:
            continue

        # ----------------------------------------------------
        # Получаем результат
        # ----------------------------------------------------

        try:

            match_data = (
                football_api.get_match_result(
                    fixture_id
                )
            )

        except Exception as e:

            logger.warning(
                f"⚠️ Ошибка получения "
                f"результата {fixture_id}: "
                f"{e}"
            )

            continue

        if not match_data:
            continue

        try:

            goals = match_data.get(
                "goals",
                {}
            )

            if not isinstance(
                goals,
                dict
            ):
                continue

            home_goals = goals.get(
                "home"
            )

            away_goals = goals.get(
                "away"
            )

        except Exception:
            continue

        if (
            home_goals is None
            or away_goals is None
        ):
            continue

        bet_type = bet.get(
            "bet",
            ""
        )

        result = determine_bet_result(
            bet_type,
            home_goals,
            away_goals
        )

        if result == "pending":
            continue

        # ----------------------------------------------------
        # Записываем результат
        # ----------------------------------------------------

        bet["result"] = result

        bet["home_goals"] = (
            home_goals
        )

        bet["away_goals"] = (
            away_goals
        )

        try:

            stake = float(
                bet.get(
                    "stake",
                    0
                )
            )

        except (
            ValueError,
            TypeError
        ):

            stake = 0.0

        try:

            odds = float(
                bet.get(
                    "odds",
                    1
                )
            )

        except (
            ValueError,
            TypeError
        ):

            odds = 1.0

        if result == "win":

            bet["profit"] = round(
                stake * (odds - 1),
                2
            )

        elif result == "loss":

            bet["profit"] = -round(
                stake,
                2
            )

        elif result == "push":

            bet["profit"] = 0

        updated += 1

        logger.info(
            f"✅ Обновлена ставка: "
            f"{bet.get('home')} vs "
            f"{bet.get('away')} → "
            f"{home_goals}-{away_goals} → "
            f"{result}"
        )

    # --------------------------------------------------------
    # Сохраняем
    # --------------------------------------------------------

    if updated > 0:

        storage.save_history(
            history
        )

        recalc_stats()

        send_telegram(
            f"✅ Автоматически обновлено "
            f"{updated} результатов!"
        )

    return updated


# ============================================================
# ПЕРЕСЧЕТ СТАТИСТИКИ
# ============================================================

def recalc_stats():

    history = storage.load_history()

    if not isinstance(
        history,
        list
    ):
        history = []

    stats = storage.load_stats()

    if not isinstance(
        stats,
        dict
    ):
        stats = {}

    total = len(
        history
    )

    wins = sum(
        1
        for b in history
        if b.get("result") == "win"
    )

    losses = sum(
        1
        for b in history
        if b.get("result") == "loss"
    )

    pushes = sum(
        1
        for b in history
        if b.get("result") == "push"
    )

    total_profit = 0.0

    total_stake = 0.0

    for b in history:

        try:

            total_profit += float(
                b.get(
                    "profit",
                    0
                )
            )

        except (
            ValueError,
            TypeError
        ):
            pass

        try:

            total_stake += float(
                b.get(
                    "stake",
                    0
                )
            )

        except (
            ValueError,
            TypeError
        ):
            pass

    stats["total"] = total

    stats["wins"] = wins

    stats["losses"] = losses

    stats["pushes"] = pushes

    stats["total_profit"] = round(
        total_profit,
        2
    )

    stats["winrate"] = round(
        (
            wins
            / (wins + losses)
            * 100
        ),
        1
    ) if (
        wins + losses
    ) > 0 else 0

    stats["roi"] = round(
        (
            total_profit
            / total_stake
            * 100
        ),
        1
    ) if total_stake > 0 else 0

    storage.save_stats(
        stats
    )

    logger.info(
        f"📊 Статистика пересчитана: "
        f"{stats}"
    )


# ============================================================
# WEBHOOK
# ============================================================

@app.route(
    "/webhook",
    methods=["POST"]
)
def webhook():

    global search_running

    try:

        data = request.get_json(
            silent=True
        )

        if not data:

            return "ok", 200

        logger.info(
            "=" * 50
        )

        logger.info(
            "📨 ПОЛУЧЕН ЗАПРОС ОТ TELEGRAM"
        )

        logger.info(
            "=" * 50
        )

        # ====================================================
        # CALLBACK QUERY
        # ====================================================

        if "callback_query" in data:

            callback = data[
                "callback_query"
            ]

            callback_data = callback.get(
                "data",
                ""
            )

            logger.info(
                f"📨 Нажата кнопка: "
                f"{callback_data}"
            )

            # Ответ Telegram
            try:

                import requests

                answer_url = (
                    f"https://api.telegram.org/"
                    f"bot{Config.TELEGRAM_TOKEN}/"
                    f"answerCallbackQuery"
                )

                requests.post(
                    answer_url,
                    json={
                        "callback_query_id":
                            callback.get(
                                "id",
                                ""
                            ),
                        "text":
                            "✅ Результат сохранён!"
                    },
                    timeout=5
                )

            except Exception as e:

                logger.error(
                    f"Ошибка ответа callback: {e}"
                )

            # ------------------------------------------------
            # RESULT
            # ------------------------------------------------

            if callback_data.startswith(
                "result_"
            ):

                parts = callback_data.split(
                    "_"
                )

                if len(parts) >= 3:

                    result_type = parts[1]

                    match_id = parts[2]

                    match = None

                    cache = storage.load_cache()

                    if not isinstance(
                        cache,
                        dict
                    ):
                        cache = {}

                    match = cache.get(
                        f"match_{match_id}"
                    )

                    # ------------------------------------------------
                    # Если нет cache — ищем файл
                    # ------------------------------------------------

                    if not match:

                        try:

                            file_path = (
                                f"data/"
                                f"match_{match_id}.json"
                            )

                            if os.path.exists(
                                file_path
                            ):

                                with open(
                                    file_path,
                                    "r",
                                    encoding="utf-8"
                                ) as f:

                                    match = json.load(
                                        f
                                    )

                        except Exception:

                            match = None

                    if match:

                        # ------------------------------------------------
                        # SKIP
                        # ------------------------------------------------

                        if result_type == "skip":

                            cache.pop(
                                f"match_{match_id}",
                                None
                            )

                            storage.save_cache(
                                cache
                            )

                            try:

                                os.remove(
                                    f"data/"
                                    f"match_{match_id}.json"
                                )

                            except Exception:
                                pass

                            return "ok", 200

                        # ------------------------------------------------
                        # РЕЗУЛЬТАТ
                        # ------------------------------------------------

                        bets = match.get(
                            "bets",
                            []
                        )

                        if bets:

                            best_bet = bets[0]

                            bet_label = str(
                                best_bet.get(
                                    "label",
                                    ""
                                )
                            ).lower()

                            if result_type == "home":

                                if (
                                    "победа хозяев"
                                    in bet_label
                                    or "п1"
                                    in bet_label
                                ):
                                    result = "win"

                                else:
                                    result = "loss"

                            elif result_type == "away":

                                if (
                                    "победа гостей"
                                    in bet_label
                                    or "п2"
                                    in bet_label
                                ):
                                    result = "win"

                                else:
                                    result = "loss"

                            elif result_type == "draw":

                                if (
                                    "1x"
                                    in bet_label
                                    or "1х"
                                    in bet_label
                                    or "x2"
                                    in bet_label
                                    or "х2"
                                    in bet_label
                                ):
                                    result = "win"

                                else:
                                    result = "loss"

                            else:

                                result = "loss"

                            # ------------------------------------------------
                            # СОХРАНЕНИЕ
                            # ------------------------------------------------

                            try:

                                history = (
                                    storage.load_history()
                                )

                                if not isinstance(
                                    history,
                                    list
                                ):
                                    history = []

                                try:

                                    stake = float(
                                        best_bet.get(
                                            "stake",
                                            0
                                        )
                                    )

                                except (
                                    ValueError,
                                    TypeError
                                ):

                                    stake = 0.0

                                try:

                                    odds = float(
                                        best_bet.get(
                                            "odds",
                                            1
                                        )
                                    )

                                except (
                                    ValueError,
                                    TypeError
                                ):

                                    odds = 1.0

                                if result == "win":

                                    profit = round(
                                        stake
                                        * (odds - 1),
                                        2
                                    )

                                elif result == "loss":

                                    profit = -round(
                                        stake,
                                        2
                                    )

                                else:

                                    profit = 0

                                # ============================================
                                # ВАЖНО:
                                # Теперь сохраняем fixture_id
                                # ============================================

                                bet_record = {

                                    "fixture_id":
                                        match.get(
                                            "fixture_id"
                                        ),

                                    "home":
                                        match.get(
                                            "home",
                                            ""
                                        ),

                                    "away":
                                        match.get(
                                            "away",
                                            ""
                                        ),

                                    "league":
                                        match.get(
                                            "league",
                                            ""
                                        ),

                                    "bet":
                                        best_bet.get(
                                            "label",
                                            ""
                                        ),

                                    "odds":
                                        odds,

                                    "stake":
                                        stake,

                                    "ev":
                                        best_bet.get(
                                            "ev",
                                            0
                                        ),

                                    "result":
                                        result,

                                    "profit":
                                        profit,

                                    "date":
                                        datetime.now().strftime(
                                            "%Y-%m-%d %H:%M"
                                        ),

                                    "home_goals":
                                        None,

                                    "away_goals":
                                        None
                                }

                                history.append(
                                    bet_record
                                )

                                storage.save_history(
                                    history
                                )

                                recalc_stats()

                                cache.pop(
                                    f"match_{match_id}",
                                    None
                                )

                                storage.save_cache(
                                    cache
                                )

                                try:

                                    os.remove(
                                        f"data/"
                                        f"match_{match_id}.json"
                                    )

                                except Exception:
                                    pass

                                msg = (
                                    "✅ "
                                    "Результат сохранён!\n"
                                )

                                msg += (
                                    f"{match.get('home')} "
                                    f"vs "
                                    f"{match.get('away')} "
                                    f"→ {result}"
                                )

                                if result == "win":

                                    msg += (
                                        f"\n💰 "
                                        f"Прибыль: "
                                        f"+${profit}"
                                    )

                                elif result == "loss":

                                    msg += (
                                        f"\n💰 "
                                        f"Проигрыш: "
                                        f"-${stake}"
                                    )

                                send_telegram(
                                    msg
                                )

                            except Exception as e:

                                logger.exception(
                                    "❌ Ошибка сохранения "
                                    "результата"
                                )

                                send_error_to_telegram(
                                    f"Ошибка сохранения "
                                    f"результата:\n{e}"
                                )

            return "ok", 200

        # ====================================================
        # MESSAGE
        # ====================================================

        if "message" in data:

            message = data[
                "message"
            ]

            text = message.get(
                "text",
                ""
            )

            chat_id = (
                message
                .get("chat", {})
                .get("id")
            )

            logger.info(
                f"👤 CHAT ID: {chat_id}"
            )

            logger.info(
                f"📝 ТЕКСТ: {text}"
            )

            # ------------------------------------------------
            # SECURITY
            # ------------------------------------------------

            if str(chat_id) != str(
                Config.ADMIN_CHAT_ID
            ):

                logger.warning(
                    f"⛔ ДОСТУП ЗАПРЕЩЕН "
                    f"для {chat_id}"
                )

                # Не отправляем "Нет доступа"
                # неизвестному пользователю через
                # admin chat_id
                return "ok", 200

            # =================================================
            # /start
            # =================================================

            if text == "/start":

                send_telegram(
                    handlers.handle_start()
                )

            # =================================================
            # /help
            # =================================================

            elif text == "/help":

                send_telegram(
                    handlers.handle_help()
                )

            # =================================================
            # /update
            # =================================================

            elif text == "/update":

                if search_running:

                    send_telegram(
                        "⚠️ Поиск уже запущен!"
                    )

                else:

                    search_running = True

                    start_time = (
                        datetime.now()
                    )

                    try:

                        send_telegram(
                            f"🔄 Поиск матчей "
                            f"в {len(Config.LEAGUES)} "
                            f"лигах..."
                        )

                        matches = (
                            get_matches_with_factors()
                        )

                        if matches:

                            send_telegram(
                                f"📊 Найдено "
                                f"{len(matches)} "
                                f"матчей. "
                                f"Анализирую..."
                            )

                            top_matches = (
                                find_top_matches(
                                    matches
                                )
                            )

                            elapsed = int(
                                (
                                    datetime.now()
                                    - start_time
                                ).total_seconds()
                            )

                            auto = get_auto_bet()

                            bets_today_count = (
                                auto.bets_today
                                if auto is not None
                                and hasattr(
                                    auto,
                                    "bets_today"
                                )
                                else 0
                            )

                            if top_matches:

                                send_telegram(
                                    f"✅ "
                                    f"<b>ПОИСК "
                                    f"ЗАВЕРШЕН!</b>\n"
                                    f"📊 Найдено "
                                    f"матчей: "
                                    f"{len(matches)}\n"
                                    f"🎯 Подходящих: "
                                    f"{len(top_matches)}\n"
                                    f"🤖 Авто-ставок: "
                                    f"{bets_today_count}\n"
                                    f"⏱️ Время: "
                                    f"{elapsed} сек."
                                )

                            else:

                                send_telegram(
                                    f"❌ Ставок "
                                    f"не найдено\n"
                                    f"⏱️ Время: "
                                    f"{elapsed} сек."
                                )

                        else:

                            send_telegram(
                                "❌ Матчей не найдено"
                            )

                    except Exception as e:

                        logger.exception(
                            "❌ Ошибка /update"
                        )

                        send_error_to_telegram(
                            f"Ошибка /update:\n{e}"
                        )

                        send_telegram(
                            f"❌ Ошибка поиска:\n{e}"
                        )

                    finally:

                        search_running = False

            # =================================================
            # /today
            # =================================================

            elif text == "/today":

                send_telegram(
                    handlers.handle_today()
                )

            # =================================================
            # /bank
            # =================================================

            elif text == "/bank":

                send_telegram(
                    handlers.handle_bank()
                )

            # =================================================
            # /stats
            # =================================================

            elif text == "/stats":

                send_telegram(
                    handlers.handle_stats()
                )

            # =================================================
            # /bettypes
            # =================================================

            elif text == "/bettypes":

                send_telegram(
                    handlers.handle_bettypes()
                )

            # =================================================
            # /timestats
            # =================================================

            elif text == "/timestats":

                send_telegram(
                    handlers.handle_timestats()
                )

            # =================================================
            # /mlstats
            # =================================================

            elif text == "/mlstats":

                send_telegram(
                    handlers.handle_mlstats()
                )

            # =================================================
            # /report
            # =================================================

            elif text == "/report":

                send_telegram(
                    handlers.handle_report()
                )

            # =================================================
            # /export
            # =================================================

            elif text == "/export":

                file, message_text = (
                    export_to_excel()
                )

                if file:

                    send_telegram(
                        message_text
                    )

                    try:

                        import requests

                        url = (
                            f"https://api.telegram.org/"
                            f"bot{Config.TELEGRAM_TOKEN}/"
                            f"sendDocument"
                        )

                        files = {
                            "document": (
                                "history.xlsx",
                                file,
                                "application/vnd.openxmlformats-"
                                "officedocument.spreadsheetml.sheet"
                            )
                        }

                        data_send = {
                            "chat_id":
                                Config.ADMIN_CHAT_ID,

                            "caption":
                                "📊 История ставок"
                        }

                        requests.post(
                            url,
                            files=files,
                            data=data_send,
                            timeout=30
                        )

                    except Exception as e:

                        logger.exception(
                            "❌ Ошибка отправки Excel"
                        )

                        send_error_to_telegram(
                            f"Ошибка отправки Excel:\n{e}"
                        )

                else:

                    send_telegram(
                        message_text
                    )

            # =================================================
            # /autobet
            # =================================================

            elif text == "/autobet":

                auto = get_auto_bet()

                if auto is None:

                    send_telegram(
                        "❌ AutoBet не загружен"
                    )

                else:

                    auto.enabled = not getattr(
                        auto,
                        "enabled",
                        True
                    )

                    send_telegram(
                        handlers.handle_autobet(
                            auto.enabled
                        )
                    )

            # =================================================
            # /train
            # =================================================

            elif text == "/train":

                send_telegram(
                    handlers.handle_train()
                )

            # =================================================
            # /arb
            # =================================================

            elif text == "/arb":

                send_telegram(
                    handlers.handle_arb()
                )

            # =================================================
            # /anomalies
            # =================================================

            elif text == "/anomalies":

                send_telegram(
                    handlers.handle_anomalies()
                )

            # =================================================
            # /security
            # =================================================

            elif text == "/security":

                send_telegram(
                    handlers.handle_security()
                )

            # =================================================
            # /stop
            # =================================================

            elif text == "/stop":

                search_running = False

                send_telegram(
                    handlers.handle_stop()
                )

            # =================================================
            # /update_results
            # =================================================

            elif text == "/update_results":

                send_telegram(
                    "🔄 Проверка результатов матчей..."
                )

                updated = (
                    update_pending_bets()
                )

                if updated > 0:

                    send_telegram(
                        f"✅ Обновлено "
                        f"{updated} результатов!"
                    )

                else:

                    send_telegram(
                        "📭 Нет завершённых "
                        "матчей для обновления"
                    )

            # =================================================
            # /team
            # =================================================

            elif text.startswith(
                "/team"
            ):

                team_name = (
                    text
                    .replace(
                        "/team",
                        "",
                        1
                    )
                    .strip()
                )

                if team_name:

                    send_telegram(
                        handlers.handle_team(
                            team_name
                        )
                    )

                else:

                    send_telegram(
                        "⚠️ Используй: "
                        "/team Название команды"
                    )

            # =================================================
            # /unblock
            # =================================================

            elif text.startswith(
                "/unblock"
            ):

                ip = (
                    text
                    .replace(
                        "/unblock",
                        "",
                        1
                    )
                    .strip()
                )

                send_telegram(
                    handlers.handle_unblock(
                        ip
                    )
                )

            # =================================================
            # /result
            # =================================================

            elif text.startswith(
                "/result"
            ):

                command = (
                    text
                    .replace(
                        "/result",
                        "",
                        1
                    )
                    .strip()
                )

                # Формат:
                # /result Fulham vs Chelsea 2-1

                if " vs " not in command:

                    send_telegram(
                        "⚠️ Используй:\n"
                        "/result Fulham vs Chelsea 2-1"
                    )

                else:

                    try:

                        match_part = (
                            command.split(
                                " vs ",
                                1
                            )
                        )

                        if len(
                            match_part
                        ) != 2:

                            raise ValueError

                        home = (
                            match_part[0]
                            .strip()
                        )

                        rest = (
                            match_part[1]
                            .strip()
                        )

                        parts_result = (
                            rest.rsplit(
                                " ",
                                1
                            )
                        )

                        if len(
                            parts_result
                        ) != 2:

                            raise ValueError

                        away = (
                            parts_result[0]
                            .strip()
                        )

                        score = (
                            parts_result[1]
                            .strip()
                        )

                        if not home or not away:

                            raise ValueError

                        send_telegram(
                            handlers.handle_result(
                                f"{home} vs {away}",
                                score
                            )
                        )

                    except Exception:

                        send_telegram(
                            "⚠️ Используй:\n"
                            "/result Fulham vs Chelsea 2-1"
                        )

            # =================================================
            # UNKNOWN
            # =================================================

            else:

                send_telegram(
                    "❌ Неизвестная команда. /help"
                )

        logger.info(
            "✅ Webhook завершен"
        )

        return "ok", 200

    except Exception as e:

        logger.exception(
            "❌ Webhook error"
        )

        send_error_to_telegram(
            f"Webhook error:\n{e}"
        )

        # Telegram должен получить 200,
        # чтобы не повторять webhook
        return "ok", 200


# ============================================================
# API: STATS
# ============================================================

@app.route(
    "/api/stats",
    methods=["GET"]
)
def api_stats():

    try:

        stats = storage.load_stats()

        bank = storage.load_bank()

        history = storage.load_history()

        if not isinstance(
            history,
            list
        ):
            history = []

        total_bets = len(
            history
        )

        wins = int(
            stats.get(
                "wins",
                0
            )
        )

        losses = int(
            stats.get(
                "losses",
                0
            )
        )

        pushes = int(
            stats.get(
                "pushes",
                0
            )
        )

        try:

            total_profit = float(
                stats.get(
                    "total_profit",
                    0
                )
            )

        except (
            ValueError,
            TypeError
        ):

            total_profit = 0.0

        winrate = round(
            wins
            / (wins + losses)
            * 100,
            1
        ) if (
            wins + losses
        ) > 0 else 0

        total_stake = 0.0

        for bet in history:

            try:

                total_stake += float(
                    bet.get(
                        "stake",
                        0
                    )
                )

            except (
                ValueError,
                TypeError
            ):
                pass

        roi = round(
            total_profit
            / total_stake
            * 100,
            1
        ) if total_stake > 0 else 0

        avg_stake = round(
            total_stake
            / total_bets,
            2
        ) if total_bets > 0 else 0

        return jsonify({

            "bank": bank,

            "total_bets":
                total_bets,

            "wins":
                wins,

            "losses":
                losses,

            "pushes":
                pushes,

            "profit":
                round(
                    total_profit,
                    2
                ),

            "winrate":
                winrate,

            "roi":
                roi,

            "avg_stake":
                avg_stake
        })

    except Exception as e:

        logger.exception(
            "❌ API stats error"
        )

        return jsonify({
            "error": str(e)
        }), 500


# ============================================================
# API: HISTORY
# ============================================================

@app.route(
    "/api/history",
    methods=["GET"]
)
def api_history():

    try:

        history = storage.load_history()

        if not isinstance(
            history,
            list
        ):
            history = []

        result = []

        for bet in history:

            # Не изменяем оригинал
            item = dict(
                bet
            )

            try:

                stake = float(
                    item.get(
                        "stake",
                        0
                    )
                )

            except (
                ValueError,
                TypeError
            ):

                stake = 0.0

            try:

                odds = float(
                    item.get(
                        "odds",
                        1
                    )
                )

            except (
                ValueError,
                TypeError
            ):

                odds = 1.0

            if item.get(
                "result"
            ) == "win":

                item["profit"] = round(
                    stake
                    * (odds - 1),
                    2
                )

            elif item.get(
                "result"
            ) == "loss":

                item["profit"] = -round(
                    stake,
                    2
                )

            else:

                item["profit"] = 0

            item["match"] = (
                f"{item.get('home', '')} "
                f"vs "
                f"{item.get('away', '')}"
            )

            result.append(
                item
            )

        return jsonify(
            result
        )

    except Exception as e:

        logger.exception(
            "❌ API history error"
        )

        return jsonify({
            "error": str(e)
        }), 500


# ============================================================
# API: BANK
# ============================================================

@app.route(
    "/api/bank",
    methods=["POST"]
)
def api_update_bank():

    try:

        data = request.get_json(
            silent=True
        )

        if not isinstance(
            data,
            dict
        ):

            return jsonify({
                "error": "Invalid JSON"
            }), 400

        if "bank" not in data:

            return jsonify({
                "error":
                    "No bank value"
            }), 400

        bank = data["bank"]

        try:

            bank = float(
                bank
            )

        except (
            ValueError,
            TypeError
        ):

            return jsonify({
                "error":
                    "Bank must be a number"
            }), 400

        storage.save_bank(
            bank
        )

        return jsonify({

            "success":
                True,

            "bank":
                bank
        })

    except Exception as e:

        logger.exception(
            "❌ API bank error"
        )

        return jsonify({
            "error": str(e)
        }), 500


# ============================================================
# API: UPDATE HISTORY
# ============================================================

@app.route(
    "/api/update_history",
    methods=["POST"]
)
def update_history():

    try:

        data = request.get_json(
            silent=True
        )

        if not isinstance(
            data,
            dict
        ):

            return jsonify({
                "error":
                    "Invalid JSON"
            }), 400

        history = data.get(
            "history",
            []
        )

        if not isinstance(
            history,
            list
        ):

            return jsonify({
                "error":
                    "history must be list"
            }), 400

        if not history:

            return jsonify({
                "error":
                    "Нет данных"
            }), 400

        storage.save_history(
            history
        )

        total = len(
            history
        )

        wins = sum(
            1
            for b in history
            if b.get("result") == "win"
        )

        losses = sum(
            1
            for b in history
            if b.get("result") == "loss"
        )

        pushes = sum(
            1
            for b in history
            if b.get("result") == "push"
        )

        total_profit = 0.0

        for b in history:

            try:

                total_profit += float(
                    b.get(
                        "profit",
                        0
                    )
                )

            except (
                ValueError,
                TypeError
            ):
                pass

        stats = storage.load_stats()

        if not isinstance(
            stats,
            dict
        ):
            stats = {}

        stats["total"] = total

        stats["wins"] = wins

        stats["losses"] = losses

        stats["pushes"] = pushes

        stats["total_profit"] = round(
            total_profit,
            2
        )

        storage.save_stats(
            stats
        )

        return jsonify({

            "success":
                True,

            "total":
                total,

            "wins":
                wins,

            "losses":
                losses,

            "pushes":
                pushes,

            "profit":
                round(
                    total_profit,
                    2
                )
        })

    except Exception as e:

        logger.exception(
            "❌ Ошибка обновления истории"
        )

        send_error_to_telegram(
            f"Ошибка обновления истории:\n{e}"
        )

        return jsonify({
            "error": str(e)
        }), 500


# ============================================================
# API: MATCHES
# ============================================================

@app.route(
    "/api/matches",
    methods=["GET"]
)
def api_matches():

    try:

        cache = storage.load_cache()

        if not isinstance(
            cache,
            dict
        ):
            cache = {}

        matches = cache.get(
            "top_matches",
            []
        )

        if not isinstance(
            matches,
            list
        ):
            matches = []

        return jsonify(
            matches
        )

    except Exception as e:

        logger.exception(
            "❌ API matches error"
        )

        return jsonify({
            "error": str(e)
        }), 500


# ============================================================
# INDEX
# ============================================================

@app.route(
    "/",
    methods=["GET"]
)
def index():

    return (
        "🤖 Quantum Bot v12 PRO | "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )


# ============================================================
# HEALTH
# ============================================================

@app.route(
    "/health",
    methods=["GET"]
)
def health():

    return {

        "status":
            "ok",

        "time":
            datetime.now().isoformat()
    }


# ============================================================
# ЗАПУСК
# ============================================================

if __name__ == "__main__":

    try:

        setup_logging()

    except Exception as e:

        print(
            f"⚠️ Ошибка setup_logging: {e}"
        )

    try:

        start_scheduler()

        logger.info(
            "✅ Scheduler запущен"
        )

    except Exception as e:

        logger.exception(
            "❌ Ошибка запуска scheduler"
        )

        send_error_to_telegram(
            f"Ошибка запуска scheduler:\n{e}"
        )

    port = int(
        os.environ.get(
            "PORT",
            10000
        )
    )

    logger.info(
        "🚀 БОТ ЗАПУЩЕН!"
    )

    logger.info(
        f"📊 Сканируется "
        f"{len(Config.LEAGUES)} лиг"
    )

    logger.info(
        f"🤖 Максимум ставок: "
        f"{Config.MAX_BETS_PER_RUN}"
    )

    logger.info(
        "✅ Мониторинг ошибок включен"
    )

    app.run(
        host="0.0.0.0",
        port=port
    )
